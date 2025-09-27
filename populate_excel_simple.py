#!/usr/bin/env python3
"""
Simple Excel Population Script
===============================

Uses the existing write_maps.py to directly populate Excel cells
with enrollment data from enrollment_automation_v6.
"""

import os
import sys
from openpyxl import load_workbook
from datetime import datetime

# Add project root to path
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

# Import from existing modules
from enrollment_automation_v6 import (
    load_block_aggregations,
    read_and_prepare_data,
    build_tier_data_from_source,
    load_plan_mappings
)

from write_maps import SHEET_WRITE_MAPS

def main():
    """Main execution"""
    print("\n" + "="*70)
    print("SIMPLE EXCEL ENROLLMENT POPULATION")
    print("="*70)
    
    # Step 1: Load enrollment data
    print("\nStep 1: Loading enrollment data...")
    plan_mappings = load_plan_mappings()
    block_aggregations = load_block_aggregations()
    
    source_data_path = os.path.join(PROJECT_ROOT, "data", "input", "source_data.xlsx")
    if not os.path.exists(source_data_path):
        print(f"✗ Source data not found: {source_data_path}")
        return 1
    
    df = read_and_prepare_data(source_data_path, plan_mappings)
    tier_data = build_tier_data_from_source(df, block_aggregations, False)
    print(f"✓ Loaded data for {len(tier_data)} facilities")
    
    # Step 2: Load Excel template
    template_path = "Prime Enrollment Funding by Facility for August.xlsx"
    if not os.path.exists(template_path):
        print(f"✗ Template not found: {template_path}")
        return 1
    
    print(f"\nStep 2: Loading Excel template...")
    wb = load_workbook(template_path)
    print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets")
    
    # Step 3: Write values using write_maps
    print("\nStep 3: Writing enrollment values...")
    
    total_writes = 0
    errors = []
    
    # Process each sheet's write map
    for sheet_name, write_map in SHEET_WRITE_MAPS.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found in workbook")
            continue
        
        ws = wb[sheet_name]
        sheet_writes = 0
        
        # Process each write instruction
        for instruction in write_map:
            client_id = instruction["client_id"]
            plan_type = instruction["plan"]
            cells = instruction["cells"]
            
            # Get data for this facility
            if client_id not in tier_data:
                continue
            
            if plan_type not in tier_data[client_id]:
                continue
            
            # Aggregate all blocks for this plan type
            tier_counts = {}
            for block_label, block_tiers in tier_data[client_id][plan_type].items():
                for tier, count in block_tiers.items():
                    if tier not in tier_counts:
                        tier_counts[tier] = 0
                    tier_counts[tier] += count
            
            # Write to cells
            for tier, cell_ref in cells.items():
                value = tier_counts.get(tier, 0)
                
                try:
                    ws[cell_ref] = value
                    sheet_writes += 1
                    total_writes += 1
                    
                    if value > 0:  # Only print non-zero writes
                        print(f"  {sheet_name}/{client_id} {plan_type} {tier}: {value} → {cell_ref}")
                        
                except Exception as e:
                    errors.append(f"Error writing to {sheet_name}!{cell_ref}: {str(e)}")
        
        if sheet_writes > 0:
            print(f"✓ {sheet_name}: {sheet_writes} cells updated")
    
    # Step 4: Save the workbook
    output_path = f"enrollment_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"\nStep 4: Saving workbook...")
    wb.save(output_path)
    print(f"✓ Saved to: {output_path}")
    
    # Print summary
    print("\n" + "="*70)
    print("SUMMARY")
    print("="*70)
    print(f"✓ Total cells updated: {total_writes}")
    
    if errors:
        print(f"\n✗ Errors encountered: {len(errors)}")
        for error in errors[:5]:
            print(f"  - {error}")
    else:
        print("\n✅ SUCCESS: All values written without errors!")
    
    return 0 if not errors else 1

if __name__ == "__main__":
    sys.exit(main())