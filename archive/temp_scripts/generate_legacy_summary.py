"""
Generate Legacy Tab Summary in Pivot Table Format
Shows EPO and VALUE enrollment counts by tier
"""

import pandas as pd
import openpyxl
from collections import defaultdict
import sys

def generate_legacy_pivot_from_source():
    """Generate pivot table summary from source data"""
    
    # Read source data
    df = pd.read_excel('/mnt/c/Users/becas/Prime_EFR/data/input/source_data.xlsx')
    
    # Legacy CLIENT IDs
    legacy_cids = [
        'H3100', 'H3105', 'H3110', 'H3115', 'H3120', 'H3130', 'H3140', 'H3150',
        'H3160', 'H3170', 'H3180', 'H3190', 'H3200', 'H3210', 'H3230', 'H3240',
        'H3280', 'H3285', 'H3290', 'H3300'
    ]
    
    # Get Legacy subscribers with active status (A or C for COBRA)
    df_legacy = df[(df['CLIENT ID'].isin(legacy_cids)) & 
                   (df['STATUS'].str.upper().isin(['A', 'C'])) &
                   (df['RELATION'].str.upper().isin(['SELF', 'EE', 'EMPLOYEE', 'SUBSCRIBER']))].copy()
    
    # Map PLAN to EPO/VALUE
    def map_plan_type(plan):
        if pd.isna(plan):
            return 'UNKNOWN'
        p = str(plan).strip().upper()
        if 'VALUE' in p or 'VAL' in p:
            return 'Value'
        elif 'EPO' in p:
            return 'EPO'
        else:
            return 'Other'
    
    df_legacy['plan_type'] = df_legacy['PLAN'].apply(map_plan_type)
    
    # Filter for EPO and VALUE only
    df_filtered = df_legacy[df_legacy['plan_type'].isin(['EPO', 'Value'])]
    
    # Debug: Check if we have data
    print(f"Debug: Legacy data rows: {len(df_filtered)}")
    print(f"Debug: Plan types: {df_filtered['plan_type'].value_counts().to_dict()}")
    print(f"Debug: BEN CODEs: {df_filtered['BEN CODE'].value_counts().head().to_dict()}")
    
    # Create pivot table
    pivot = pd.pivot_table(
        df_filtered,
        values='DEP SSN',
        index='BEN CODE',
        columns='plan_type',
        aggfunc='count',
        fill_value=0,
        margins=True,
        margins_name='Grand Total'
    )
    
    return pivot

def generate_legacy_pivot_from_excel(filepath):
    """Generate pivot table summary from updated Excel file"""
    
    # Read the updated workbook
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Legacy']
    
    # Extract values from Legacy tab
    legacy_data = defaultdict(lambda: defaultdict(int))
    
    # Map of facilities in Legacy tab with their cell ranges
    facilities = [
        ('H3170', 'San Dimas', [(4,7), (10,13)]),  # EPO and VALUE ranges
        ('H3130', 'Bio-Med', [(20,23), (26,29)]),
        ('H3100', 'Chino', [(36,39), (42,45)]),
        ('H3300', 'Chino RN', [(53,56), (59,62)]),
        ('H3140', 'Desert Valley', [(69,72), (75,78)]),
        ('H3150', 'Desert Med', [(85,88), (91,94)]),
        ('H3210', 'Huntington', [(101,104), (107,110)]),
        ('H3200', 'La Palma', [(133,136), (139,142)]),
        ('H3160', 'Montclair', [(149,152), (155,158)]),
        ('H3115', 'Premiere', [(165,168), None]),  # EPO only
        ('H3110', 'Prime Mgmt', [(175,178), (181,184)]),
        ('H3230', 'Paradise', [(191,194), (197,200)]),
        ('H3240', 'Paradise Med', [(207,210), (213,216)]),
        ('H3180', 'Sherman', [(223,226), (229,232)]),
        ('H3280', 'Shasta', [(271,274), (277,280)]),
        ('H3285', 'Shasta Med', [(287,290), (293,296)])
    ]
    
    # Tier mapping
    tier_map = {0: 'EMP', 1: 'ESP', 2: 'ECH', 3: 'FAM'}
    
    # Collect all values
    for cid, name, ranges in facilities:
        # EPO values
        if ranges[0]:
            start, end = ranges[0]
            for i, row in enumerate(range(start, end+1)):
                val = ws[f'G{row}'].value
                if val and val != 0:
                    tier = tier_map.get(i, 'UNKNOWN')
                    legacy_data[tier]['EPO'] += val
        
        # VALUE values  
        if ranges[1]:
            start, end = ranges[1]
            for i, row in enumerate(range(start, end+1)):
                val = ws[f'G{row}'].value
                if val and val != 0:
                    tier = tier_map.get(i, 'UNKNOWN')
                    legacy_data[tier]['Value'] += val
    
    return legacy_data

def print_pivot_table(pivot_data, source="Excel"):
    """Print pivot table in requested format"""
    
    if isinstance(pivot_data, pd.DataFrame):
        # From pandas pivot
        print("\nLegacy Tab Pivot Table (from source data):")
        print("Row Labels\tEPO\tValue\tGrand Total")
        
        for tier in ['EMP', 'ESP', 'ECH', 'FAM']:
            if tier in pivot_data.index:
                epo = int(pivot_data.loc[tier, 'EPO']) if 'EPO' in pivot_data.columns else 0
                value = int(pivot_data.loc[tier, 'Value']) if 'Value' in pivot_data.columns else 0
                total = epo + value
                print(f'{tier}\t{epo}\t{value}\t{total}')
        
        # Grand Total
        if 'Grand Total' in pivot_data.index:
            epo = int(pivot_data.loc['Grand Total', 'EPO']) if 'EPO' in pivot_data.columns else 0
            value = int(pivot_data.loc['Grand Total', 'Value']) if 'Value' in pivot_data.columns else 0
            total = int(pivot_data.loc['Grand Total', 'Grand Total'])
            print(f'Grand Total\t{epo}\t{value}\t{total}')
    else:
        # From dictionary
        print(f"\nLegacy Tab Pivot Table (from {source}):")
        print("Row Labels\tEPO\tValue\tGrand Total")
        
        grand_total_epo = 0
        grand_total_value = 0
        
        for tier in ['EMP', 'ESP', 'ECH', 'FAM']:
            epo = pivot_data[tier]['EPO']
            value = pivot_data[tier]['Value']
            total = epo + value
            grand_total_epo += epo
            grand_total_value += value
            print(f'{tier}\t{epo}\t{value}\t{total}')
        
        print(f'Grand Total\t{grand_total_epo}\t{grand_total_value}\t{grand_total_epo + grand_total_value}')

def main():
    print("="*60)
    print("LEGACY TAB SUMMARY GENERATOR")
    print("="*60)
    
    # Generate from source data
    print("\n1. FROM SOURCE DATA (source_data.xlsx):")
    pivot_source = generate_legacy_pivot_from_source()
    print_pivot_table(pivot_source, "source data")
    
    # Generate from Excel file
    print("\n2. FROM UPDATED EXCEL FILE:")
    excel_file = '/mnt/c/Users/becas/Prime_EFR/Prime Enrollment Funding by Facility for August_updated.xlsx'
    pivot_excel = generate_legacy_pivot_from_excel(excel_file)
    print_pivot_table(pivot_excel, "Excel file")
    
    print("\n" + "="*60)
    print("EXPECTED VALUES (for comparison):")
    print("Row Labels\tEPO\tValue\tGrand Total")
    print("EMP\t3053\t290\t3343")
    print("ESP\t481\t34\t515")
    print("ECH\t833\t60\t893")
    print("FAM\t596\t48\t644")
    print("Grand Total\t4963\t432\t5395")

if __name__ == "__main__":
    main()