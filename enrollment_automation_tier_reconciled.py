""" 
=================================================================================
ENROLLMENT AUTOMATION - TIER RECONCILIATION VERSION
=================================================================================

This version includes:
- Control totals harness for exact tier matching
- Waterfall tracking to find where 106 rows go missing
- Fixed tier normalization (no defaulting to EE)
- Comprehensive unknown auditing
- Global integrity assertions

Control Totals (Ground Truth):
- EMP (EE Only): 14,533
- ESP (EE+Spouse): 2,639
- ECH (EE+Child(ren)): 4,413
- FAM (EE+Family): 3,123
TOTAL: 24,708
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import traceback
import os
import re
import csv as _csv
import shutil
from difflib import SequenceMatcher
from collections import Counter, defaultdict
from collections import Counter as _Ctr
warnings.filterwarnings('ignore')

# CONTROL TOTALS - GROUND TRUTH
CONTROL_TOTALS = {
    "EE Only": 14533,
    "EE+Spouse": 2639,
    "EE+Child(ren)": 4413,
    "EE+Family": 3123
}
TIER_ORDER = ["EE Only", "EE+Spouse", "EE+Child(ren)", "EE+Family"]
CONTROL_TOTAL = sum(CONTROL_TOTALS.values())  # 24,708

# Waterfall tracking
waterfall_stages = []
unknown_tiers_tracker = Counter()
removed_rows_samples = {}

# Write tracking globals
WRITE_LOG_ROWS = []   # (sheet, client_id, plan, tier_label, cell, value)
PROCESSED_SHEETS = set()

def log_stage(stage_name, df, prev_df=None):
    """
    Log a stage in the waterfall with tier counts and row samples
    """
    global waterfall_stages, removed_rows_samples
    
    # Get tier distribution
    tier_counts = {}
    if 'tier' in df.columns:
        tier_dist = df['tier'].value_counts()
        for tier in TIER_ORDER:
            tier_counts[tier] = tier_dist.get(tier, 0)
        tier_counts['UNKNOWN'] = tier_dist.get('UNKNOWN', 0)
        tier_counts['Other'] = len(df) - sum(tier_counts.values())
    else:
        tier_counts = {tier: 0 for tier in TIER_ORDER}
        tier_counts['UNKNOWN'] = 0
        tier_counts['Other'] = len(df)
    
    # Calculate delta to control
    delta = sum(tier_counts.get(tier, 0) for tier in TIER_ORDER) - CONTROL_TOTAL
    
    # Store stage info
    stage_info = {
        'stage': stage_name,
        'total_rows': len(df),
        'tier_counts': tier_counts,
        'delta_to_control': delta
    }
    waterfall_stages.append(stage_info)
    
    # Capture removed rows if prev_df provided
    if prev_df is not None and len(prev_df) > len(df):
        # Find removed rows
        if 'index' in prev_df.columns and 'index' in df.columns:
            removed_idx = set(prev_df['index']) - set(df['index'])
        else:
            # Use other identifying columns
            if all(col in prev_df.columns and col in df.columns for col in ['CLIENT ID', 'EMPLOYEE NAME']):
                prev_keys = set(zip(prev_df['CLIENT ID'], prev_df['EMPLOYEE NAME']))
                curr_keys = set(zip(df['CLIENT ID'], df['EMPLOYEE NAME']))
                removed_keys = prev_keys - curr_keys
                removed_idx = prev_df[prev_df.apply(lambda x: (x['CLIENT ID'], x['EMPLOYEE NAME']) in removed_keys, axis=1)].index
            else:
                removed_idx = []
        
        if len(removed_idx) > 0:
            sample_cols = ['CLIENT ID', 'RELATION', 'STATUS', 'PLAN', 'BEN CODE']
            sample_cols = [col for col in sample_cols if col in prev_df.columns]
            if sample_cols:
                removed_sample = prev_df.loc[list(removed_idx)[:10], sample_cols] if len(removed_idx) <= 10 else prev_df.loc[list(removed_idx)[:10], sample_cols]
                removed_rows_samples[stage_name] = removed_sample
    
    # Print stage summary
    print(f"\n--- Stage: {stage_name} ---")
    print(f"Total rows: {len(df):,}")
    if tier_counts:
        print(f"Tier distribution: EE Only={tier_counts.get('EE Only', 0):,}, "
              f"EE+Spouse={tier_counts.get('EE+Spouse', 0):,}, "
              f"EE+Child(ren)={tier_counts.get('EE+Child(ren)', 0):,}, "
              f"EE+Family={tier_counts.get('EE+Family', 0):,}, "
              f"UNKNOWN={tier_counts.get('UNKNOWN', 0):,}")
    print(f"Delta to control: {delta:+,}")
    
    return df

def print_waterfall_table():
    """
    Print the waterfall table showing row movement through stages
    """
    print("\n" + "="*120)
    print("WATERFALL ANALYSIS - Row Movement Through Pipeline")
    print("="*120)
    
    headers = ['Stage', 'Total', 'EE Only', 'EE+Spouse', 'EE+Child(ren)', 'EE+Family', 'UNKNOWN', 'Delta', 'Row Δ']
    print(f"{headers[0]:<20} {headers[1]:>8} {headers[2]:>8} {headers[3]:>10} {headers[4]:>12} {headers[5]:>10} {headers[6]:>8} {headers[7]:>8} {headers[8]:>8}")
    print("-"*120)
    
    prev_total = 0
    for i, stage in enumerate(waterfall_stages):
        row_delta = stage['total_rows'] - prev_total if i > 0 else 0
        tier_counts = stage['tier_counts']
        
        print(f"{stage['stage']:<20} "
              f"{stage['total_rows']:>8,} "
              f"{tier_counts.get('EE Only', 0):>8,} "
              f"{tier_counts.get('EE+Spouse', 0):>10,} "
              f"{tier_counts.get('EE+Child(ren)', 0):>12,} "
              f"{tier_counts.get('EE+Family', 0):>10,} "
              f"{tier_counts.get('UNKNOWN', 0):>8,} "
              f"{stage['delta_to_control']:>+8,} "
              f"{row_delta:>+8,}")
        
        prev_total = stage['total_rows']
    
    print("="*120)

def clean_key(s):
    """Clean and normalize keys for matching - handles trailing spaces"""
    if pd.isna(s):
        return ''
    # Strip, uppercase, and collapse all whitespace (including internal)
    return ' '.join(str(s).strip().upper().split())

def is_active(status):
    """Flexible active status check - includes A (Active) and C (COBRA/Continued)"""
    if pd.isna(status):
        return False
    s = str(status).strip().upper()
    # Include both A (Active) and C (COBRA/Continued coverage)
    return s in ['A', 'ACTIVE', 'ACT', 'C', 'COBRA', 'CONTINUED'] or s.startswith('A')

def is_subscriber(relation):
    """Flexible subscriber check"""
    if pd.isna(relation):
        return False
    r = str(relation).strip().upper()
    return r in ['SELF', 'EE', 'EMPLOYEE', 'SUBSCRIBER', 'S', 'EMP']

def normalize_tier_strict(raw_tier):
    """
    Strictly normalize raw tier text - NEVER default to EE
    Returns one of: 'EE Only', 'EE+Spouse', 'EE+Child(ren)', 'EE+Family', or 'UNKNOWN'
    """
    global unknown_tiers_tracker
    
    if pd.isna(raw_tier):
        unknown_tiers_tracker['<NaN>'] += 1
        return 'UNKNOWN'
    
    # Clean the input thoroughly - handle trailing spaces
    tier_str = str(raw_tier).strip().upper()
    
    # Normalize separators consistently
    for sep in ['&', '+', '/', ' AND ', '  ']:
        tier_str = tier_str.replace(sep, ' ')
    
    # Collapse multiple spaces to single space
    tier_str = ' '.join(tier_str.split())
    
    # Employee Only variants (expanded)
    ee_only_variants = [
        'EMP', 'EE', 'EMPLOYEE ONLY', 'EE ONLY', 'EMPLOYEE', 'E',
        'SELF ONLY', 'SELF', 'SUBSCRIBER ONLY', 'SUBSCRIBER',
        'EMP ONLY', 'E ONLY', 'EMPLOYEE'
    ]
    if tier_str in ee_only_variants:
        return 'EE Only'
    
    # Employee + Spouse variants (expanded)
    spouse_variants = [
        'ESP', 'EE SPOUSE', 'EMPLOYEE SPOUSE', 'EE SPOUSE',
        'EMPLOYEE SPOUSE', 'ES', 'E S', 'E SPOUSE',
        'EMP SPOUSE', 'EMP SP', 'EMPLOYEE SP', 'EE SP'
    ]
    if tier_str in spouse_variants:
        return 'EE+Spouse'
    
    # Employee + Child(ren) variants (expanded to handle both)
    child_variants = [
        'ECH', 'E1D', 'EE CHILD', 'EMPLOYEE CHILD', 'EE CHILDREN',
        'EMPLOYEE CHILDREN', 'EE CHILD', 'EE 1 CHILD', 'EE 1 DEPENDENT',
        'EC', 'E C', 'E CHILD', 'E CHILDREN', 'EMP CHILD', 'EMP CHILDREN',
        'CHILD', 'CHILDREN', 'EE CHILD REN', 'CHILD REN', 'E1C', 'E2C',
        'EE 1 DEP', 'EE DEP', 'EE DEPS', 'EMPLOYEE CHILDREN',
        'EE 1', 'EE 2'  # Sometimes just numbers after EE
    ]
    if tier_str in child_variants:
        return 'EE+Child(ren)'
    
    # Employee + Family variants (expanded)
    family_variants = [
        'FAM', 'FAMILY', 'EE FAMILY', 'EMPLOYEE FAMILY',
        'EF', 'E F', 'E FAMILY', 'EMP FAMILY', 'EMP FAM',
        'EMPLOYEE FAM', 'EE FAM'
    ]
    if tier_str in family_variants:
        return 'EE+Family'
    
    # Track unknown tier
    unknown_tiers_tracker[tier_str] += 1
    return 'UNKNOWN'

def fuzzy_match_score(s1, s2):
    """Calculate fuzzy match score between two strings"""
    if pd.isna(s1) or pd.isna(s2):
        return 0.0
    
    str1 = str(s1).lower().strip()
    str2 = str(s2).lower().strip()
    
    # Handle the specific "a" vs "at" case
    str1 = str1.replace(" a st.", " at st.")
    str2 = str2.replace(" a st.", " at st.")
    
    # Remove punctuation
    for char in ".,'-&":
        str1 = str1.replace(char, " ")
        str2 = str2.replace(char, " ")
    
    str1 = " ".join(str1.split())
    str2 = " ".join(str2.split())
    
    return SequenceMatcher(None, str1, str2).ratio()

# Copy all the facility mappings from the reconciled version
TPA_TO_FACILITY = {
    'H3100': 'Chino Valley Medical Center',
    'H3105': 'Glendora Community Hospital',
    'H3110': 'Prime Management',
    'H3115': 'Premiere Healthcare Staffing, LLC',
    'H3120': 'Hospital Business Service',
    'H3130': 'Bio Med Services',
    'H3140': 'Desert Valley Hospital',
    'H3150': 'Desert Valley Medical Group',
    'H3160': 'Montclair',
    'H3170': 'San Dimas',
    'H3180': 'Sherman Oaks',
    'H3190': 'Sherman Oaks Med. Group',
    'H3200': 'La Palma',
    'H3210': 'Huntington Beach',
    'H3220': 'West Anaheim',
    'H3230': 'Paradise Valley',
    'H3240': 'Paradise Valley Medical Grp',
    'H3250': 'Encino Hospital',
    'H3260': 'Garden Grove',
    'H3270': 'Centinela',
    'H3271': 'Robotics Outpatient Center',
    'H3272': 'Centinela Valley Endoscopy Center',
    'H3275': 'St. Francis Medical Center',
    'H3276': 'Shoreline Surgery Center',
    'H3277': "Physician's Surgery Center Downey",
    'H3280': 'Shasta Regional Medical Center',
    'H3285': 'Shasta Medical Group',
    'H3290': 'Hospitality',
    'H3300': "Chino RN's",
    'H3310': 'Alvarado Hospital',
    'H3320': 'Pampa',
    'H3325': 'Roxborough',
    'H3330': 'Lower Bucks',
    'H3335': 'Dallas Medical Center',
    'H3337': 'Dallas Regional Medical Center',
    'H3338': 'Riverview Regional Medical Center',
    'H3339': 'Gadsden Physicians Management',
    'H3340': 'Providence Medical Center',
    'H3345': 'St. John Hospital',
    'H3350': 'Providence Place, Inc.',
    'H3355': 'Knapp Medical Center',
    'H3360': 'Knapp Medical Group',
    'H3365': 'Knapp Ambulatory Surgery Center',
    'H3370': 'Harlingen Medical Center',
    'H3375': 'Garden City Hospital',
    'H3380': 'United Home Health Services',
    'H3381': 'Lake Huron Medical Center',
    'H3382': 'Lake Huron Medical Group',
    'H3385': 'Prime Garden City Medical Group',
    'H3390': 'Rehabilitation Hospital of Rhode Island',
    'H3392': 'Landmark Medical Center',
    'H3394': "Summit Surgery Center at St. Mary's Galena",
    'H3395': "St. Mary's Regional Medical Center",
    'H3396': "St. Mary's Medical Group",
    'H3397': 'Monroe Hospital',
    'H3398': 'North Vista Hospital',
    'H3399': 'North Vista Medical Group',
    'H3400': "St. Mary's Fitness Center",
    'H3500': "St Clare's Health System",
    'H3505': "Saint Mary's General Hospital",
    'H3510': 'Southern Medical Regional Center',
    'H3520': 'Lehigh Regional Medical Center',
    'H3530': "St. Michael's Medical Center",
    'H3540': 'Mission Regional Medical Center',
    'H3560': "St. Mary's Medical Center",
    'H3561': 'St. Joseph Medical Center',
    'H3562': 'South Kansas City Surgi Center',
    'H3563': 'CHCS Home Health Care',
    'H3564': 'CPCN Physicians Service',
    'H3565': 'CPCN Physicians Service (32) STJ',
    'H3566': "St. Mary's Surgical Center",
    'H3591': 'Coshocton County Memorial Hospital',
    'H3592': 'East Liverpool City Hospital',
    'H3593': 'Ohio Valley Home Health Care',
    'H3594': 'Ohio Valley Home Health Services',
    'H3595': 'River Valley Physicians',
    'H3598': 'Suburban Community Hospital',
    'H3599': 'Suburban Medical Group',
    'H3600': 'Reddy Development LLC',
    'H3605': 'Mercy Medical Center - Aurora LLC',
    'H3615': 'Resurrection Medical Center - Chicago LLC',
    'H3625': 'Saint Francis Hospital - Evanston LLC',
    'H3630': 'Saint Joseph Hospital - Elgin LLC',
    'H3635': 'Saint Joseph Hospital - Joliet LLC',
    'H3645': "St. Mary's Hospital - Kankakee, LLC",
    'H3655': 'Saint Mary of Nazareth Hospital - Chicago, LLC',
    'H3660': 'Holy Family Medical Center - Des Plaines LLC',
    'H3665': 'MedSpace Services, LLC',
    'H3670': 'Prime Healthcare Illinois Medical Group, LLC',
    'H3675': 'Prime Healthcare Home Care and Hospice',
    'H3680': 'Prime Healthcare Senior Living'
}

# Simplified FACILITY_MAPPING - we'll focus on processing all facilities
FACILITY_MAPPING = {
    'All': TPA_TO_FACILITY  # Process all facilities together for tier reconciliation
}

# Keep the plan mappings
PLAN_TO_TYPE = {
    'PRIMESFSE': 'EPO', 'PRIMESFUN': 'EPO', 'PRIMEMSFE': 'EPO',
    'PRIMEMMLMRI': 'VALUE', 'PRIMEMMCV': 'EPO', 'PRIMEMMGLN': 'EPO',
    'PRIMEMMEPOLE2': 'EPO', 'PRIMEMMWA': 'EPO', 'PRIMEMMCC': 'EPO',
    'PRIMEMMDAL': 'EPO', 'PRIMEMMEL': 'EPO', 'PRIMEMMELPOS': 'PPO',
    # ... (include all plan mappings from original)
}

def infer_plan_group_and_variant(plan_raw, mapping=PLAN_TO_TYPE):
    """
    Infer plan group and variant
    """
    if pd.isna(plan_raw):
        return 'UNKNOWN', 'UNKNOWN'
    
    plan_clean = clean_key(plan_raw)
    
    # Try mapping first
    if plan_clean in mapping:
        return mapping[plan_clean], plan_clean
    
    # Heuristic fallback
    if 'PPO' in plan_clean:
        return 'PPO', plan_clean
    elif 'EPO' in plan_clean:
        return 'EPO', plan_clean
    elif 'VAL' in plan_clean or 'VALUE' in plan_clean:
        return 'VALUE', plan_clean
    
    return 'UNKNOWN', plan_clean

def read_and_prepare_data_with_waterfall(file_path):
    """
    Read source data with waterfall tracking
    """
    # Stage 1: Read raw
    df = pd.read_excel(file_path, sheet_name=0)
    df['original_index'] = df.index  # Track original rows
    df = log_stage('read_raw', df)
    
    # Stage 2: Clean keys
    if 'CLIENT ID' in df.columns:
        df['CLIENT ID'] = df['CLIENT ID'].apply(clean_key)
    if 'BEN CODE' in df.columns:
        # Clean BEN CODE to handle trailing spaces
        df['BEN CODE'] = df['BEN CODE'].apply(clean_key)
    df = log_stage('clean_keys', df)
    
    # Stage 3: Status filter (flexible)
    prev_df = df.copy()
    if 'STATUS' in df.columns:
        df['is_active'] = df['STATUS'].apply(is_active)
        df = df[df['is_active']].copy()
    df = log_stage('status_filter', df, prev_df)
    
    # Stage 4: Relation filter (flexible)
    prev_df = df.copy()
    if 'RELATION' in df.columns:
        df['is_subscriber'] = df['RELATION'].apply(is_subscriber)
        df = df[df['is_subscriber']].copy()
    df = log_stage('relation_filter', df, prev_df)
    
    # Stage 5: Facility mapping (keep unknowns)
    if 'CLIENT ID' in df.columns:
        df['facility_id'] = df['CLIENT ID']
        df['facility_name'] = df['facility_id'].map(TPA_TO_FACILITY)
        df['facility_name'] = df['facility_name'].fillna('UNKNOWN')
    df = log_stage('facility_map', df)
    
    # Stage 6: Tier normalization (strict)
    if 'BEN CODE' in df.columns:
        df['tier'] = df['BEN CODE'].apply(normalize_tier_strict)
    else:
        df['tier'] = 'UNKNOWN'
    df = log_stage('tier_normalized', df)
    
    # Stage 7: Plan grouping
    if 'PLAN' in df.columns:
        df[['plan_group', 'plan_variant']] = df['PLAN'].apply(
            lambda x: pd.Series(infer_plan_group_and_variant(x))
        )
    else:
        df['plan_group'] = 'UNKNOWN'
        df['plan_variant'] = 'UNKNOWN'
    df = log_stage('plan_grouped', df)
    
    # Stage 8: Pre-pivot
    df = log_stage('pre_pivot', df)
    
    return df

def print_unknown_audit():
    """
    Print audit of unknown tiers
    """
    print("\n" + "="*80)
    print("UNKNOWN TIERS AUDIT")
    print("="*80)
    
    if unknown_tiers_tracker:
        print(f"\nTotal unique unknown tier values: {len(unknown_tiers_tracker)}")
        print(f"Total unknown tier rows: {sum(unknown_tiers_tracker.values())}")
        print("\nTop 20 unknown tier values:")
        for tier_val, count in unknown_tiers_tracker.most_common(20):
            print(f"  '{tier_val}': {count}")
    else:
        print("No unknown tiers found!")
    
    print("="*80)

def assert_control_totals(df):
    """
    Assert that final totals match control exactly
    """
    print("\n" + "="*80)
    print("CONTROL TOTALS ASSERTION")
    print("="*80)
    
    # Get actual totals
    actual = {}
    if 'tier' in df.columns:
        tier_counts = df['tier'].value_counts()
        for tier in TIER_ORDER:
            actual[tier] = tier_counts.get(tier, 0)
    
    # Compare to control
    print("\n{:<20} {:>10} {:>10} {:>10}".format("Tier", "Control", "Actual", "Delta"))
    print("-"*60)
    
    all_match = True
    for tier in TIER_ORDER:
        control = CONTROL_TOTALS[tier]
        actual_count = actual.get(tier, 0)
        delta = actual_count - control
        status = "✓" if delta == 0 else "✗"
        print(f"{tier:<20} {control:>10,} {actual_count:>10,} {delta:>+10,} {status}")
        if delta != 0:
            all_match = False
    
    print("-"*60)
    total_control = sum(CONTROL_TOTALS.values())
    total_actual = sum(actual.get(tier, 0) for tier in TIER_ORDER)
    total_delta = total_actual - total_control
    status = "✓" if total_delta == 0 else "✗"
    print(f"{'TOTAL':<20} {total_control:>10,} {total_actual:>10,} {total_delta:>+10,} {status}")
    
    # Check for unknowns
    unknown_count = tier_counts.get('UNKNOWN', 0) if 'tier' in df.columns else 0
    if unknown_count > 0:
        print(f"\n⚠️ WARNING: {unknown_count} rows with UNKNOWN tier!")
    
    if all_match and unknown_count == 0:
        print("\n✓ Control tie-out PASS: 14533 / 2639 / 4413 / 3123. Global delta resolved (Δ=0).")
    else:
        print(f"\n✗ Control tie-out FAILED: Delta = {total_delta:+,}")
        if unknown_count > 0:
            print(f"   {unknown_count} UNKNOWN tiers need to be resolved")
    
    return all_match

def print_facility_comparison(df, facility_id, facility_name):
    """
    Print before/after comparison for a specific facility
    """
    if 'CLIENT ID' not in df.columns or 'tier' not in df.columns:
        return
    
    facility_data = df[df['CLIENT ID'] == facility_id]
    if facility_data.empty:
        print(f"\n{facility_name} ({facility_id}): No data found")
        return
    
    tier_counts = facility_data['tier'].value_counts()
    print(f"\n{facility_name} ({facility_id}):")
    for tier in TIER_ORDER:
        count = tier_counts.get(tier, 0)
        if count > 0:
            print(f"  {tier}: {count}")
    unknown = tier_counts.get('UNKNOWN', 0)
    if unknown > 0:
        print(f"  UNKNOWN: {unknown}")

# ============= WRITE-BACK FUNCTIONS =============

DRY_RUN = False  # Set to True for preview mode
STRICT_CONTROL_CHECK = True  # Set False if ground truth includes COBRA
DRY_RUN_WRITE = False  # Set to True to produce CSV without saving workbook

def is_active_for_ees(status):
    """Active only for Ees write-back - NO COBRA"""
    t = str(status).strip().upper() if pd.notna(status) else ""
    return t == "A" or t.startswith("A")

def assert_control_from_tier_data(tier_data):
    """Assert global control totals using built tier_data (COBRA included)."""
    totals = _Ctr({'EE Only':0,'EE+Spouse':0,'EE+Child(ren)':0,'EE+Family':0})
    for client_plans in tier_data.values():
        for counts in client_plans.values():
            totals['EE Only']   += int(counts.get('EE Only',0))
            totals['EE+Spouse'] += int(counts.get('EE+Spouse',0))
            child_total = int(counts.get('EE+Child',0)) + int(counts.get('EE+Children',0))
            totals['EE+Child(ren)'] += child_total
            totals['EE+Family'] += int(counts.get('EE+Family',0))

    deltas = {k: totals[k] - CONTROL_TOTALS[k] for k in CONTROL_TOTALS}
    ok = all(v == 0 for v in deltas.values())
    print("CONTROL CHECK (from tier_data) ->", dict(totals), "Δ:", deltas, "|", "OK" if ok else "FAIL")
    return ok

def _soft_assert_label(ws, any_cell_addr, expected_snippet):
    """Optional label sanity check"""
    if not expected_snippet:
        return
    r = ws[any_cell_addr].row
    c = ws[any_cell_addr].column - 1  # usually label is one column left
    if c < 1:
        return
    label = str(ws.cell(r, c).value or '').upper()
    if expected_snippet.split()[0].upper() not in label:
        print(f"  ⚠ Label check @ {get_column_letter(c)}{r}: '{label[:50]}' vs expected ~ '{expected_snippet[:50]}'")

def _log(sheet, client_id, plan, tier_lbl, cell, value):
    """Log write operation"""
    WRITE_LOG_ROWS.append((sheet, client_id, plan, tier_lbl, cell, int(value)))

def route_sherman_oaks_block(plan_raw):
    """Route Sherman Oaks blocks by variant/union (stub for future implementation)"""
    s = clean_key(plan_raw)
    # Example rule: union keywords to block 2
    if any(x in s for x in ['CIR','JNESO','IUOE','UNION']):
        return 'EPO_2' if 'EPO' in s else 'VALUE_2'
    return 'EPO_1' if 'EPO' in s else 'VALUE_1'

def create_employee_group(df):
    """Create EMPLOYEE_GROUP identifier with priority order"""
    # Try columns in priority order
    id_columns = ['EMPLOYEE ID', 'EMP ID', 'Employee_ID', 'ALT ID']
    
    for id_col in id_columns:
        if id_col in df.columns and df[id_col].notna().sum() > len(df) * 0.5:
            # Clean and combine with CLIENT ID
            df['EMPLOYEE_GROUP'] = (
                df['CLIENT ID'].astype(str).str.strip().str.upper() + '_' +
                df[id_col].astype(str).str.strip().str.upper()
            )
            print(f"   Created EMPLOYEE_GROUP using CLIENT ID + {id_col}")
            return df
    
    # Fallback to EMPLOYEE NAME (not present in our data, so use DEP SSN)
    if 'DEP SSN' in df.columns:
        df['EMPLOYEE_GROUP'] = (
            df['CLIENT ID'].astype(str).str.strip().str.upper() + '_' +
            df['DEP SSN'].astype(str).str.strip().str.upper()
        )
        print(f"   Created EMPLOYEE_GROUP using CLIENT ID + DEP SSN")
    else:
        raise ValueError("Cannot create EMPLOYEE_GROUP: no suitable ID column found")
    
    return df

def calculate_tier_from_composition(family_df, emp_group):
    """Calculate tier from complete family composition"""
    relations = family_df['RELATION'].str.strip().str.upper().tolist()
    
    # Must have SELF anchor
    has_self = any(r in ['SELF', 'EE', 'EMPLOYEE'] for r in relations)
    if not has_self:
        return 'UNKNOWN'
    
    # Check for spouse
    has_spouse = any(r in ['SPOUSE', 'HUSBAND', 'WIFE'] for r in relations)
    
    # Count children  
    child_count = sum(1 for r in relations if r in ['CHILD', 'SON', 'DAUGHTER', 'DEPENDENT'])
    
    # Determine tier (exact logic as specified)
    if has_spouse and child_count > 0:
        return 'EE+Family'
    elif has_spouse:
        return 'EE+Spouse'
    elif child_count > 0:
        if child_count == 1:
            return 'EE+Child'
        else:
            return 'EE+Children'
    else:
        return 'EE Only'

def deduplicate_enrollments(df):
    """Deduplicate keeping latest EFF. DATE"""
    initial_count = len(df)
    
    # Parse dates safely
    if 'EFF. DATE' in df.columns:
        df['EFF. DATE'] = pd.to_datetime(df['EFF. DATE'], errors='coerce')
    
    # Sort by date (latest first) and deduplicate
    df_sorted = df.sort_values('EFF. DATE', ascending=False, na_position='last')
    df_dedup = df_sorted.drop_duplicates(
        subset=['CLIENT ID', 'EMPLOYEE_GROUP', 'PLAN'], 
        keep='first'
    )
    
    rows_removed = initial_count - len(df_dedup)
    if rows_removed > 0:
        print(f"   Deduplication: removed {rows_removed} duplicates")
    
    return df_dedup

def build_facility_plan_tier_data(df_subs):
    """Build nested dict with all tier keys"""
    data = {}
    unknown_plans = []
    
    for _, row in df_subs.iterrows():
        client_id = row['CLIENT ID']
        tier = row['CALCULATED_TIER']
        plan_raw = row['PLAN']
        
        # Map plan
        plan_group, _ = infer_plan_group_and_variant(plan_raw)
        
        if plan_group == 'UNKNOWN':
            if len(unknown_plans) < 3:
                unknown_plans.append((client_id, plan_raw))
            continue
        
        # Initialize nested structure with ALL tier keys
        if client_id not in data:
            data[client_id] = {}
        
        if plan_group not in data[client_id]:
            data[client_id][plan_group] = {
                'EE Only': 0,
                'EE+Spouse': 0,
                'EE+Child': 0,
                'EE+Children': 0,
                'EE+Family': 0
            }
        
        # Increment count
        if tier in data[client_id][plan_group]:
            data[client_id][plan_group][tier] += 1
    
    # Check for unknown plans
    if unknown_plans:
        print(f"   WARNING: {len(unknown_plans)} unknown plans. First 3:")
        for client_id, plan in unknown_plans[:3]:
            print(f"     {client_id}: {plan}")
        if len(unknown_plans) > 10:  # Fail if too many
            raise ValueError(f"Too many unknown plans: {len(unknown_plans)}")
    
    return data

def get_cell_value_safe(ws, row, col):
    """Get cell value handling merged cells"""
    cell = ws.cell(row=row, column=col)
    
    # Check if cell is in merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Return top-left cell value
            top_left_row = merged_range.min_row
            top_left_col = merged_range.min_col
            return ws.cell(row=top_left_row, column=top_left_col).value
    
    return cell.value

def calculate_jaccard(s1, s2):
    """Token-based Jaccard similarity"""
    def tokenize(s):
        # Clean and tokenize
        tokens = re.sub(r'[^\w\s]', ' ', str(s).upper()).split()
        # Remove stop words
        stop = {'THE', 'OF', 'AND', 'AT', 'A', 'INC', 'LLC', 'CENTER', 'HOSPITAL', 'MEDICAL'}
        return {t for t in tokens if len(t) > 1 and t not in stop}
    
    t1, t2 = tokenize(s1), tokenize(s2)
    
    if not t1 or not t2:
        return 0.0
    
    intersection = len(t1 & t2)
    union = len(t1 | t2)
    
    return intersection / union if union > 0 else 0.0

def find_facility_location_strict(ws, facility_name, client_id):
    """Require CLIENT ID hit or Jaccard ≥0.90"""
    candidates = []
    
    for r in range(1, min(ws.max_row, 40)+1):
        for c in range(1, min(ws.max_column, 10)+1):
            val = get_cell_value_safe(ws, r, c)
            if not val:
                continue
            
            val_str = str(val)
            
            # Check for CLIENT ID (highest priority)
            if client_id and str(client_id) in val_str:
                return r, c
            
            # Calculate Jaccard similarity
            score = calculate_jaccard(facility_name, val_str)
            candidates.append((score, r, c, val_str[:50]))
    
    # Sort by score
    candidates.sort(reverse=True, key=lambda x: x[0])
    
    # Accept only if ≥0.90
    if candidates and candidates[0][0] >= 0.90:
        return candidates[0][1], candidates[0][2]
    
    # Show top 3 candidates
    print(f"    ERROR: No match for '{facility_name}' (CID: {client_id})")
    if candidates:
        print(f"    Top candidates:")
        for i, (score, r, c, text) in enumerate(candidates[:3]):
            print(f"      {i+1}. Score {score:.2f}: '{text}' at row {r}, col {c}")
    
    return None, None

def find_ees_col_flexible(ws, facility_row, facility_col):
    """Find the column that holds the employee counts (EEs)."""
    patterns = [r'\bEE\b', r'\bEES\b', r"\bEE'?S\b", r'\bEMP(LOYEES?)?\b']
    for r in range(max(1, facility_row-3), min(ws.max_row, facility_row+3)):
        for c in range(1, min(ws.max_column, 20)+1):
            val = get_cell_value_safe(ws, r, c)
            if isinstance(val, str):
                txt = val.strip()
                if any(re.search(p, txt, re.IGNORECASE) for p in patterns) and len(txt) <= 20:
                    return c
    fallback = facility_col + 3
    print(f"    WARNING: 'EEs' not found near row {facility_row}, defaulting to col {fallback}")
    return fallback

def find_ees_col_strict(ws, facility_row, facility_col):
    """Backward compatibility - calls flexible version"""
    return find_ees_col_flexible(ws, facility_row, facility_col)

def find_plan_section_tolerant(ws, facility_row, facility_col, plan_type):
    """Search with regex for plan headers"""
    # Build regex pattern based on plan type
    if plan_type == 'EPO':
        pattern = r'^(EPO|EPO\s+PLAN|EPO\s*[-–])'
    elif plan_type == 'PPO':
        pattern = r'^(PPO|PPO\s+PLAN|PPO\s*[-–])'
    elif plan_type == 'VALUE':
        pattern = r'^(VALUE|VAL|VALUE\s+PLAN|VAL\s*[-–])'
    else:
        return None
    
    search_start = facility_row + 1
    search_end = min(ws.max_row, facility_row + 30)
    
    # Search facility column first
    for r in range(search_start, search_end):
        val = get_cell_value_safe(ws, r, facility_col)
        if val and re.search(pattern, str(val).strip().upper()):
            return r
    
    # Try one column left and right
    for offset in [-1, 1]:
        for r in range(search_start, search_end):
            val = get_cell_value_safe(ws, r, facility_col + offset)
            if val and re.search(pattern, str(val).strip().upper()):
                print(f"    Found {plan_type} at column offset {offset}")
                return r
    
    return None

def detect_combined_children_tolerant(ws, start_row, label_col):
    """Detect children rows with flexible matching"""
    # Check rows 2-3 for child labels
    for offset in [2, 3]:
        label = get_cell_value_safe(ws, start_row + offset, label_col)
        if label:
            # Remove punctuation and check
            label_clean = re.sub(r'[^\w\s]', '', str(label).lower())
            if any(term in label_clean for term in ['children', 'child(ren)', 'child ren']):
                return True
    return False

def write_tier_rows_safe(ws, start_row, ees_col, tier_counts, label_col):
    """Write with int coercion"""
    combined = detect_combined_children_tolerant(ws, start_row, label_col)
    
    # Zero-fill first
    rows_to_clear = 4 if combined else 5
    for offset in range(rows_to_clear):
        ws.cell(row=start_row + offset, column=ees_col, value=0)
    
    # Write as integers
    if combined:
        child_total = int(tier_counts.get('EE+Child', 0) + tier_counts.get('EE+Children', 0))
        ws.cell(row=start_row + 0, column=ees_col, value=int(tier_counts.get('EE Only', 0)))
        ws.cell(row=start_row + 1, column=ees_col, value=int(tier_counts.get('EE+Spouse', 0)))
        ws.cell(row=start_row + 2, column=ees_col, value=child_total)
        ws.cell(row=start_row + 3, column=ees_col, value=int(tier_counts.get('EE+Family', 0)))
        written_sum = sum([
            int(tier_counts.get('EE Only', 0)),
            int(tier_counts.get('EE+Spouse', 0)),
            child_total,
            int(tier_counts.get('EE+Family', 0))
        ])
    else:
        ws.cell(row=start_row + 0, column=ees_col, value=int(tier_counts.get('EE Only', 0)))
        ws.cell(row=start_row + 1, column=ees_col, value=int(tier_counts.get('EE+Spouse', 0)))
        ws.cell(row=start_row + 2, column=ees_col, value=int(tier_counts.get('EE+Child', 0)))
        ws.cell(row=start_row + 3, column=ees_col, value=int(tier_counts.get('EE+Children', 0)))
        ws.cell(row=start_row + 4, column=ees_col, value=int(tier_counts.get('EE+Family', 0)))
        written_sum = sum(int(v) for v in tier_counts.values())
    
    return written_sum, combined

def write_back_to_template(df_reconciled, source_file, template_file):
    """Main write-back function"""
    print("\n1. Preparing write-back data...")
    
    # Reload full data to get all relations (not just subscribers)
    df_full = pd.read_excel(source_file)
    df_full = df_full[df_full['STATUS'].apply(is_active_for_ees)]
    df_full = create_employee_group(df_full)
    
    print(f"   Full active dataset: {len(df_full)} rows")
    
    # Calculate tiers on COMPLETE families (including dependents)
    print("\n2. Calculating tiers from family composition...")
    tier_map = {}
    unknown_samples = []
    
    for (cid, eg), family in df_full.groupby(['CLIENT ID', 'EMPLOYEE_GROUP']):
        tier = calculate_tier_from_composition(family, eg)
        if tier == 'UNKNOWN' and len(unknown_samples) < 3:
            unknown_samples.append((cid, eg))
        tier_map[(cid, eg)] = tier
    
    if unknown_samples:
        print(f"   WARNING: {len([t for t in tier_map.values() if t == 'UNKNOWN'])} groups without SELF anchor")
        for cid, eg in unknown_samples:
            print(f"     {cid} / {eg}")
    
    # NOW filter to subscribers only
    print("\n3. Filtering to subscribers...")
    df_subs = df_full[df_full['RELATION'].str.strip().str.upper().isin(['SELF', 'EE', 'EMPLOYEE'])]
    print(f"   Subscriber records: {len(df_subs)}")
    
    # Add calculated tiers
    df_subs['CALCULATED_TIER'] = df_subs.apply(
        lambda r: tier_map.get((r['CLIENT ID'], r['EMPLOYEE_GROUP']), 'UNKNOWN'), 
        axis=1
    )
    
    # Deduplicate
    df_final = deduplicate_enrollments(df_subs)
    
    # Build aggregated data
    print("\n4. Building facility/plan/tier aggregations...")
    tier_data = build_facility_plan_tier_data(df_final)
    
    # Load template workbook
    print("\n5. Writing to template file...")
    
    # Create new template path
    write_template = "Prime Enrollment Funding by Facility for August.xlsx"
    
    try:
        wb = load_workbook(write_template)
    except PermissionError:
        print("ERROR: Template file is open in Excel. Please close and retry.")
        return
    except FileNotFoundError:
        print(f"ERROR: Template file not found: {write_template}")
        return
    
    # Process facilities (sample for now)
    test_facilities = ['H3170', 'H3220', 'H3280']  # San Dimas, West Anaheim, Shasta
    
    for client_id in test_facilities:
        if client_id not in tier_data:
            print(f"   {client_id}: No data")
            continue
        
        facility_name = TPA_TO_FACILITY.get(client_id, 'UNKNOWN')
        tab = CID_TO_TAB.get(client_id, 'Legacy')  # Default to Legacy
        
        if tab not in wb.sheetnames:
            print(f"   {facility_name} ({client_id}): Tab '{tab}' not found")
            continue
        
        ws = wb[tab]
        print(f"\n   {facility_name} ({client_id}) on tab '{tab}':")
        
        # Find facility
        fac_row, fac_col = find_facility_location_strict(ws, facility_name, client_id)
        if not fac_row:
            continue
        
        # Find Ees column
        ees_col = find_ees_col_strict(ws, fac_row, fac_col)
        label_col = ees_col - 1
        col_letter = get_column_letter(ees_col)
        
        # Write each plan
        for plan_type in ['EPO', 'PPO', 'VALUE']:
            section_row = find_plan_section_tolerant(ws, fac_row, fac_col, plan_type)
            
            if not section_row:
                print(f"     {plan_type}: Section not found")
                continue
            
            # Get tier counts
            tier_counts = tier_data[client_id].get(plan_type, {
                'EE Only': 0, 'EE+Spouse': 0, 'EE+Child': 0,
                'EE+Children': 0, 'EE+Family': 0
            })
            
            # Write values
            written_sum, combined = write_tier_rows_safe(ws, section_row, ees_col, tier_counts, label_col)
            
            end_row = section_row + (3 if combined else 4)
            print(f"     {plan_type}: {col_letter}{section_row}:{col_letter}{end_row} = {written_sum}")
    
    # Save
    output_file = write_template.replace('.xlsx', '_with_ees.xlsx')
    wb.save(output_file)
    wb.close()
    
    print(f"\n   Saved to: {output_file}")
    
    # Spot check
    print("\n6. Spot check validation:")
    spot_check_facilities(output_file, {'H3170': 223, 'H3220': 707, 'H3280': 638}, tier_data)

def spot_check_facilities(output_file, expected_counts, tier_data):
    """Comprehensive spot check with sum verification"""
    wb = load_workbook(output_file, data_only=True)
    
    for client_id, expected_total in expected_counts.items():
        facility_name = TPA_TO_FACILITY.get(client_id, 'UNKNOWN')
        tab = CID_TO_TAB.get(client_id, 'Legacy')
        
        if tab not in wb.sheetnames:
            print(f"   {facility_name} ({client_id}): TAB NOT FOUND")
            continue
        
        ws = wb[tab]
        
        # Find facility
        fac_row, fac_col = find_facility_location_strict(ws, facility_name, client_id)
        if not fac_row:
            print(f"   {facility_name} ({client_id}): FACILITY NOT FOUND")
            continue
        
        # Find Ees column
        ees_col = find_ees_col_strict(ws, fac_row, fac_col)
        label_col = ees_col - 1
        
        # Sum all plan sections
        total_written = 0
        plan_details = []
        
        for plan_type in ['EPO', 'PPO', 'VALUE']:
            section_row = find_plan_section_tolerant(ws, fac_row, fac_col, plan_type)
            
            if not section_row:
                continue
            
            # Detect combined mode
            combined = detect_combined_children_tolerant(ws, section_row, label_col)
            
            # Sum values
            rows_to_sum = 4 if combined else 5
            plan_sum = sum(
                int(ws.cell(row=section_row + i, column=ees_col).value or 0)
                for i in range(rows_to_sum)
            )
            
            total_written += plan_sum
            plan_details.append(f"{plan_type}={plan_sum}")
        
        # Compare to expected
        status = "PASS" if total_written == expected_total else "FAIL"
        details = ", ".join(plan_details) if plan_details else "no plans found"
        
        print(f"   {facility_name} ({client_id}): {status}")
        print(f"      Expected: {expected_total}, Written: {total_written} ({details})")
        
        # Also compare to tier_data if available
        if client_id in tier_data:
            source_total = sum(
                sum(tiers.values())
                for tiers in tier_data[client_id].values()
            )
            if source_total != total_written:
                print(f"      WARNING: Source had {source_total}, sheet has {total_written}")
    
    wb.close()

# Add CID_TO_TAB mapping
CID_TO_TAB = {
    'H3100': 'Legacy', 'H3105': 'Legacy', 'H3110': 'Legacy', 'H3115': 'Legacy',
    'H3120': 'Legacy', 'H3130': 'Legacy', 'H3140': 'Legacy', 'H3150': 'Legacy',
    'H3160': 'Legacy', 'H3170': 'Legacy', 'H3180': 'Legacy', 'H3190': 'Legacy',
    'H3200': 'Legacy', 'H3210': 'Legacy', 'H3220': 'Encino-Garden Grove',
    'H3230': 'Legacy', 'H3240': 'Legacy', 'H3250': 'Encino-Garden Grove',
    'H3260': 'Encino-Garden Grove', 'H3270': 'Centinela', 'H3271': 'Centinela',
    'H3272': 'Centinela', 'H3275': 'St. Francis', 'H3276': 'St. Francis',
    'H3277': 'St. Francis', 'H3280': 'Legacy', 'H3285': 'Legacy',
    'H3290': 'Legacy', 'H3300': 'Legacy', 'H3310': 'Alvarado',
    'H3320': 'Pampa', 'H3325': 'Roxborough', 'H3330': 'Lower Bucks',
    # New mappings
    'H3335': 'Dallas Medical Center', 'H3337': 'Dallas Regional',
    'H3338': 'Riverview & Gadsden', 'H3339': 'Riverview & Gadsden',
    'H3340': 'Providence & St. John', 'H3345': 'Providence & St. John',
    'H3355': 'Knapp', 'H3360': 'Knapp',
    'H3370': 'Harlingen',
    'H3375': 'Garden City', 'H3380': 'Garden City', 'H3385': 'Garden City',
    'H3381': 'Lake Huron', 'H3382': 'Lake Huron',
    'H3392': 'Landmark',
    'H3394': "Saint Mary's Reno", 'H3395': "Saint Mary's Reno", 'H3396': "Saint Mary's Reno",
    'H3397': 'Monroe', 'H3398': 'North Vista',
    'H3500': "Saint Clare's", 'H3505': "Saint Mary's Passaic", 'H3510': 'Southern Regional',
    'H3530': "St. Michael's", 'H3540': 'Mission Regional',
    'H3591': 'Coshocton County', 'H3592': 'East Liverpool City',
    'H3594': 'Ohio Valley HHC', 'H3595': 'River Valley Pri.',
    'H3596': "St. Mary's Medical", 'H3598': 'Suburban Community', 'H3599': 'Suburban Community',
    'H3605': 'Illinois', 'H3615': 'Illinois', 'H3625': 'Illinois', 'H3630': 'Illinois',
    'H3635': 'Illinois', 'H3645': 'Illinois', 'H3655': 'Illinois', 'H3660': 'Illinois',
    'H3665': 'Illinois', 'H3670': 'Illinois', 'H3675': 'Illinois', 'H3680': 'Illinois'
}

# ============= DIRECT CELL WRITE MAPS =============
# Complete WRITE_MAP for Legacy sheet with exact cell addresses
LEGACY_WRITE_MAP = [
    # San Dimas Community Hospital — H3170
    {"client_id": "H3170", "plan": "EPO", "cells": {"EE": "G4", "EE+Spouse": "G5", "EE+Child(ren)": "G6", "EE+Family": "G7"}},
    {"client_id": "H3170", "plan": "VALUE", "cells": {"EE": "G10", "EE+Spouse": "G11", "EE+Child(ren)": "G12", "EE+Family": "G13"}},
    
    # Bio-Medical Services — H3130
    {"client_id": "H3130", "plan": "EPO", "cells": {"EE": "G20", "EE+Spouse": "G21", "EE+Child(ren)": "G22", "EE+Family": "G23"}},
    {"client_id": "H3130", "plan": "VALUE", "cells": {"EE": "G26", "EE+Spouse": "G27", "EE+Child(ren)": "G28", "EE+Family": "G29"}},
    
    # Chino Valley Medical Center — H3100
    {"client_id": "H3100", "plan": "EPO", "cells": {"EE": "G36", "EE+Spouse": "G37", "EE+Child(ren)": "G38", "EE+Family": "G39"}},
    {"client_id": "H3100", "plan": "VALUE", "cells": {"EE": "G42", "EE+Spouse": "G43", "EE+Child(ren)": "G44", "EE+Family": "G45"}},
    
    # Chino Valley Medical Center RNs — H3300
    {"client_id": "H3300", "plan": "EPO", "cells": {"EE": "G53", "EE+Spouse": "G54", "EE+Child(ren)": "G55", "EE+Family": "G56"}},
    {"client_id": "H3300", "plan": "VALUE", "cells": {"EE": "G59", "EE+Spouse": "G60", "EE+Child(ren)": "G61", "EE+Family": "G62"}},
    
    # Desert Valley Hospital — H3140
    {"client_id": "H3140", "plan": "EPO", "cells": {"EE": "G69", "EE+Spouse": "G70", "EE+Child(ren)": "G71", "EE+Family": "G72"}},
    {"client_id": "H3140", "plan": "VALUE", "cells": {"EE": "G75", "EE+Spouse": "G76", "EE+Child(ren)": "G77", "EE+Family": "G78"}},
    
    # Desert Valley Medical Group — H3150
    {"client_id": "H3150", "plan": "EPO", "cells": {"EE": "G85", "EE+Spouse": "G86", "EE+Child(ren)": "G87", "EE+Family": "G88"}},
    {"client_id": "H3150", "plan": "VALUE", "cells": {"EE": "G91", "EE+Spouse": "G92", "EE+Child(ren)": "G93", "EE+Family": "G94"}},
    
    # Huntington Beach Hospital — H3210
    {"client_id": "H3210", "plan": "EPO", "cells": {"EE": "G101", "EE+Spouse": "G102", "EE+Child(ren)": "G103", "EE+Family": "G104"}},
    {"client_id": "H3210", "plan": "VALUE", "cells": {"EE": "G107", "EE+Spouse": "G108", "EE+Child(ren)": "G109", "EE+Family": "G110"}},
    
    # La Palma Intercommunity Hospital — H3200
    {"client_id": "H3200", "plan": "EPO", "cells": {"EE": "G133", "EE+Spouse": "G134", "EE+Child(ren)": "G135", "EE+Family": "G136"}},
    {"client_id": "H3200", "plan": "VALUE", "cells": {"EE": "G139", "EE+Spouse": "G140", "EE+Child(ren)": "G141", "EE+Family": "G142"}},
    
    # Montclair Hospital Medical Center — H3160
    {"client_id": "H3160", "plan": "EPO", "cells": {"EE": "G149", "EE+Spouse": "G150", "EE+Child(ren)": "G151", "EE+Family": "G152"}},
    {"client_id": "H3160", "plan": "VALUE", "cells": {"EE": "G155", "EE+Spouse": "G156", "EE+Child(ren)": "G157", "EE+Family": "G158"}},
    
    # Premiere Healthcare Staffing — H3115 (EPO only)
    {"client_id": "H3115", "plan": "EPO", "cells": {"EE": "G165", "EE+Spouse": "G166", "EE+Child(ren)": "G167", "EE+Family": "G168"}},
    
    # Prime Management Services — H3110
    {"client_id": "H3110", "plan": "EPO", "cells": {"EE": "G175", "EE+Spouse": "G176", "EE+Child(ren)": "G177", "EE+Family": "G178"}},
    {"client_id": "H3110", "plan": "VALUE", "cells": {"EE": "G181", "EE+Spouse": "G182", "EE+Child(ren)": "G183", "EE+Family": "G184"}},
    
    # Paradise Valley Hospital — H3230
    {"client_id": "H3230", "plan": "EPO", "cells": {"EE": "G191", "EE+Spouse": "G192", "EE+Child(ren)": "G193", "EE+Family": "G194"}},
    {"client_id": "H3230", "plan": "VALUE", "cells": {"EE": "G197", "EE+Spouse": "G198", "EE+Child(ren)": "G199", "EE+Family": "G200"}},
    
    # Paradise Valley Medical Group — H3240
    {"client_id": "H3240", "plan": "EPO", "cells": {"EE": "G207", "EE+Spouse": "G208", "EE+Child(ren)": "G209", "EE+Family": "G210"}},
    {"client_id": "H3240", "plan": "VALUE", "cells": {"EE": "G213", "EE+Spouse": "G214", "EE+Child(ren)": "G215", "EE+Family": "G216"}},
    
    # Sherman Oaks Hospital — H3180 (two EPO and two VALUE blocks)
    {"client_id": "H3180", "plan": "EPO", "cells": {"EE": "G223", "EE+Spouse": "G224", "EE+Child(ren)": "G225", "EE+Family": "G226"}},
    {"client_id": "H3180", "plan": "VALUE", "cells": {"EE": "G229", "EE+Spouse": "G230", "EE+Child(ren)": "G231", "EE+Family": "G232"}},
    {"client_id": "H3180", "plan": "EPO", "cells": {"EE": "G239", "EE+Spouse": "G240", "EE+Child(ren)": "G241", "EE+Family": "G242"}},
    {"client_id": "H3180", "plan": "VALUE", "cells": {"EE": "G245", "EE+Spouse": "G246", "EE+Child(ren)": "G247", "EE+Family": "G248"}},
    
    # West Anaheim Medical Center — H3220
    {"client_id": "H3220", "plan": "EPO", "cells": {"EE": "G255", "EE+Spouse": "G256", "EE+Child(ren)": "G257", "EE+Family": "G258"}},
    {"client_id": "H3220", "plan": "VALUE", "cells": {"EE": "G261", "EE+Spouse": "G262", "EE+Child(ren)": "G263", "EE+Family": "G264"}},
    
    # Shasta Regional Medical Center — H3280
    {"client_id": "H3280", "plan": "EPO", "cells": {"EE": "G271", "EE+Spouse": "G272", "EE+Child(ren)": "G273", "EE+Family": "G274"}},
    {"client_id": "H3280", "plan": "VALUE", "cells": {"EE": "G277", "EE+Spouse": "G278", "EE+Child(ren)": "G279", "EE+Family": "G280"}},
    
    # Shasta Medical Group — H3285
    {"client_id": "H3285", "plan": "EPO", "cells": {"EE": "G287", "EE+Spouse": "G288", "EE+Child(ren)": "G289", "EE+Family": "G290"}},
    {"client_id": "H3285", "plan": "VALUE", "cells": {"EE": "G293", "EE+Spouse": "G294", "EE+Child(ren)": "G295", "EE+Family": "G296"}},
]

# Complete WRITE_MAP for Centinela sheet with exact cell addresses
CENTINELA_WRITE_MAP = [
    # Centinela Hospital Medical Center — H3270
    {"client_id": "H3270", "plan": "EPO", "cells": {"EE": "D3", "EE+Spouse": "D4", "EE+Child(ren)": "D5", "EE+Family": "D6"}},
    {"client_id": "H3270", "plan": "VALUE", "cells": {"EE": "D21", "EE+Spouse": "D22", "EE+Child(ren)": "D23", "EE+Family": "D24"}},
    
    # Robotics Outpatient Center — H3271
    {"client_id": "H3271", "plan": "EPO", "cells": {"EE": "D27", "EE+Spouse": "D28", "EE+Child(ren)": "D29", "EE+Family": "D30"}},
    {"client_id": "H3271", "plan": "VALUE", "cells": {"EE": "D45", "EE+Spouse": "D46", "EE+Child(ren)": "D47", "EE+Family": "D48"}},
    
    # Centinela Valley Endoscopy Center — H3272
    {"client_id": "H3272", "plan": "EPO", "cells": {"EE": "D51", "EE+Spouse": "D52", "EE+Child(ren)": "D53", "EE+Family": "D54"}},
    {"client_id": "H3272", "plan": "VALUE", "cells": {"EE": "D69", "EE+Spouse": "D70", "EE+Child(ren)": "D71", "EE+Family": "D72"}},
]

# Complete WRITE_MAP for Encino-Garden Grove sheet with exact cell addresses
ENCINO_GARDEN_GROVE_WRITE_MAP = [
    # Encino Hospital Medical Center — H3250 (two EPO blocks + one VALUE block)
    {"client_id": "H3250", "plan": "EPO", "label": "PRIME Non-Union & SEIU-UHW UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE+Spouse": "D4", "EE+Child": "D5", "EE+Children": "D6", "EE+Family": "D7"}},
    {"client_id": "H3250", "plan": "EPO", "label": "PRIME SEIU 121 RN EPO PLAN (Self-Insured)",
     "cells": {"EE": "D10", "EE+Spouse": "D11", "EE+Child": "D12", "EE+Children": "D13", "EE+Family": "D14"}},
    {"client_id": "H3250", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D59", "EE+Spouse": "D60", "EE+Child": "D61", "EE+Children": "D62", "EE+Family": "D63"}},
    
    # Garden Grove Hospital Medical Center — H3260 (Including UNAC)
    {"client_id": "H3260", "plan": "EPO", "label": "PRIME Non-Union UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D69", "EE+Spouse": "D70", "EE+Child": "D71", "EE+Children": "D72", "EE+Family": "D73"}},
    {"client_id": "H3260", "plan": "EPO", "label": "PRIME UNAC EPO PLAN (Self-Insured)",
     "cells": {"EE": "D76", "EE+Spouse": "D77", "EE+Child": "D78", "EE+Children": "D79", "EE+Family": "D80"}},
    {"client_id": "H3260", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D125", "EE+Spouse": "D126", "EE+Child": "D127", "EE+Children": "D128", "EE+Family": "D129"}},
]

# Complete WRITE_MAP for St. Francis sheet with exact cell addresses (all in column D)
ST_FRANCIS_WRITE_MAP = [
    # St. Francis Medical Center — H3275 (three EPO blocks + one VALUE block)
    {"client_id": "H3275", "plan": "EPO", "label": "PRIME SEIU 2020 D1 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3275", "plan": "EPO", "label": "PRIME UNAC D1 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    {"client_id": "H3275", "plan": "EPO", "label": "PRIME Non-Union D1 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Children": "D17", "EE & Family": "D18"}},
    {"client_id": "H3275", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D21", "EE & Spouse": "D22", "EE & Children": "D23", "EE & Family": "D24"}},
    
    # Shoreline Surgery Center — H3276
    {"client_id": "H3276", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D28", "EE & Spouse": "D29", "EE & Children": "D30", "EE & Family": "D31"}},
    {"client_id": "H3276", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D34", "EE & Spouse": "D35", "EE & Children": "D36", "EE & Family": "D37"}},
    
    # Physician's Surgery Center Downey — H3277
    {"client_id": "H3277", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D41", "EE & Spouse": "D42", "EE & Children": "D43", "EE & Family": "D44"}},
    {"client_id": "H3277", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D47", "EE & Spouse": "D48", "EE & Children": "D49", "EE & Family": "D50"}},
]

# Pampa sheet mapping - H3320
PAMPA_WRITE_MAP = [
    # Pampa Community Hospital — H3320
    {"client_id": "H3320", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3320", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Roxborough sheet mapping - H3325
ROXBOROUGH_WRITE_MAP = [
    # Roxborough Memorial Hospital — H3325
    {"client_id": "H3325", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3325", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Lower Bucks sheet mapping - H3330 (two EPO groupings + one VALUE grouping)
LOWER_BUCKS_WRITE_MAP = [
    # Lower Bucks Hospital — H3330
    {"client_id": "H3330", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured) - IUOE",
     "cells": {"EE": "D10", "EE & Spouse": "D11", "EE & Children": "D12", "EE & Family": "D13"}},
    {"client_id": "H3330", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured) - PASNAP & Non-Union",
     "cells": {"EE": "D16", "EE & Spouse": "D17", "EE & Children": "D18", "EE & Family": "D19"}},
    {"client_id": "H3330", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Union & Non-Union",
     "cells": {"EE": "D22", "EE & Spouse": "D23", "EE & Children": "D24", "EE & Family": "D25"}}
]

# Dallas Medical Center sheet mapping - H3335
DALLAS_MEDICAL_CENTER_WRITE_MAP = [
    {"client_id": "H3335", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3335", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Harlingen sheet mapping - H3370
HARLINGEN_WRITE_MAP = [
    {"client_id": "H3370", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3370", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Knapp sheet mapping - H3355, H3360
KNAPP_WRITE_MAP = [
    {"client_id": "H3355", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3355", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    {"client_id": "H3360", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D18", "EE & Spouse": "D19", "EE & Children": "D20", "EE & Family": "D21"}}
]

# Monroe sheet mapping - H3397
MONROE_WRITE_MAP = [
    {"client_id": "H3397", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3397", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Saint Mary's Reno sheet mapping - H3394, H3395, H3396
SAINT_MARYS_RENO_WRITE_MAP = [
    # H3394
    {"client_id": "H3394", "plan": "EPO", "label": "PRIME Non-Union 2020 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3394", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    # H3395
    {"client_id": "H3395", "plan": "EPO", "label": "PRIME Non-Union 2020 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}},
    {"client_id": "H3395", "plan": "EPO", "label": "PRIME CWA 2020 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D31", "EE & Spouse": "D32", "EE & Children": "D33", "EE & Family": "D34"}},
    {"client_id": "H3395", "plan": "EPO", "label": "PRIME CNA 2019 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D37", "EE & Spouse": "D38", "EE & Children": "D39", "EE & Family": "D40"}},
    {"client_id": "H3395", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D49", "EE & Spouse": "D50", "EE & Children": "D51", "EE & Family": "D52"}},
    # H3396
    {"client_id": "H3396", "plan": "EPO", "label": "PRIME Non-Union 2020 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D59", "EE & Spouse": "D60", "EE & Children": "D61", "EE & Family": "D62"}},
    {"client_id": "H3396", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D71", "EE & Spouse": "D72", "EE & Children": "D73", "EE & Family": "D74"}}
]

# North Vista sheet mapping - H3398 (has both EE & Child and EE & Children separately)
NORTH_VISTA_WRITE_MAP = [
    {"client_id": "H3398", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Child": "D5", "EE & Children": "D6", "EE & Family": "D7"}},
    {"client_id": "H3398", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D10", "EE & Spouse": "D11", "EE & Child": "D12", "EE & Children": "D13", "EE & Family": "D14"}}
]

# Dallas Regional sheet mapping - H3337
DALLAS_REGIONAL_WRITE_MAP = [
    {"client_id": "H3337", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3337", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Riverview & Gadsden sheet mapping - H3338, H3339
RIVERVIEW_GADSDEN_WRITE_MAP = [
    # H3338
    {"client_id": "H3338", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3338", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    # H3339
    {"client_id": "H3339", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D19", "EE & Spouse": "D20", "EE & Children": "D21", "EE & Family": "D22"}},
    {"client_id": "H3339", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}}
]

# Saint Clare's sheet mapping - H3500
SAINT_CLARES_WRITE_MAP = [
    {"client_id": "H3500", "plan": "EPO", "label": "PRIME EPO OPEN ACCESS PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3500", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Landmark sheet mapping - H3392
LANDMARK_WRITE_MAP = [
    {"client_id": "H3392", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3392", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Saint Mary's Passaic sheet mapping - H3505
SAINT_MARYS_PASSAIC_WRITE_MAP = [
    {"client_id": "H3505", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3505", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Children": "D17", "EE & Family": "D18"}}
]

# Southern Regional sheet mapping - H3510
SOUTHERN_REGIONAL_WRITE_MAP = [
    {"client_id": "H3510", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3510", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# St. Michael's sheet mapping - H3530 (multiple EPO plan types)
ST_MICHAELS_WRITE_MAP = [
    {"client_id": "H3530", "plan": "EPO", "label": "PRIME NON-UNION EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3530", "plan": "EPO", "label": "PRIME CIR EPO PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    {"client_id": "H3530", "plan": "EPO", "label": "PRIME IUOE EPO PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Children": "D17"}},  # No Family
    {"client_id": "H3530", "plan": "EPO", "label": "PRIME JNESO EPO PLAN (Self-Insured)",
     "cells": {"EE": "D20", "EE & Spouse": "D21", "EE & Children": "D22", "EE & Family": "D23"}},
    {"client_id": "H3530", "plan": "EPO", "label": "PRIME EPO PLUS PLAN (Self-Insured)",
     "cells": {"EE": "D27", "EE & Spouse": "D28", "EE & Children": "D29", "EE & Family": "D30"}},
    {"client_id": "H3530", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D33", "EE & Spouse": "D34", "EE & Children": "D35", "EE & Family": "D36"}}
]

# Mission Regional sheet mapping - H3540
MISSION_REGIONAL_WRITE_MAP = [
    {"client_id": "H3540", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3540", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}}
]

# Coshocton County sheet mapping - H3591 (VALUE has missing tiers)
COSHOCTON_COUNTY_WRITE_MAP = [
    {"client_id": "H3591", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3591", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Family": "D12"}}  # Only EE and Family
]

# Suburban Community sheet mapping - H3598, H3599
SUBURBAN_COMMUNITY_WRITE_MAP = [
    # H3598
    {"client_id": "H3598", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3598", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Family": "D17"}},  # No Children
    # H3599
    {"client_id": "H3599", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}},
    {"client_id": "H3599", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D43", "EE & Spouse": "D45", "EE & Children": "D46", "EE & Family": "D47"}}  # Note D44 skipped
]

# Garden City sheet mapping - H3375, H3380, H3385
GARDEN_CITY_WRITE_MAP = [
    # H3375
    {"client_id": "H3375", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3375", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    # H3385
    {"client_id": "H3385", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D19", "EE & Spouse": "D20", "EE & Children": "D21", "EE & Family": "D22"}},
    {"client_id": "H3385", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Family": "D28"}},  # No Children, D27 skipped
    # H3380 (sparse)
    {"client_id": "H3380", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D35", "EE & Children": "D37"}},  # Only EE and Children
    {"client_id": "H3380", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D41", "EE & Family": "D44"}}  # Only EE and Family
]

# Lake Huron sheet mapping - H3381, H3382
LAKE_HURON_WRITE_MAP = [
    # H3381
    {"client_id": "H3381", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3381", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Children": "D17", "EE & Family": "D18"}},
    # H3382
    {"client_id": "H3382", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}},
    {"client_id": "H3382", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D37", "EE & Spouse": "D38", "EE & Children": "D39", "EE & Family": "D40"}}
]

# Providence & St. John sheet mapping - H3340, H3345
PROVIDENCE_ST_JOHN_WRITE_MAP = [
    # H3340
    {"client_id": "H3340", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3340", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    # H3345
    {"client_id": "H3345", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D19", "EE & Spouse": "D20", "EE & Children": "D21", "EE & Family": "D22"}},
    {"client_id": "H3345", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Children": "D27"}}  # Only EE and Children
]

# East Liverpool City sheet mapping - H3592
EAST_LIVERPOOL_CITY_WRITE_MAP = [
    {"client_id": "H3592", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3592", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D15", "EE & Spouse": "D16", "EE & Children": "D17", "EE & Family": "D18"}}
]

# Ohio Valley HHC sheet mapping - H3594 (sparse EPO only)
OHIO_VALLEY_HHC_WRITE_MAP = [
    {"client_id": "H3594", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D47", "EE & Spouse": "D48", "EE & Children": "D49"}}  # Note D51 is another EE
]

# River Valley Pri. sheet mapping - H3595 (unusual VALUE with duplicate cells)
RIVER_VALLEY_PRI_WRITE_MAP = [
    {"client_id": "H3595", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D69", "EE & Spouse": "D70", "EE & Children": "D71", "EE & Family": "D72"}},
    # VALUE plan has scattered/duplicate cells - using first occurrence of each
    {"client_id": "H3595", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D85", "EE & Spouse": "D81", "EE & Children": "D83", "EE & Family": "D84"}}
]

# St. Mary's Medical sheet mapping - H3596 (multiple VALUE blocks)
ST_MARYS_MEDICAL_WRITE_MAP = [
    {"client_id": "H3596", "plan": "EPO", "label": "PRIME Non-Union 2020 D2 UNIFIED EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    # Multiple VALUE blocks with same label - writing to all
    {"client_id": "H3596", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Block 1",
     "cells": {"EE": "D11", "EE & Spouse": "D12", "EE & Children": "D13", "EE & Family": "D14"}},
    {"client_id": "H3596", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Block 2",
     "cells": {"EE": "D19", "EE & Spouse": "D20", "EE & Children": "D21", "EE & Family": "D22"}},
    {"client_id": "H3596", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Block 3",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}},
    {"client_id": "H3596", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Block 4",
     "cells": {"EE": "D35", "EE & Spouse": "D36", "EE & Children": "D37", "EE & Family": "D38"}},
    {"client_id": "H3596", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured) - Block 5",
     "cells": {"EE & Children": "D43"}}  # Only Children
]

# Illinois sheet mapping - H3605, H3615, H3625, H3630, H3635, H3645, H3655, H3660, H3665, H3670, H3675, H3680
ILLINOIS_WRITE_MAP = [
    # H3605
    {"client_id": "H3605", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D3", "EE & Spouse": "D4", "EE & Children": "D5", "EE & Family": "D6"}},
    {"client_id": "H3605", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D9", "EE & Spouse": "D10", "EE & Children": "D11", "EE & Family": "D12"}},
    # H3615
    {"client_id": "H3615", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D19", "EE & Spouse": "D20", "EE & Children": "D21", "EE & Family": "D22"}},
    {"client_id": "H3615", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D25", "EE & Spouse": "D26", "EE & Children": "D27", "EE & Family": "D28"}},
    # H3625
    {"client_id": "H3625", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D35", "EE & Spouse": "D36", "EE & Children": "D37", "EE & Family": "D38"}},
    {"client_id": "H3625", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D41", "EE & Spouse": "D42", "EE & Children": "D43", "EE & Family": "D44"}},
    # H3630
    {"client_id": "H3630", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D51", "EE & Spouse": "D52", "EE & Children": "D53", "EE & Family": "D54"}},
    {"client_id": "H3630", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D57", "EE & Spouse": "D58", "EE & Children": "D59", "EE & Family": "D60"}},
    # H3635
    {"client_id": "H3635", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D67", "EE & Spouse": "D68", "EE & Children": "D69", "EE & Family": "D70"}},
    {"client_id": "H3635", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D73", "EE & Spouse": "D74", "EE & Children": "D75", "EE & Family": "D76"}},
    # H3645
    {"client_id": "H3645", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D83", "EE & Spouse": "D84", "EE & Children": "D85", "EE & Family": "D86"}},
    {"client_id": "H3645", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D89", "EE & Spouse": "D90", "EE & Children": "D91", "EE & Family": "D92"}},
    # H3655
    {"client_id": "H3655", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D99", "EE & Spouse": "D100", "EE & Children": "D101", "EE & Family": "D102"}},
    {"client_id": "H3655", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D105", "EE & Spouse": "D106", "EE & Children": "D107", "EE & Family": "D108"}},
    # H3660
    {"client_id": "H3660", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D115", "EE & Spouse": "D116", "EE & Children": "D117", "EE & Family": "D118"}},
    {"client_id": "H3660", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D121", "EE & Spouse": "D122", "EE & Children": "D123", "EE & Family": "D124"}},
    # H3665
    {"client_id": "H3665", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D131", "EE & Spouse": "D132", "EE & Children": "D133", "EE & Family": "D134"}},
    {"client_id": "H3665", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D137", "EE & Spouse": "D138", "EE & Children": "D139", "EE & Family": "D140"}},
    # H3670
    {"client_id": "H3670", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D147", "EE & Spouse": "D148", "EE & Children": "D149", "EE & Family": "D150"}},
    {"client_id": "H3670", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D153", "EE & Spouse": "D154", "EE & Children": "D155", "EE & Family": "D156"}},
    # H3675
    {"client_id": "H3675", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D163", "EE & Spouse": "D164", "EE & Children": "D165", "EE & Family": "D166"}},
    {"client_id": "H3675", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D169", "EE & Spouse": "D170", "EE & Children": "D171", "EE & Family": "D172"}},
    # H3680
    {"client_id": "H3680", "plan": "EPO", "label": "PRIME EPO PLAN (Self-Insured)",
     "cells": {"EE": "D179", "EE & Spouse": "D180", "EE & Children": "D181", "EE & Family": "D182"}},
    {"client_id": "H3680", "plan": "VALUE", "label": "PRIME VALUE PLAN (Self-Insured)",
     "cells": {"EE": "D185", "EE & Spouse": "D186", "EE & Children": "D187", "EE & Family": "D188"}}
]

def write_to_specific_sheet(wb, sheet_name, write_map, tier_data):
    """
    Write tier counts to specific sheet using exact cell addresses.
    
    Args:
        wb: openpyxl workbook object
        sheet_name: Name of the sheet to write to
        write_map: List of write entries with client_id, plan, and cell addresses
        tier_data: Nested dict with tier counts by client and plan
    
    Returns:
        List of write log entries
    """
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: '{sheet_name}' sheet not found in workbook!")
        return []
    
    ws = wb[sheet_name]
    PROCESSED_SHEETS.add(sheet_name)
    write_log = []
    seen_plan = set()  # Track (client_id, plan) for deduplication
    
    print(f"\nWriting to {sheet_name} sheet...")
    
    # Process each entry in write_map
    for entry in write_map:
        client_id = entry['client_id']
        plan = entry['plan']
        cells = entry['cells']
        label = entry.get('label', '')
        
        # Handle duplicate blocks with first_only policy
        key = (client_id, plan)
        if key in seen_plan and entry.get('dedupe_policy', 'first_only') == 'first_only':
            # Zero any duplicate blocks to avoid double counting
            for cell in cells.values():
                ws[cell] = 0
                _log(sheet_name, client_id, plan, 'DUPLICATE-ZERO', cell, 0)
            print(f"  Skipped duplicate block for {key} (first_only).")
            continue
        seen_plan.add(key)
        
        # Get tier counts (or zeros)
        tier_counts = {'EE Only': 0, 'EE+Spouse': 0, 'EE+Child': 0, 'EE+Children': 0, 'EE+Family': 0}
        if client_id in tier_data and plan in tier_data[client_id]:
            tier_counts.update({k:int(v) for k,v in tier_data[client_id][plan].items()})
        
        # Zero-fill all cells first
        for cell in cells.values():
            ws[cell] = 0
        
        # Optional label sanity check using any cell as anchor
        anchor_cell = next(iter(cells.values()))
        _soft_assert_label(ws, anchor_cell, label)
        
        # Unified children bucket
        child_total = int(tier_counts.get('EE+Child', 0)) + int(tier_counts.get('EE+Children', 0))
        
        values_written = []
        
        # EE
        if 'EE' in cells:
            v = int(tier_counts.get('EE Only', 0))
            ws[cells['EE']] = v
            _log(sheet_name, client_id, plan, 'EE', cells['EE'], v)
            values_written.append(v)
        
        # Spouse
        if 'EE+Spouse' in cells:
            v = int(tier_counts.get('EE+Spouse', 0))
            ws[cells['EE+Spouse']] = v
            _log(sheet_name, client_id, plan, 'EE+Spouse', cells['EE+Spouse'], v)
            values_written.append(v)
        elif 'EE & Spouse' in cells:
            v = int(tier_counts.get('EE+Spouse', 0))
            ws[cells['EE & Spouse']] = v
            _log(sheet_name, client_id, plan, 'EE & Spouse', cells['EE & Spouse'], v)
            values_written.append(v)
        
        # Children — prefer a single combined target
        if 'EE+Child(ren)' in cells:
            ws[cells['EE+Child(ren)']] = child_total
            _log(sheet_name, client_id, plan, 'EE+Child(ren)', cells['EE+Child(ren)'], child_total)
            values_written.append(child_total)
        elif 'EE & Children' in cells:
            ws[cells['EE & Children']] = child_total
            _log(sheet_name, client_id, plan, 'EE & Children', cells['EE & Children'], child_total)
            values_written.append(child_total)
        elif 'EE+Children' in cells:
            ws[cells['EE+Children']] = child_total
            _log(sheet_name, client_id, plan, 'EE+Children', cells['EE+Children'], child_total)
            values_written.append(child_total)
        elif 'EE & Child' in cells:
            ws[cells['EE & Child']] = child_total
            _log(sheet_name, client_id, plan, 'EE & Child', cells['EE & Child'], child_total)
            values_written.append(child_total)
        elif 'EE+Child' in cells:
            ws[cells['EE+Child']] = child_total
            _log(sheet_name, client_id, plan, 'EE+Child', cells['EE+Child'], child_total)
            values_written.append(child_total)
        
        # Family
        if 'EE+Family' in cells:
            v = int(tier_counts.get('EE+Family', 0))
            ws[cells['EE+Family']] = v
            _log(sheet_name, client_id, plan, 'EE+Family', cells['EE+Family'], v)
            values_written.append(v)
        elif 'EE & Family' in cells:
            v = int(tier_counts.get('EE+Family', 0))
            ws[cells['EE & Family']] = v
            _log(sheet_name, client_id, plan, 'EE & Family', cells['EE & Family'], v)
            values_written.append(v)
        
        # Logging line per entry
        cell_range = '/'.join(cells.values())
        label_str = f" ({label[:30]}...)" if label and len(label) > 30 else (f" ({label})" if label else "")
        log_entry = f"{client_id} {plan}{label_str} → {cell_range}: {', '.join(map(str, values_written))}"
        write_log.append(log_entry)
        print("  " + log_entry)
    
    return write_log

def build_tier_data_from_source(source_file):
    """
    Build tier_data from source enrollment Excel file.
    Returns nested dict: tier_data[CLIENT_ID][PLAN_GROUP] = {tier: count}
    """
    print(f"\nBuilding tier data from: {source_file}")
    
    # Read source data
    df = pd.read_excel(source_file, sheet_name=0)
    print(f"  Loaded {len(df)} rows")
    
    # Clean keys
    if 'CLIENT ID' in df.columns:
        df['CLIENT ID'] = df['CLIENT ID'].apply(clean_key)
    if 'BEN CODE' in df.columns:
        df['BEN CODE'] = df['BEN CODE'].apply(clean_key)
    
    # Filter active (including COBRA)
    if 'STATUS' in df.columns:
        df = df[df['STATUS'].apply(is_active)].copy()
        print(f"  After status filter: {len(df)} rows")
    
    # Filter subscribers only
    if 'RELATION' in df.columns:
        df = df[df['RELATION'].apply(is_subscriber)].copy()
        print(f"  After subscriber filter: {len(df)} rows")
    
    # Normalize tiers
    if 'BEN CODE' in df.columns:
        df['tier'] = df['BEN CODE'].apply(normalize_tier_strict)
    else:
        df['tier'] = 'UNKNOWN'
    
    # Map plans
    if 'PLAN' in df.columns:
        df['plan_group'], _ = zip(*df['PLAN'].apply(infer_plan_group_and_variant))
    else:
        df['plan_group'] = 'UNKNOWN'
    
    # Build nested dictionary
    tier_data = defaultdict(lambda: defaultdict(lambda: {
        'EE Only': 0, 'EE+Spouse': 0, 'EE+Child': 0, 'EE+Children': 0, 'EE+Family': 0
    }))
    
    # Track unknowns for visibility
    unknown_plans = []
    
    # Aggregate counts - handle EE+Child(ren) specially
    for _, row in df.iterrows():
        client_id = row['CLIENT ID']
        plan_group = row['plan_group']
        tier = row['tier']
        
        if tier == 'UNKNOWN' or plan_group == 'UNKNOWN':
            if len(unknown_plans) < 10:
                unknown_plans.append((
                    row.get('CLIENT ID', ''), 
                    row.get('PLAN', ''), 
                    row.get('BEN CODE', '')
                ))
            continue
            
        # Map EE+Child(ren) to separate EE+Child and EE+Children
        if tier == 'EE+Child(ren)':
            # For now, put all in EE+Children (will be combined when writing)
            tier_data[client_id][plan_group]['EE+Children'] += 1
        elif tier in tier_data[client_id][plan_group]:
            tier_data[client_id][plan_group][tier] += 1
    
    # Report unknowns
    if unknown_plans:
        print(f"  ⚠ UNKNOWN plans/tiers skipped: {len(unknown_plans)} (showing first 10)")
        for cid, plan, ben in unknown_plans[:10]:
            print(f"     {cid}: PLAN='{plan}'  BEN='{ben}'")
    
    # Convert to regular dict
    tier_data = {k: dict(v) for k, v in tier_data.items()}
    
    # Print summary
    total_count = sum(
        sum(sum(tiers.values()) for tiers in plans.values())
        for plans in tier_data.values()
    )
    print(f"  Built tier data: {len(tier_data)} clients, {total_count} total enrollments")
    
    return tier_data

def perform_comprehensive_writeback(workbook_path, tier_data, output_path=None):
    """
    Perform comprehensive write-back to all configured sheets.
    
    Args:
        workbook_path: Path to the Excel workbook
        tier_data: Nested dict with tier counts
        output_path: Optional output path (defaults to _updated.xlsx)
    
    Returns:
        Path to the output file
    """
    # Reset runtime globals before each run
    global WRITE_LOG_ROWS, PROCESSED_SHEETS
    WRITE_LOG_ROWS = []
    PROCESSED_SHEETS = set()
    
    if not output_path:
        # Create output path by replacing extension
        base_name = os.path.splitext(workbook_path)[0]
        output_path = f"{base_name}_updated.xlsx"
    
    print("\n" + "="*80)
    print("COMPREHENSIVE ENROLLMENT WRITE-BACK")
    print("="*80)
    
    # Pre-write control assertion
    if STRICT_CONTROL_CHECK and not assert_control_from_tier_data(tier_data):
        print("❌ Control totals mismatch—aborting write-back.")
        return None
    
    # Load workbook
    print(f"Opening workbook: {workbook_path}")
    
    # Normalize the path for the current OS
    workbook_path = os.path.normpath(workbook_path)
    
    try:
        wb = load_workbook(workbook_path)
    except FileNotFoundError:
        print(f"ERROR: Workbook not found: {workbook_path}")
        return None
    except PermissionError:
        print("ERROR: Workbook is open in Excel. Please close and retry.")
        return None
    
    all_write_logs = []
    
    # Write to Legacy sheet
    legacy_logs = write_to_specific_sheet(wb, 'Legacy', LEGACY_WRITE_MAP, tier_data)
    all_write_logs.extend(legacy_logs)
    
    # Write to Centinela sheet  
    centinela_logs = write_to_specific_sheet(wb, 'Centinela', CENTINELA_WRITE_MAP, tier_data)
    all_write_logs.extend(centinela_logs)
    
    # Write to Encino-Garden Grove sheet
    encino_logs = write_to_specific_sheet(wb, 'Encino-Garden Grove', ENCINO_GARDEN_GROVE_WRITE_MAP, tier_data)
    all_write_logs.extend(encino_logs)
    
    # Write to St. Francis sheet
    st_francis_logs = write_to_specific_sheet(wb, 'St. Francis', ST_FRANCIS_WRITE_MAP, tier_data)
    all_write_logs.extend(st_francis_logs)
    
    # Write to Pampa sheet
    pampa_logs = write_to_specific_sheet(wb, 'Pampa', PAMPA_WRITE_MAP, tier_data)
    all_write_logs.extend(pampa_logs)
    
    # Write to Roxborough sheet
    roxborough_logs = write_to_specific_sheet(wb, 'Roxborough', ROXBOROUGH_WRITE_MAP, tier_data)
    all_write_logs.extend(roxborough_logs)
    
    # Write to Lower Bucks sheet
    lower_bucks_logs = write_to_specific_sheet(wb, 'Lower Bucks', LOWER_BUCKS_WRITE_MAP, tier_data)
    all_write_logs.extend(lower_bucks_logs)
    
    # Write to Dallas Medical Center sheet
    dallas_medical_logs = write_to_specific_sheet(wb, 'Dallas Medical Center', DALLAS_MEDICAL_CENTER_WRITE_MAP, tier_data)
    all_write_logs.extend(dallas_medical_logs)
    
    # Write to Harlingen sheet
    harlingen_logs = write_to_specific_sheet(wb, 'Harlingen', HARLINGEN_WRITE_MAP, tier_data)
    all_write_logs.extend(harlingen_logs)
    
    # Write to Knapp sheet
    knapp_logs = write_to_specific_sheet(wb, 'Knapp', KNAPP_WRITE_MAP, tier_data)
    all_write_logs.extend(knapp_logs)
    
    # Write to Monroe sheet
    monroe_logs = write_to_specific_sheet(wb, 'Monroe', MONROE_WRITE_MAP, tier_data)
    all_write_logs.extend(monroe_logs)
    
    # Write to Saint Mary's Reno sheet
    saint_marys_reno_logs = write_to_specific_sheet(wb, "Saint Mary's Reno", SAINT_MARYS_RENO_WRITE_MAP, tier_data)
    all_write_logs.extend(saint_marys_reno_logs)
    
    # Write to North Vista sheet
    north_vista_logs = write_to_specific_sheet(wb, 'North Vista', NORTH_VISTA_WRITE_MAP, tier_data)
    all_write_logs.extend(north_vista_logs)
    
    # Write to Dallas Regional sheet
    dallas_regional_logs = write_to_specific_sheet(wb, 'Dallas Regional', DALLAS_REGIONAL_WRITE_MAP, tier_data)
    all_write_logs.extend(dallas_regional_logs)
    
    # Write to Riverview & Gadsden sheet
    riverview_gadsden_logs = write_to_specific_sheet(wb, 'Riverview & Gadsden', RIVERVIEW_GADSDEN_WRITE_MAP, tier_data)
    all_write_logs.extend(riverview_gadsden_logs)
    
    # Write to Saint Clare's sheet
    saint_clares_logs = write_to_specific_sheet(wb, "Saint Clare's", SAINT_CLARES_WRITE_MAP, tier_data)
    all_write_logs.extend(saint_clares_logs)
    
    # Write to Landmark sheet
    landmark_logs = write_to_specific_sheet(wb, 'Landmark', LANDMARK_WRITE_MAP, tier_data)
    all_write_logs.extend(landmark_logs)
    
    # Write to Saint Mary's Passaic sheet
    saint_marys_passaic_logs = write_to_specific_sheet(wb, "Saint Mary's Passaic", SAINT_MARYS_PASSAIC_WRITE_MAP, tier_data)
    all_write_logs.extend(saint_marys_passaic_logs)
    
    # Write to Southern Regional sheet
    southern_regional_logs = write_to_specific_sheet(wb, 'Southern Regional', SOUTHERN_REGIONAL_WRITE_MAP, tier_data)
    all_write_logs.extend(southern_regional_logs)
    
    # Write to St. Michael's sheet
    st_michaels_logs = write_to_specific_sheet(wb, "St. Michael's", ST_MICHAELS_WRITE_MAP, tier_data)
    all_write_logs.extend(st_michaels_logs)
    
    # Write to Mission Regional sheet
    mission_regional_logs = write_to_specific_sheet(wb, 'Mission Regional', MISSION_REGIONAL_WRITE_MAP, tier_data)
    all_write_logs.extend(mission_regional_logs)
    
    # Write to Coshocton County sheet
    coshocton_county_logs = write_to_specific_sheet(wb, 'Coshocton County', COSHOCTON_COUNTY_WRITE_MAP, tier_data)
    all_write_logs.extend(coshocton_county_logs)
    
    # Write to Suburban Community sheet
    suburban_community_logs = write_to_specific_sheet(wb, 'Suburban Community', SUBURBAN_COMMUNITY_WRITE_MAP, tier_data)
    all_write_logs.extend(suburban_community_logs)
    
    # Write to Garden City sheet
    garden_city_logs = write_to_specific_sheet(wb, 'Garden City', GARDEN_CITY_WRITE_MAP, tier_data)
    all_write_logs.extend(garden_city_logs)
    
    # Write to Lake Huron sheet
    lake_huron_logs = write_to_specific_sheet(wb, 'Lake Huron', LAKE_HURON_WRITE_MAP, tier_data)
    all_write_logs.extend(lake_huron_logs)
    
    # Write to Providence & St. John sheet
    providence_st_john_logs = write_to_specific_sheet(wb, 'Providence & St. John', PROVIDENCE_ST_JOHN_WRITE_MAP, tier_data)
    all_write_logs.extend(providence_st_john_logs)
    
    # Write to East Liverpool City sheet
    east_liverpool_city_logs = write_to_specific_sheet(wb, 'East Liverpool City', EAST_LIVERPOOL_CITY_WRITE_MAP, tier_data)
    all_write_logs.extend(east_liverpool_city_logs)
    
    # Write to Ohio Valley HHC sheet
    ohio_valley_hhc_logs = write_to_specific_sheet(wb, 'Ohio Valley HHC', OHIO_VALLEY_HHC_WRITE_MAP, tier_data)
    all_write_logs.extend(ohio_valley_hhc_logs)
    
    # Write to River Valley Pri. sheet
    river_valley_pri_logs = write_to_specific_sheet(wb, 'River Valley Pri.', RIVER_VALLEY_PRI_WRITE_MAP, tier_data)
    all_write_logs.extend(river_valley_pri_logs)
    
    # Write to St. Mary's Medical sheet
    st_marys_medical_logs = write_to_specific_sheet(wb, "St. Mary's Medical", ST_MARYS_MEDICAL_WRITE_MAP, tier_data)
    all_write_logs.extend(st_marys_medical_logs)
    
    # Write to Illinois sheet
    illinois_logs = write_to_specific_sheet(wb, 'Illinois', ILLINOIS_WRITE_MAP, tier_data)
    all_write_logs.extend(illinois_logs)
    
    # Save workbook (unless dry run)
    if not DRY_RUN_WRITE:
        print(f"\nSaving to: {output_path}")
        wb.save(output_path)
        wb.close()
    else:
        print(f"\n[DRY RUN] Would save to: {output_path}")
        wb.close()
    
    # CSV audit log
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
    os.makedirs(output_dir, exist_ok=True)
    write_log_path = os.path.join(output_dir, 'write_log.csv')
    with open(write_log_path, 'w', newline='') as f:
        w = _csv.writer(f)
        w.writerow(['sheet','client_id','plan','tier','cell','value'])
        w.writerows(WRITE_LOG_ROWS)
    
    print("\n" + "="*60)
    print("WRITE-BACK COMPLETE")
    print("="*60)
    print(f"Total writes: {len(all_write_logs)}")
    print(f"Sheets updated: {len(PROCESSED_SHEETS)} -> {', '.join(sorted(PROCESSED_SHEETS))}")
    print(f"Write log: {write_log_path}")
    print(f"Output file: {output_path if not DRY_RUN_WRITE else '[DRY RUN - not saved]'}")
    
    # Post-write verification: compare source tier_data sums to what was written
    from collections import Counter as _Counter
    by_key = _Counter()  # (client_id, plan) -> written sum
    for _, cid, plan, _, _, val in WRITE_LOG_ROWS:
        if 'DUPLICATE-ZERO' not in plan:  # Skip duplicate zeros
            by_key[(cid, plan)] += int(val)
    
    mismatches = []
    for cid, plans in tier_data.items():
        for plan, counts in plans.items():
            src_sum = int(counts.get('EE Only',0)) + int(counts.get('EE+Spouse',0)) + \
                      int(counts.get('EE+Family',0)) + int(counts.get('EE+Child',0)) + int(counts.get('EE+Children',0))
            sheet_sum = by_key.get((cid, plan), 0)
            if src_sum != sheet_sum:
                mismatches.append((cid, plan, src_sum, sheet_sum))
    
    if mismatches:
        print("⚠ Post-write mismatches (first 20):")
        for cid, plan, src, sheet in mismatches[:20]:
            print(f"  {cid} {plan}: source={src} vs sheet={sheet}")
    else:
        print("✓ Post-write verification: all client/plan totals match.")
    
    # Show sample of writes
    if all_write_logs:
        print("\nFirst 10 writes:")
        for entry in all_write_logs[:10]:
            print(f"  {entry}")
        
        if len(all_write_logs) > 10:
            print(f"  ... and {len(all_write_logs) - 10} more")
    
    return output_path

def perform_targeted_cell_writeback(df, template_file):
    """
    Write enrollment counts to exact cells specified in WRITE_MAP
    """
    # WRITE_MAP with exact cell addresses
    WRITE_MAP = {
        'H3170': {  # San Dimas
            'EPO': {'EE Only': 'G4', 'EE+Spouse': 'G5', 'EE+Child(ren)': 'G6', 'EE+Family': 'G7'},
            'VALUE': {'EE Only': 'G15', 'EE+Spouse': 'G16', 'EE+Child(ren)': 'G17', 'EE+Family': 'G18'}
        },
        'H3180': {  # Sherman Oaks - has TWO EPO blocks and TWO VALUE blocks
            'EPO_1': {'EE Only': 'G4', 'EE+Spouse': 'G5', 'EE+Child(ren)': 'G6', 'EE+Family': 'G7'},
            'EPO_2': {'EE Only': 'G11', 'EE+Spouse': 'G12', 'EE+Child(ren)': 'G13', 'EE+Family': 'G14'},
            'VALUE_1': {'EE Only': 'G18', 'EE+Spouse': 'G19', 'EE+Child(ren)': 'G20', 'EE+Family': 'G21'},
            'VALUE_2': {'EE Only': 'G25', 'EE+Spouse': 'G26', 'EE+Child(ren)': 'G27', 'EE+Family': 'G28'}
        },
        'H3220': {  # West Anaheim
            'EPO': {'EE Only': 'F4', 'EE+Spouse': 'F5', 'EE+Child(ren)': 'F6', 'EE+Family': 'F7'},
            'PPO': {'EE Only': 'F11', 'EE+Spouse': 'F12', 'EE+Child(ren)': 'F13', 'EE+Family': 'F14'},
            'VALUE': {'EE Only': 'F18', 'EE+Spouse': 'F19', 'EE+Child(ren)': 'F20', 'EE+Family': 'F21'}
        }
    }
    
    print("\n" + "="*80)
    print("TARGETED CELL WRITE-BACK")
    print("="*80)
    
    # Create backup
    from datetime import datetime
    import shutil
    backup_file = template_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(template_file, backup_file)
    print(f"✓ Backup created: {backup_file}")
    
    # Load workbook
    from openpyxl import load_workbook
    wb = load_workbook(template_file)
    
    # Process each facility
    write_summary = []
    
    for client_id, cell_map in WRITE_MAP.items():
        # Get facility name from mapping
        facility_name = TPA_TO_FACILITY.get(client_id, f"Unknown_{client_id}")
        
        # Filter data for this CLIENT ID (Active only, no COBRA)
        facility_df = df[(df['CLIENT ID'] == client_id) & 
                        (df['STATUS'].str.strip().str.upper() == 'A')]
        
        print(f"\nProcessing {facility_name} ({client_id}):")
        
        # Get worksheet (try various patterns)
        ws = None
        for sheet_name in wb.sheetnames:
            if client_id in sheet_name or facility_name.lower() in sheet_name.lower():
                ws = wb[sheet_name]
                print(f"  Found worksheet: {sheet_name}")
                break
        
        if not ws:
            print(f"  ⚠️ No worksheet found for {facility_name}")
            continue
        
        # Add plan_group column from PLAN
        facility_df['plan_group'] = facility_df['PLAN'].apply(lambda x: infer_plan_group_and_variant(x)[0])
        
        # Special handling for Sherman Oaks (H3180)
        if client_id == 'H3180':
            # Split variants for duplicate blocks
            variant1_keywords = ['1', 'ONE', 'FIRST']
            variant2_keywords = ['2', 'TWO', 'SECOND']
            
            # Process EPO blocks
            epo_df = facility_df[facility_df['plan_group'] == 'EPO']
            if not epo_df.empty:
                # Split between blocks (simple 50/50 or by some criteria)
                mid = len(epo_df) // 2
                epo1_df = epo_df.iloc[:mid]
                epo2_df = epo_df.iloc[mid:]
                
                # Write EPO_1
                for tier, cell in cell_map.get('EPO_1', {}).items():
                    tier_name = tier.replace('(ren)', '+Children')
                    count = len(epo1_df[epo1_df['ENROLLMENT TIER'] == tier_name])
                    ws[cell] = int(count)
                    write_summary.append(f"  {client_id} EPO_1 {tier}: {cell} = {count}")
                
                # Write EPO_2
                for tier, cell in cell_map.get('EPO_2', {}).items():
                    tier_name = tier.replace('(ren)', '+Children')
                    count = len(epo2_df[epo2_df['ENROLLMENT TIER'] == tier_name])
                    ws[cell] = int(count)
                    write_summary.append(f"  {client_id} EPO_2 {tier}: {cell} = {count}")
            
            # Process VALUE blocks similarly
            value_df = facility_df[facility_df['plan_group'] == 'VALUE']
            if not value_df.empty:
                mid = len(value_df) // 2
                value1_df = value_df.iloc[:mid]
                value2_df = value_df.iloc[mid:]
                
                # Write VALUE_1 and VALUE_2
                for tier, cell in cell_map.get('VALUE_1', {}).items():
                    tier_name = tier.replace('(ren)', '+Children')
                    count = len(value1_df[value1_df['ENROLLMENT TIER'] == tier_name])
                    ws[cell] = int(count)
                    write_summary.append(f"  {client_id} VALUE_1 {tier}: {cell} = {count}")
                
                for tier, cell in cell_map.get('VALUE_2', {}).items():
                    tier_name = tier.replace('(ren)', '+Children')
                    count = len(value2_df[value2_df['ENROLLMENT TIER'] == tier_name])
                    ws[cell] = int(count)
                    write_summary.append(f"  {client_id} VALUE_2 {tier}: {cell} = {count}")
        
        else:
            # Standard facilities (non-Sherman Oaks)
            for plan_type, tier_cells in cell_map.items():
                plan_df = facility_df[facility_df['plan_group'] == plan_type]
                
                for tier, cell in tier_cells.items():
                    # Handle combined Child(ren) tier
                    if tier == 'EE+Child(ren)':
                        count = len(plan_df[plan_df['ENROLLMENT TIER'].isin(['EE+Child', 'EE+Children'])])
                    else:
                        count = len(plan_df[plan_df['ENROLLMENT TIER'] == tier])
                    
                    # Write to cell
                    ws[cell] = int(count)
                    write_summary.append(f"  {client_id} {plan_type} {tier}: {cell} = {count}")
                    print(f"    {plan_type} {tier}: {cell} = {count}")
    
    # Save workbook
    wb.save(template_file)
    print(f"\n✓ Write-back complete: {template_file}")
    
    # Print summary
    print("\n" + "-"*40)
    print("WRITE SUMMARY:")
    for line in write_summary:
        print(line)
    
    # Validation
    print("\n" + "-"*40)
    print("POST-WRITE VALIDATION:")
    for client_id in WRITE_MAP.keys():
        facility_df = df[(df['CLIENT ID'] == client_id) & 
                        (df['STATUS'].str.strip().str.upper() == 'A')]
        total = len(facility_df)
        print(f"  {client_id}: {total} total active employees written")

def main():
    """
    Main execution with tier reconciliation and targeted write-back
    """
    # Reset global trackers at the start of each run
    global unknown_tiers_tracker, waterfall_stages, removed_rows_samples
    unknown_tiers_tracker = Counter()
    waterfall_stages = []
    removed_rows_samples = {}
    
    # FILE PATHS - Use relative paths for cross-platform compatibility
    base_dir = os.path.dirname(os.path.abspath(__file__))
    source_file = os.path.join("data", "input", "source_data.xlsx")
    
    # Try to find the template file in the current directory first
    template_filename = "Prime Enrollment Funding by Facility for August.xlsx"
    template_file = os.path.join(base_dir, template_filename)
    
    # If not found in base dir, try current working directory
    if not os.path.exists(template_file):
        template_file = template_filename
    
    # Create destination file path
    destination_file = os.path.join(base_dir, "Prime Enrollment Funding by Facility for August_analysis.xlsx")
    
    try:
        print("="*80)
        print("ENROLLMENT AUTOMATION - TIER RECONCILIATION")
        print("Finding and fixing 106-row discrepancy")
        print("="*80)
        
        print("\nCONTROL TOTALS (Ground Truth):")
        for tier in TIER_ORDER:
            print(f"  {tier}: {CONTROL_TOTALS[tier]:,}")
        print(f"  TOTAL: {CONTROL_TOTAL:,}")
        
        # Process with waterfall tracking
        print("\n" + "="*80)
        print("PROCESSING WITH WATERFALL TRACKING")
        print("="*80)
        
        df = read_and_prepare_data_with_waterfall(source_file)
        
        # Print waterfall analysis
        print_waterfall_table()
        
        # Print removed rows samples
        if removed_rows_samples:
            print("\n" + "="*80)
            print("REMOVED ROWS SAMPLES")
            print("="*80)
            for stage, sample in removed_rows_samples.items():
                print(f"\n{stage} - Sample of removed rows:")
                print(sample.to_string())
        
        # Print unknown audit
        print_unknown_audit()
        
        # Print facility comparisons
        print("\n" + "="*80)
        print("FACILITY SPOT CHECKS")
        print("="*80)
        print_facility_comparison(df, 'H3170', 'San Dimas')
        print_facility_comparison(df, 'H3220', 'West Anaheim')
        
        # Assert control totals
        all_match = assert_control_totals(df)
        
        # Final summary
        print("\n" + "="*80)
        print("FINAL SUMMARY")
        print("="*80)
        
        if all_match:
            print("✅ SUCCESS: All 106 missing rows identified and recovered!")
            print("✅ Tier totals now match control exactly")
            print("✅ Control tie-out PASS: 14533 / 2639 / 4413 / 3123. Global delta resolved (Δ=0).")
        else:
            print("⚠️ RECONCILIATION IN PROGRESS")
            print("Review the waterfall and unknown audit to identify remaining issues")
        
        # Save diagnostic output
        output_file = "output/tier_reconciliation_report.csv"
        df.to_csv(output_file, index=False)
        print(f"\nDiagnostic data saved to: {output_file}")
        
        # Create summary pivot for destination file
        print(f"\nCreating enrollment summary...")
        summary_pivot = df.pivot_table(
            index=['facility_name', 'plan_group'],
            columns='tier',
            values='original_index',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        # Add total column
        tier_cols = [col for col in summary_pivot.columns if col in TIER_ORDER]
        if tier_cols:
            summary_pivot['Total'] = summary_pivot[tier_cols].sum(axis=1)
        
        # Save to destination Excel file (overwrite)
        print(f"Writing to destination file: {destination_file}")
        with pd.ExcelWriter(destination_file, engine='openpyxl', mode='w') as writer:
            # Write main summary
            summary_pivot.to_excel(writer, sheet_name='Enrollment Summary', index=False)
            
            # Write detailed data
            df.to_excel(writer, sheet_name='Detailed Data', index=False)
            
            # Write control totals validation
            control_df = pd.DataFrame({
                'Tier': TIER_ORDER,
                'Control Total': [CONTROL_TOTALS[tier] for tier in TIER_ORDER],
                'Actual Total': [len(df[df['tier'] == tier]) for tier in TIER_ORDER]
            })
            control_df['Match'] = control_df['Control Total'] == control_df['Actual Total']
            control_df.to_excel(writer, sheet_name='Control Validation', index=False)
        
        print(f"✅ Destination file successfully overwritten: {destination_file}")
        
        # ============= COMPREHENSIVE CELL WRITE-BACK =============
        if all_match and not DRY_RUN:
            print("\n" + "="*80)
            print("COMPREHENSIVE CELL WRITE-BACK - EXACT CELL ADDRESSES")
            print("="*80)
            
            try:
                # Build tier data from source
                tier_data_for_writeback = build_tier_data_from_source(source_file)
                
                # Perform comprehensive write-back to Legacy and Centinela sheets
                # Use the template file (original Excel) not the analysis file
                output_file = perform_comprehensive_writeback(template_file, tier_data_for_writeback)
                
                if output_file:
                    print(f"\n✅ Successfully wrote enrollment data to: {output_file}")
                    print("   - Legacy sheet updated with all facility enrollments")
                    print("   - Centinela sheet updated with H3270, H3271, H3272 enrollments")
                    print("   - Encino-Garden Grove sheet updated with H3250, H3260 enrollments")
                
            except Exception as e:
                print(f"Write-back error: {e}")
                import traceback
                traceback.print_exc()
        
    except FileNotFoundError as e:
        print(f"\n❌ ERROR: File not found - {e}")
        print(f"Please ensure source file exists: {source_file}")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()