#!/usr/bin/env python3
"""
Excel Cell Discovery Script
============================

This script helps find where enrollment values should go in the Excel template
by searching for Client IDs and plan indicators.
"""

from openpyxl import load_workbook
import json

def discover_cell_locations(workbook_path, output_json="cell_mappings_discovered.json"):
    """Discover where enrollment values should be written"""
    
    print(f"Loading workbook: {workbook_path}")
    wb = load_workbook(workbook_path, data_only=True)
    
    # Dictionary to store discovered mappings
    mappings = {}
    
    # Tabs to search
    tabs_to_search = [
        'Monroe', 'Lower Bucks', 'Legacy', 'St. Francis', 
        'Encino-Garden Grove', 'North Vista', 'Centinela',
        'Illinois', 'Saint Mary\'s Reno'
    ]
    
    for tab_name in tabs_to_search:
        if tab_name not in wb.sheetnames:
            print(f"⚠️ Tab '{tab_name}' not found")
            continue
            
        print(f"\nSearching tab: {tab_name}")
        ws = wb[tab_name]
        mappings[tab_name] = {}
        
        # Search for Client IDs (like H3397, H3330, etc.)
        for row in range(1, min(200, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                
                if cell_value and isinstance(cell_value, str):
                    # Check if it looks like a Client ID
                    if cell_value.startswith('H3') and len(cell_value) == 5:
                        client_id = cell_value
                        print(f"  Found Client ID {client_id} at row {row}, col {col}")
                        
                        if client_id not in mappings[tab_name]:
                            mappings[tab_name][client_id] = {
                                'location': f"Row {row}, Col {col}",
                                'EPO': {},
                                'VALUE': {}
                            }
                        
                        # Look for EPO section below
                        for search_row in range(row + 1, min(row + 30, ws.max_row + 1)):
                            row_text = ""
                            for search_col in range(1, min(10, ws.max_column + 1)):
                                val = ws.cell(row=search_row, column=search_col).value
                                if val:
                                    row_text += str(val).upper()
                            
                            if 'EPO' in row_text and 'VALUE' not in row_text:
                                # Found EPO section, now find tiers
                                print(f"    EPO section at row {search_row}")
                                
                                # Look for tier rows
                                for tier_row in range(search_row + 1, min(search_row + 10, ws.max_row + 1)):
                                    tier_col = 3  # Usually column C
                                    value_col = 4  # Usually column D
                                    
                                    tier_label = ws.cell(row=tier_row, column=tier_col).value
                                    if tier_label and any(x in str(tier_label).upper() for x in ['EE', 'EMPLOYEE', 'SPOUSE', 'CHILD', 'FAMILY']):
                                        cell_ref = ws.cell(row=tier_row, column=value_col).coordinate
                                        
                                        # Normalize tier name
                                        tier_upper = str(tier_label).upper()
                                        if 'FAMILY' in tier_upper:
                                            tier_key = 'EE+Family'
                                        elif 'SPOUSE' in tier_upper:
                                            tier_key = 'EE+Spouse'
                                        elif 'CHILD' in tier_upper:
                                            tier_key = 'EE+Child(ren)'
                                        else:
                                            tier_key = 'EE Only'
                                        
                                        mappings[tab_name][client_id]['EPO'][tier_key] = cell_ref
                                        print(f"      {tier_key}: {cell_ref}")
                            
                            elif 'VALUE' in row_text:
                                # Found VALUE section
                                print(f"    VALUE section at row {search_row}")
                                
                                # Look for tier rows
                                for tier_row in range(search_row + 1, min(search_row + 10, ws.max_row + 1)):
                                    tier_col = 3  # Usually column C
                                    value_col = 4  # Usually column D
                                    
                                    tier_label = ws.cell(row=tier_row, column=tier_col).value
                                    if tier_label and any(x in str(tier_label).upper() for x in ['EE', 'EMPLOYEE', 'SPOUSE', 'CHILD', 'FAMILY']):
                                        cell_ref = ws.cell(row=tier_row, column=value_col).coordinate
                                        
                                        # Normalize tier name
                                        tier_upper = str(tier_label).upper()
                                        if 'FAMILY' in tier_upper:
                                            tier_key = 'EE+Family'
                                        elif 'SPOUSE' in tier_upper:
                                            tier_key = 'EE+Spouse'
                                        elif 'CHILD' in tier_upper:
                                            tier_key = 'EE+Child(ren)'
                                        else:
                                            tier_key = 'EE Only'
                                        
                                        mappings[tab_name][client_id]['VALUE'][tier_key] = cell_ref
                                        print(f"      {tier_key}: {cell_ref}")
    
    # Save mappings to JSON
    with open(output_json, 'w') as f:
        json.dump(mappings, f, indent=2)
    
    print(f"\n✓ Mappings saved to: {output_json}")
    print(f"  Found {sum(len(v) for v in mappings.values())} facility mappings")
    
    return mappings

if __name__ == "__main__":
    workbook_path = "Prime Enrollment Funding by Facility for August.xlsx"
    mappings = discover_cell_locations(workbook_path)
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    for tab, facilities in mappings.items():
        if facilities:
            print(f"\n{tab}:")
            for client_id, data in facilities.items():
                epo_count = len(data.get('EPO', {}))
                value_count = len(data.get('VALUE', {}))
                print(f"  {client_id}: {epo_count} EPO cells, {value_count} VALUE cells")