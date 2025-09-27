#!/usr/bin/env python3
"""
Fix Enrollment Discrepancies
============================
Comprehensive script to fix all enrollment count discrepancies and data quality issues.
"""

import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

# Add project root to path
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

# Import validation data
from pdf_validation_data import PDF_VALIDATION_DATA, TOTAL_CELL_FORMULAS

# Import from existing modules
from enrollment_automation_v6 import (
    load_block_aggregations,
    read_and_prepare_data,
    build_tier_data_from_source,
    load_plan_mappings
)

from write_maps import SHEET_WRITE_MAPS

# Define cells that should have formulas instead of hard-coded values
FORMULA_CELLS = {
    "Providence & St John": {
        "D7": "=SUM(D3:D6)",   # Providence EPO total
        "D13": "=SUM(D9:D12)",  # Providence PPO total
        "D19": "=SUM(D15:D18)", # Providence VALUE total
        "D25": "=SUM(D21:D24)", # St John EPO total
        "D31": "=SUM(D27:D30)"  # St John VALUE total
    },
    "Illinois": {
        # To be expanded based on actual Illinois tab layout
        # Will add formulas for all facility total rows
    },
    "Legacy": {
        # Check for any total rows that should have formulas
        "G8": "=SUM(G4:G7)",    # San Dimas EPO total
        "G14": "=SUM(G10:G13)",  # San Dimas VALUE total
        "G24": "=SUM(G20:G23)",  # Bio-Med EPO total
        "G30": "=SUM(G26:G29)",  # Bio-Med VALUE total
        # Add more as needed
    }
}

# Define cells that should be blank (no values)
CELLS_TO_CLEAR = {
    "Legacy": [
        # Hidden/inactive rows that might have orphaned values
        "G117", "G118", "G119", "G120",  # If there are hidden PPO rows
    ],
    "Illinois": [
        # Cells identified as having stray values
        # Will be populated based on inspection
    ]
}

def load_enrollment_data():
    """Load enrollment data using existing infrastructure."""
    print("Loading enrollment data...")
    plan_mappings = load_plan_mappings()
    block_aggregations = load_block_aggregations()
    
    source_data_path = os.path.join(PROJECT_ROOT, "data", "input", "source_data.xlsx")
    if not os.path.exists(source_data_path):
        print(f"✗ Source data not found: {source_data_path}")
        return None
    
    df = read_and_prepare_data(source_data_path, plan_mappings)
    tier_data = build_tier_data_from_source(df, block_aggregations, False)
    print(f"✓ Loaded data for {len(tier_data)} facilities")
    return tier_data

def validate_against_pdf(client_id, plan_type, tier, value):
    """Validate a value against PDF data."""
    if client_id not in PDF_VALIDATION_DATA:
        return None, "No validation data"
    
    # Map tier labels from system format to PDF format
    tier_mapping = {
        "EE Only": "EE",
        "EE+Spouse": "EE & Spouse",
        "EE+Child(ren)": "EE & Children",
        "EE+Family": "EE & Family",
        # 5-tier mappings
        "EE+Child": "EE & Child",
        "EE+Children": "EE & Children",
        "EE+1 Dep": "EE & Child"
    }
    
    pdf_tier = tier_mapping.get(tier, tier)
    
    pdf_data = PDF_VALIDATION_DATA[client_id].get(plan_type, {})
    
    # Handle multi-block facilities
    if isinstance(pdf_data, dict) and any(isinstance(v, dict) for v in pdf_data.values()):
        # Multi-block - sum all blocks for this tier
        total = 0
        for block_data in pdf_data.values():
            if isinstance(block_data, dict):
                total += block_data.get(pdf_tier, 0)
        expected = total
    else:
        expected = pdf_data.get(pdf_tier, 0)
    
    if value == expected:
        return True, f"✓"
    else:
        return False, f"Expected {expected}, got {value}"

def write_enrollment_with_validation(wb, tier_data):
    """Write enrollment data with validation against PDF values."""
    print("\n" + "="*70)
    print("WRITING AND VALIDATING ENROLLMENT DATA")
    print("="*70)
    
    discrepancies = []
    total_writes = 0
    total_matches = 0
    
    for sheet_name, write_map in SHEET_WRITE_MAPS.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found in workbook")
            continue
        
        ws = wb[sheet_name]
        sheet_writes = 0
        sheet_matches = 0
        sheet_issues = []
        
        # Process each write instruction
        for instruction in write_map:
            client_id = instruction["client_id"]
            plan_type = instruction["plan"]
            cells = instruction["cells"]
            
            # Get expected values from PDF
            pdf_expected = PDF_VALIDATION_DATA.get(client_id, {}).get(plan_type, {})
            
            # Get actual enrollment data
            if client_id not in tier_data:
                if any(v > 0 for v in (pdf_expected.values() if isinstance(pdf_expected, dict) else [])):
                    sheet_issues.append(f"Missing data for {client_id} {plan_type}")
                continue
            
            if plan_type not in tier_data[client_id]:
                if any(v > 0 for v in (pdf_expected.values() if isinstance(pdf_expected, dict) else [])):
                    sheet_issues.append(f"Missing {plan_type} data for {client_id}")
                continue
            
            # Aggregate all blocks for this plan type
            tier_counts = {}
            for block_label, block_tiers in tier_data[client_id][plan_type].items():
                for tier, count in block_tiers.items():
                    if tier not in tier_counts:
                        tier_counts[tier] = 0
                    tier_counts[tier] += count
            
            # Write to cells with validation
            for tier, cell_ref in cells.items():
                value = tier_counts.get(tier, 0)
                
                # Validate against PDF
                is_valid, msg = validate_against_pdf(client_id, plan_type, tier, value)
                
                try:
                    ws[cell_ref] = value
                    sheet_writes += 1
                    total_writes += 1
                    
                    if is_valid:
                        sheet_matches += 1
                        total_matches += 1
                        if value > 0:
                            print(f"  ✓ {sheet_name}/{client_id} {plan_type} {tier}: {value} → {cell_ref}")
                    else:
                        sheet_issues.append(f"{client_id} {plan_type} {tier} @ {cell_ref}: {msg}")
                        print(f"  ✗ {sheet_name}/{client_id} {plan_type} {tier}: {msg}")
                        
                except Exception as e:
                    sheet_issues.append(f"Error writing to {cell_ref}: {str(e)}")
        
        if sheet_writes > 0:
            match_rate = (sheet_matches / sheet_writes * 100) if sheet_writes > 0 else 0
            status = "✓" if match_rate == 100 else "⚠️"
            print(f"{status} {sheet_name}: {sheet_writes} cells, {match_rate:.1f}% match rate")
            
            if sheet_issues:
                discrepancies.append({
                    "sheet": sheet_name,
                    "issues": sheet_issues,
                    "match_rate": match_rate
                })
    
    return total_writes, total_matches, discrepancies

def apply_formulas(wb):
    """Replace hard-coded values with formulas in total cells."""
    print("\n" + "="*70)
    print("APPLYING FORMULAS TO TOTAL CELLS")
    print("="*70)
    
    formulas_applied = 0
    
    for sheet_name, formulas in FORMULA_CELLS.items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        for cell_ref, formula in formulas.items():
            try:
                ws[cell_ref] = formula
                formulas_applied += 1
                print(f"  ✓ {sheet_name}!{cell_ref} = {formula}")
            except Exception as e:
                print(f"  ✗ Error applying formula to {sheet_name}!{cell_ref}: {e}")
    
    print(f"\n✓ Applied {formulas_applied} formulas")
    return formulas_applied

def clear_blank_cells(wb):
    """Clear cells that should be blank but contain stray values."""
    print("\n" + "="*70)
    print("CLEARING CONTAMINATED BLANK CELLS")
    print("="*70)
    
    cells_cleared = 0
    
    for sheet_name, cell_list in CELLS_TO_CLEAR.items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        for cell_ref in cell_list:
            try:
                current_value = ws[cell_ref].value
                if current_value is not None and current_value != 0:
                    ws[cell_ref] = None
                    cells_cleared += 1
                    print(f"  ✓ Cleared {sheet_name}!{cell_ref} (was {current_value})")
            except Exception as e:
                print(f"  ✗ Error clearing {sheet_name}!{cell_ref}: {e}")
    
    print(f"\n✓ Cleared {cells_cleared} contaminated cells")
    return cells_cleared

def check_illinois_tab(wb):
    """Special handling for Illinois tab data quality issues."""
    print("\n" + "="*70)
    print("CHECKING ILLINOIS TAB")
    print("="*70)
    
    if "Illinois" not in wb.sheetnames:
        print("✗ Illinois tab not found")
        return
    
    ws = wb["Illinois"]
    
    # Expected Illinois facilities from PDF
    illinois_facilities = {
        "H3560": "St. Mary's Hospital",
        "H3561": "St. Mary's Physicians", 
        "H3562": "St. Mary's Behavioral Health",
        "H3564": "St. Joseph Medical Center",
        "H3565": "St. Joseph Physicians",
        "H3595": "St. Joseph Behavioral Health"
    }
    
    print("Illinois facilities to validate:")
    for client_id, name in illinois_facilities.items():
        pdf_data = PDF_VALIDATION_DATA.get(client_id, {})
        if pdf_data:
            total_enrollment = 0
            for plan_data in pdf_data.values():
                if isinstance(plan_data, dict):
                    total_enrollment += sum(plan_data.values())
            print(f"  {client_id} {name}: {total_enrollment} total enrollees")

def generate_fix_report(discrepancies, total_writes, total_matches):
    """Generate a detailed fix report."""
    print("\n" + "="*70)
    print("FIX REPORT SUMMARY")
    print("="*70)
    
    match_rate = (total_matches / total_writes * 100) if total_writes > 0 else 0
    
    print(f"Total cells written: {total_writes}")
    print(f"Matching PDF values: {total_matches}")
    print(f"Overall match rate: {match_rate:.1f}%")
    
    if discrepancies:
        print(f"\n⚠️ Sheets with discrepancies: {len(discrepancies)}")
        for disc in discrepancies:
            print(f"\n{disc['sheet']} (Match rate: {disc['match_rate']:.1f}%):")
            for issue in disc['issues'][:5]:  # Show first 5 issues
                print(f"  - {issue}")
            if len(disc['issues']) > 5:
                print(f"  ... and {len(disc['issues']) - 5} more issues")
    
    # Save detailed report to file
    report_path = f"fix_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(report_path, 'w') as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "total_writes": total_writes,
            "total_matches": total_matches,
            "match_rate": match_rate,
            "discrepancies": discrepancies
        }, f, indent=2)
    
    print(f"\n✓ Detailed report saved to: {report_path}")

def main():
    """Main execution."""
    print("\n" + "="*80)
    print("ENROLLMENT DISCREPANCY FIX SCRIPT")
    print("="*80)
    
    # Step 1: Load enrollment data
    tier_data = load_enrollment_data()
    if not tier_data:
        return 1
    
    # Step 2: Load Excel template
    template_path = "Prime Enrollment Funding by Facility for August.xlsx"
    if not os.path.exists(template_path):
        print(f"✗ Template not found: {template_path}")
        return 1
    
    print(f"\nLoading Excel template...")
    wb = load_workbook(template_path)
    print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets")
    
    # Step 3: Write enrollment data with validation
    total_writes, total_matches, discrepancies = write_enrollment_with_validation(wb, tier_data)
    
    # Step 4: Apply formulas to total cells
    formulas_applied = apply_formulas(wb)
    
    # Step 5: Clear contaminated blank cells
    cells_cleared = clear_blank_cells(wb)
    
    # Step 6: Special check for Illinois tab
    check_illinois_tab(wb)
    
    # Step 7: Save the fixed workbook
    output_path = f"enrollment_fixed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"\nSaving fixed workbook...")
    wb.save(output_path)
    print(f"✓ Saved to: {output_path}")
    
    # Step 8: Generate fix report
    generate_fix_report(discrepancies, total_writes, total_matches)
    
    # Final summary
    print("\n" + "="*80)
    print("EXECUTION COMPLETE")
    print("="*80)
    
    match_rate = (total_matches / total_writes * 100) if total_writes > 0 else 0
    
    if match_rate == 100:
        print("✅ SUCCESS: All values match PDF validation data!")
    elif match_rate >= 95:
        print(f"✓ MOSTLY COMPLETE: {match_rate:.1f}% match rate")
        print("Review the fix report for remaining discrepancies.")
    else:
        print(f"⚠️ NEEDS ATTENTION: {match_rate:.1f}% match rate")
        print("Significant discrepancies found. Review the fix report for details.")
    
    return 0 if match_rate >= 95 else 1

if __name__ == "__main__":
    sys.exit(main())