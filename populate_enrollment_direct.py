#!/usr/bin/env python3
"""
Direct Excel Population Script
==============================

This script directly writes enrollment values to specific Excel cells
based on the tier data from enrollment_automation_v6.

No VBA, no CSV, just direct cell writes.
"""

import os
import sys
from openpyxl import load_workbook
from datetime import datetime
import json

# Add project root to path
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

from enrollment_automation_v6 import (
    CID_TO_TAB,
    load_block_aggregations,
    read_and_prepare_data,
    build_tier_data_from_source,
    load_plan_mappings
)

# Define cell mappings for each tab/facility/plan combination
# This is a simplified example - you'll need to expand this based on your Excel template
CELL_MAPPINGS = {
    'Monroe': {
        'H3397': {
            'EPO': {
                'EE Only': 'D15',
                'EE+Spouse': 'D16', 
                'EE+Child(ren)': 'D17',
                'EE+Family': 'D18'
            },
            'VALUE': {
                'EE Only': 'D20',
                'EE+Spouse': 'D21',
                'EE+Child(ren)': 'D22', 
                'EE+Family': 'D23'
            }
        }
    },
    'Lower Bucks': {
        'H3330': {
            'EPO': {
                # These would be the actual cell references in your Excel
                'EE Only': 'D10',
                'EE+Spouse': 'D11',
                'EE+Child(ren)': 'D12',
                'EE+Family': 'D13'
            },
            'VALUE': {
                'EE Only': 'D15',
                'EE+Spouse': 'D16',
                'EE+Child(ren)': 'D17',
                'EE+Family': 'D18'
            }
        }
    },
    # Add more mappings as needed...
}

def get_enrollment_data():
    """Load and process enrollment data using existing logic"""
    print("Loading enrollment data...")
    
    # Load configurations
    plan_mappings = load_plan_mappings()
    block_aggregations = load_block_aggregations()
    
    # Load source data
    source_data_path = os.path.join(PROJECT_ROOT, "data", "input", "source_data.xlsx")
    df = read_and_prepare_data(source_data_path, plan_mappings)
    
    # Build tier data
    tier_data = build_tier_data_from_source(df, block_aggregations, False)
    
    return tier_data

def write_to_excel(workbook_path, tier_data, output_path=None):
    """Write enrollment data directly to Excel cells"""
    
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"enrollment_output_{timestamp}.xlsx"
    
    print(f"\nLoading workbook: {workbook_path}")
    wb = load_workbook(workbook_path)
    
    # Track what we update
    updates = []
    errors = []
    
    # Process each facility in tier_data
    for client_id, facility_data in tier_data.items():
        if client_id not in CID_TO_TAB:
            continue
            
        tab_name = CID_TO_TAB[client_id]
        
        # Check if we have cell mappings for this tab/facility
        if tab_name not in CELL_MAPPINGS:
            print(f"⚠️ No cell mappings defined for tab: {tab_name}")
            continue
            
        if client_id not in CELL_MAPPINGS[tab_name]:
            print(f"⚠️ No cell mappings defined for {client_id} in {tab_name}")
            continue
        
        # Get the worksheet
        if tab_name not in wb.sheetnames:
            errors.append(f"Tab '{tab_name}' not found in workbook")
            continue
            
        ws = wb[tab_name]
        
        # Process each plan type (EPO, VALUE)
        for plan_type in ['EPO', 'VALUE']:
            if plan_type not in facility_data:
                continue
                
            if plan_type not in CELL_MAPPINGS[tab_name][client_id]:
                continue
            
            # Get tier counts (aggregate all blocks for this plan type)
            tier_counts = {}
            for block_label, block_tiers in facility_data[plan_type].items():
                for tier, count in block_tiers.items():
                    if tier not in tier_counts:
                        tier_counts[tier] = 0
                    tier_counts[tier] += count
            
            # Write to specific cells
            cell_map = CELL_MAPPINGS[tab_name][client_id][plan_type]
            for tier, cell_ref in cell_map.items():
                value = tier_counts.get(tier, 0)
                
                try:
                    ws[cell_ref] = value
                    updates.append({
                        'tab': tab_name,
                        'client_id': client_id,
                        'plan': plan_type,
                        'tier': tier,
                        'cell': cell_ref,
                        'value': value
                    })
                    print(f"✓ {tab_name}/{client_id} {plan_type} {tier}: {value} → {cell_ref}")
                except Exception as e:
                    errors.append(f"Error writing to {cell_ref}: {str(e)}")
    
    # Save the workbook
    print(f"\nSaving to: {output_path}")
    wb.save(output_path)
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"✓ Updates made: {len(updates)}")
    if errors:
        print(f"✗ Errors: {len(errors)}")
        for error in errors[:5]:  # Show first 5 errors
            print(f"  - {error}")
    
    # Save update log
    log_file = f"direct_write_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(log_file, 'w') as f:
        json.dump({
            'updates': updates,
            'errors': errors,
            'timestamp': datetime.now().isoformat()
        }, f, indent=2)
    print(f"\nLog saved to: {log_file}")
    
    return len(updates), len(errors)

def main():
    """Main execution"""
    print("\n" + "="*60)
    print("DIRECT EXCEL ENROLLMENT POPULATION")
    print("="*60)
    
    # Get enrollment data
    tier_data = get_enrollment_data()
    print(f"✓ Loaded data for {len(tier_data)} facilities")
    
    # Define Excel paths
    template_path = "Prime Enrollment Funding by Facility for August.xlsx"
    
    # Check if template exists
    if not os.path.exists(template_path):
        print(f"✗ Template not found: {template_path}")
        return 1
    
    # Write to Excel
    updates, errors = write_to_excel(template_path, tier_data)
    
    if errors == 0:
        print("\n✅ SUCCESS: All values written successfully!")
    else:
        print(f"\n⚠️ Completed with {errors} errors")
    
    return 0 if errors == 0 else 1

if __name__ == "__main__":
    sys.exit(main())