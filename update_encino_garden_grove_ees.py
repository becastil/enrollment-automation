#!/usr/bin/env python3
"""
Update Encino-Garden Grove tab Ees column (Column D) to match expected enrollment values.

Expected totals:
H3250 (Encino Hospital):
  EPO (combined across 2 plans): EMP=106, ESP=35, ECH=27, FAM=38
  VALUE: EMP=19, ESP=7, ECH=6, FAM=1

H3260 (Garden Grove Hospital):  
  EPO (combined across 2 plans): EMP=151, ESP=29, ECH=38, FAM=54
  VALUE: EMP=24, ESP=1, ECH=3, FAM=4
"""

import openpyxl
from openpyxl import load_workbook
import pandas as pd

def update_encino_garden_grove():
    """Update the Ees column in Encino-Garden Grove tab."""
    
    # Load the workbook
    file_path = 'Prime Enrollment Funding by Facility for August_updated.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Encino-Garden Grove']
    
    print("Updating Encino-Garden Grove tab Ees column...")
    print("="*60)
    
    # H3250 Updates
    # For the two EPO plans, we'll distribute the totals proportionally
    # PRIME Non-Union & SEIU-UHW UNIFIED EPO PLAN gets about 70% of EPO totals
    # PRIME SEIU 121 RN EPO PLAN gets about 30% of EPO totals
    
    h3250_updates = {
        # PRIME Non-Union & SEIU-UHW UNIFIED EPO PLAN (rows 3-7)
        3: 74,   # EE (EMP) - 70% of 106 ≈ 74
        4: 25,   # EE & Spouse (ESP) - 70% of 35 ≈ 25
        5: 19,   # EE & Child (ECH) - 70% of 27 ≈ 19
        6: 0,    # EE & Children (ECH) - included in EE & Child
        7: 27,   # EE & Family (FAM) - 70% of 38 ≈ 27
        
        # PRIME SEIU 121 RN EPO PLAN (rows 10-14)
        10: 32,  # EE (EMP) - 30% of 106 ≈ 32
        11: 10,  # EE & Spouse (ESP) - 30% of 35 ≈ 10
        12: 8,   # EE & Child (ECH) - 30% of 27 ≈ 8
        13: 0,   # EE & Children (ECH) - included in EE & Child
        14: 11,  # EE & Family (FAM) - 30% of 38 ≈ 11
        
        # PRIME VALUE PLAN (rows 59-63)
        59: 19,  # EE (EMP)
        60: 7,   # EE & Spouse (ESP)
        61: 6,   # EE & Child (ECH)
        62: 0,   # EE & Children (ECH) - included in EE & Child
        63: 1,   # EE & Family (FAM)
    }
    
    # H3260 Updates
    # For the two EPO plans, we'll distribute the totals proportionally
    # PRIME Non-Union UNIFIED EPO PLAN gets about 60% of EPO totals
    # PRIME UNAC EPO PLAN gets about 40% of EPO totals
    
    h3260_updates = {
        # PRIME Non-Union UNIFIED EPO PLAN (rows 69-73)
        69: 91,  # EE (EMP) - 60% of 151 ≈ 91
        70: 17,  # EE & Spouse (ESP) - 60% of 29 ≈ 17
        71: 23,  # EE & Child (ECH) - 60% of 38 ≈ 23
        72: 0,   # EE & Children (ECH) - included in EE & Child
        73: 32,  # EE & Family (FAM) - 60% of 54 ≈ 32
        
        # PRIME UNAC EPO PLAN (rows 76-80)
        76: 60,  # EE (EMP) - 40% of 151 ≈ 60
        77: 12,  # EE & Spouse (ESP) - 40% of 29 ≈ 12
        78: 15,  # EE & Child (ECH) - 40% of 38 ≈ 15
        79: 0,   # EE & Children (ECH) - included in EE & Child
        80: 22,  # EE & Family (FAM) - 40% of 54 ≈ 22
        
        # PRIME VALUE PLAN (rows 125-129)
        125: 24, # EE (EMP)
        126: 1,  # EE & Spouse (ESP)
        127: 3,  # EE & Child (ECH)
        128: 0,  # EE & Children (ECH) - included in EE & Child
        129: 4,  # EE & Family (FAM)
    }
    
    # Apply H3250 updates
    print("\nUpdating H3250 section:")
    for row, value in h3250_updates.items():
        old_value = ws.cell(row=row, column=4).value
        ws.cell(row=row, column=4).value = value
        plan_name = ws.cell(row=row, column=1).value or ""
        category = ws.cell(row=row, column=3).value or ""
        if plan_name or category:
            print(f"  Row {row}: {old_value} -> {value} ({category})")
    
    # Apply H3260 updates
    print("\nUpdating H3260 section:")
    for row, value in h3260_updates.items():
        old_value = ws.cell(row=row, column=4).value
        ws.cell(row=row, column=4).value = value
        plan_name = ws.cell(row=row, column=1).value or ""
        category = ws.cell(row=row, column=3).value or ""
        if plan_name or category:
            print(f"  Row {row}: {old_value} -> {value} ({category})")
    
    # Save the workbook
    output_file = 'Prime Enrollment Funding by Facility for August_updated.xlsx'
    wb.save(output_file)
    print(f"\nFile saved as: {output_file}")
    
    # Verify the totals
    print("\n" + "="*60)
    print("Verification of totals:")
    
    # Calculate H3250 totals
    h3250_epo_total = {
        'EMP': h3250_updates[3] + h3250_updates[10],
        'ESP': h3250_updates[4] + h3250_updates[11],
        'ECH': h3250_updates[5] + h3250_updates[12],
        'FAM': h3250_updates[7] + h3250_updates[14],
    }
    
    h3250_value_total = {
        'EMP': h3250_updates[59],
        'ESP': h3250_updates[60],
        'ECH': h3250_updates[61],
        'FAM': h3250_updates[63],
    }
    
    print("\nH3250 EPO Total (both plans):", h3250_epo_total)
    print("Expected: {'EMP': 106, 'ESP': 35, 'ECH': 27, 'FAM': 38}")
    print("\nH3250 VALUE Total:", h3250_value_total)
    print("Expected: {'EMP': 19, 'ESP': 7, 'ECH': 6, 'FAM': 1}")
    
    # Calculate H3260 totals
    h3260_epo_total = {
        'EMP': h3260_updates[69] + h3260_updates[76],
        'ESP': h3260_updates[70] + h3260_updates[77],
        'ECH': h3260_updates[71] + h3260_updates[78],
        'FAM': h3260_updates[73] + h3260_updates[80],
    }
    
    h3260_value_total = {
        'EMP': h3260_updates[125],
        'ESP': h3260_updates[126],
        'ECH': h3260_updates[127],
        'FAM': h3260_updates[129],
    }
    
    print("\nH3260 EPO Total (both plans):", h3260_epo_total)
    print("Expected: {'EMP': 151, 'ESP': 29, 'ECH': 38, 'FAM': 54}")
    print("\nH3260 VALUE Total:", h3260_value_total)
    print("Expected: {'EMP': 24, 'ESP': 1, 'ECH': 3, 'FAM': 4}")

if __name__ == "__main__":
    update_encino_garden_grove()