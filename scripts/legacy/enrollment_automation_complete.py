""" 
=================================================================================
ENROLLMENT AUTOMATION SCRIPT - Complete Version with Excel Update
=================================================================================

This version includes:
- Tier collapse bug fix with direct normalization
- Excel template update functionality
- Local file paths for testing
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import traceback
import os
warnings.filterwarnings('ignore')

# FACILITY MAPPINGS (keeping existing from original)
TPA_TO_FACILITY = {
    'H3100': 'Chino Valley Medical Center',
    'H3105': 'Glendora Community Hospital',
    'H3110': 'Prime Management',
    'H3115': 'Premiere Healthcare Staffing, LLC',
    'H3120': 'Hospital Business Service',
    'H3130': 'Bio Med Services',
    'H3140': 'Desert Valley Hospital',
    'H3150': 'Desert Valley Medical Group',
    'H3160': 'Montclair',
    'H3170': 'San Dimas',
    'H3180': 'Sherman Oaks',
    'H3190': 'Sherman Oaks Med. Group',
    'H3200': 'La Palma',
    'H3210': 'Huntington Beach',
    'H3220': 'West Anaheim',
    'H3230': 'Paradise Valley',
    'H3240': 'Paradise Valley Medical Grp',
    'H3250': 'Encino Hospital',
    'H3260': 'Garden Grove',
    'H3270': 'Centinela',
    'H3271': 'Robotics Outpatient Center',
    'H3272': 'Centinela Valley Endoscopy Center',
    'H3275': 'St. Francis Medical Center',
    'H3276': 'Shoreline Surgery Center',
    'H3277': "Physician's Surgery Center Downey",
    'H3280': 'Shasta Regional Medical Center',
    'H3285': 'Shasta Medical Group',
    'H3290': 'Hospitality',
    'H3300': "Chino RN's",
    'H3310': 'Alvarado Hospital',
    'H3320': 'Pampa',
    'H3325': 'Roxborough',
    'H3330': 'Lower Bucks',
    'H3335': 'Dallas Medical Center',
    'H3337': 'Dallas Regional Medical Center',
    'H3338': 'Riverview Regional Medical Center',
    'H3339': 'Gadsden Physicians Management',
    'H3340': 'Providence Medical Center',
    'H3345': 'St. John Hospital',
    'H3350': 'Providence Place, Inc.',
    'H3355': 'Knapp Medical Center',
    'H3360': 'Knapp Medical Group',
    'H3365': 'Knapp Ambulatory Surgery Center',
    'H3370': 'Harlingen Medical Center',
    'H3375': 'Garden City Hospital',
    'H3380': 'United Home Health Services',
    'H3381': 'Lake Huron Medical Center',
    'H3382': 'Lake Huron Medical Group',
    'H3385': 'Prime Garden City Medical Group',
    'H3390': 'Rehabilitation Hospital of Rhode Island',
    'H3392': 'Landmark Medical Center',
    'H3394': "Summit Surgery Center a St. Mary's Galena",
    'H3395': "St. Mary's Regional Medical Center",
    'H3396': "St. Mary's Medical Group",
    'H3397': 'Monroe Hospital',
    'H3398': 'North Vista Hospital',
    'H3399': 'North Vista Medical Group',
    'H3400': "St. Mary's Fitness Center",
    'H3500': "St Clare's Health System",
    'H3505': "Saint Mary's General Hospital",
    'H3510': 'Southern Medical Regional Center',
    'H3520': 'Lehigh Regional Medical Center',
    'H3530': "St. Michael's Medical Center",
    'H3540': 'Mission Regional Medical Center',
    'H3560': "St. Mary's Medical Center",
    'H3561': 'St. Joseph Medical Center',
    'H3562': 'South Kansas City Surgi Center',
    'H3563': 'CHCS Home Health Care',
    'H3564': 'CPCN Physicians Service',
    'H3565': 'CPCN Physicians Service (32) STJ',
    'H3566': "St. Mary's Surgical Center",
    'H3591': 'Coshocton County Memorial Hospital',
    'H3592': 'East Liverpool City Hospital',
    'H3593': 'Ohio Valley Home Health Care',
    'H3594': 'Ohio Valley Home Health Services',
    'H3595': 'River Valley Physicians',
    'H3598': 'Suburban Community Hospital',
    'H3599': 'Suburban Medical Group',
    'H3600': 'Reddy Development LLC',
    'H3605': 'Mercy Medical Center - Aurora LLC',
    'H3615': 'Resurrection Medical Center - Chicago LLC',
    'H3625': 'Saint Francis Hospital - Evanston LLC',
    'H3630': 'Saint Joseph Hospital - Elgin LLC',
    'H3635': 'Saint Joseph Hospital - Joliet LLC',
    'H3645': "St. Mary's Hospital - Kankakee, LLC",
    'H3655': 'Saint Mary of Nazareth Hospital - Chicago, LLC',
    'H3660': 'Holy Family Medical Center - Des Plaines LLC',
    'H3665': 'MedSpace Services, LLC',
    'H3670': 'Prime Healthcare Illinois Medical Group, LLC',
    'H3675': 'Prime Healthcare Home Care and Hospice',
    'H3680': 'Prime Healthcare Senior Living'
}

# FACILITY MAPPING BY TABS (from original)
FACILITY_MAPPING = {
    'Legacy': {
        'H3170': 'San Dimas Community Hospital',
        'H3130': 'Bio-Medical Services',
        'H3100': 'Chino Valley Medical Center',
        'H3300': 'Chino Valley Medical Center RNs',
        'H3140': 'Desert Valley Hospital',
        'H3150': 'Desert Valley Medical Group',
        'H3210': 'Huntington Beach Hospital',
        'H3200': 'La Palma Intercommunity Hospital',
        'H3160': 'Montclair Hospital Medical Center',
        'H3115': 'Premiere Healthcare Staffing',
        'H3110': 'Prime Management Services',
        'H3230': 'Paradise Valley Hospital',
        'H3240': 'Paradise Valley Medical Group',
        'H3180': 'Sherman Oaks Hospital',
        'H3190': 'Sherman Oaks Medical Group',
        'H3220': 'West Anaheim Medical Center',
        'H3285': 'Shasta Medical Group',
        'H3290': 'Hospitality'
    },
    'Centinela': {
        'H3270': 'Centinela Hospital Medical Center',
        'H3271': 'Robotics Outpatient Center',
        'H3272': 'Centinela Valley Endoscopy Center'
    },
    'Encino-Garden Grove': {
        'H3250': 'Encino Hospital Medical Center',
        'H3260': 'Garden Grove Hospital Medical Center'
    },
    'St. Francis': {
        'H3275': 'St. Francis Medical Center',
        'H3276': 'Shoreline Surgery Center',
        'H3277': "Physician's Surgery Center Downey"
    },
    'Pampa': {
        'H3320': 'Pampa Regional Medical Center'
    },
    'Roxborough': {
        'H3325': 'Roxborough Memorial Hospital'
    },
    'Lower Bucks': {
        'H3330': 'Lower Bucks Hospital'
    },
    'Dallas Medical Center': {
        'H3335': 'Dallas Medical Center'
    },
    'Harlingen': {
        'H3370': 'Harlingen Medical Center'
    },
    'Knapp': {
        'H3355': 'Knapp Medical Center',
        'H3360': 'Knapp Medical Group'
    },
    'Monroe': {
        'H3397': 'Monroe Hospital'
    },
    "Saint Mary's Reno": {
        'H3394': "Summit Surgery Center at St. Mary's Galena",
        'H3395': "Saint Mary's Regional Medical Center",
        'H3396': "Saint Mary's Medical Group",
        'H3400': "Saint Mary's Fitness Center"
    },
    'North Vista': {
        'H3398': 'North Vista Hospital',
        'H3399': 'North Vista Medical Group'
    },
    'Dallas Regional': {
        'H3337': 'Dallas Regional Medical Center'
    },
    'Riverview & Gadsden': {
        'H3338': 'Riverview Regional Medical Center',
        'H3339': 'Gadsden Physicians Management'
    },
    "Saint Clare's": {
        'H3500': "Saint Clare's Health System"
    },
    'Landmark': {
        'H3392': 'Landmark Medical Center'
    },
    "Saint Mary's Passaic": {
        'H3505': "Saint Mary's General Hospital - Passaic, NJ"
    },
    'Southern Regional': {
        'H3510': 'Southern Regional Medical Center'
    },
    "St Michael's": {
        'H3530': "St. Michael's Medical Center"
    },
    'Mission': {
        'H3540': 'Mission Regional Medical Center'
    },
    'Coshocton': {
        'H3591': 'Coshocton County Memorial Hospital'
    },
    'Suburban': {
        'H3598': 'Suburban Community Hospital',
        'H3599': 'Suburban Medical Group'
    },
    'Garden City': {
        'H3375': 'Garden City Hospital',
        'H3385': 'Garden City Medical Group',
        'H3380': 'United Home Health Services'
    },
    'Lake Huron': {
        'H3381': 'Lake Huron Medical Center',
        'H3382': 'Lake Huron Medical Group'
    },
    'Providence & St John': {
        'H3340': 'Providence Medical Center',
        'H3345': 'St. John Hospital'
    },
    'East Liverpool': {
        'H3592': 'East Liverpool City Hospital',
        'H3594': 'Ohio Valley Home Health Services',
        'H3595': 'River Valley Physicians'
    },
    'St Joe & St Mary\'s': {
        'H3560': "St. Mary's Medical Center",
        'H3561': 'St. Joseph Medical Center',
        'H3562': 'South Kansas City Surgical Center',
        'H3563': 'CHCS Home Health Care',
        'H3564': 'CPCN Physicians Service (30) STM',
        'H3565': 'CPCN Physicians Service (32) STJ',
        'H3566': "St. Mary's Surgical Center"
    },
    'Illinois': {
        'H3605': 'Mercy Medical Center - Aurora LLC',
        'H3615': 'Resurrection Medical Center - Chicago LLC',
        'H3625': 'Saint Francis Hospital - Evanston LLC',
        'H3630': 'Saint Joseph Hospital - Elgin LLC',
        'H3635': 'Saint Joseph Hospital - Joliet LLC',
        'H3645': "St. Mary's Hospital - Kankakee, LLC",
        'H3655': 'Saint Mary of Nazareth Hospital - Chicago, LLC',
        'H3660': 'Holy Family Medical Center',
        'H3665': 'MedSpace Services, LLC',
        'H3670': 'Prime Healthcare Illinois Medical Group, LLC',
        'H3675': 'Prime Healthcare Home Care and Hospice',
        'H3680': 'Prime Healthcare Senior Living'
    },
    'Alvarado': {
        'H3310': 'Alvarado Hospital'
    }
}

# PLAN TYPE MAPPING (keeping existing)
PLAN_TO_TYPE = {
    'PRIMESFSE': 'EPO',
    'PRIMESFUN': 'EPO',
    'PRIMEMSFE': 'EPO',
    'PRIMEMMLMRI': 'VALUE',
    'PRIMEMMCV': 'EPO',
    'PRIMEMMGLN': 'EPO',
    'PRIMEMMEPOLE2': 'EPO',
    'PRIMEMMWA': 'EPO',
    'PRIMEMMCC': 'EPO',
    'PRIMEMMDAL': 'EPO',
    'PRIMEMMEL': 'EPO',
    'PRIMEMMELPOS': 'PPO',
    'PRIMEMMEPOCEN': 'EPO',
    'PRIMEMMEPOEA': 'EPO',
    'PRIMEMMEPOES': 'EPO',
    'PRIMEMMEPOLE': 'EPO',
    'PRIMEMMEPOPA': 'EPO',
    'PRIMEMMEPOROX': 'EPO',
    'PRIMEMMEPOUEE': 'EPO',
    'PRIMEMMGCH': 'EPO',
    'PRIMEMMHAR': 'EPO',
    'PRIMEMMKN': 'EPO',
    'PRIMEMMKNVAL': 'VALUE',
    'PRIMEMMKS': 'EPO',
    'PRIMEMMLB': 'EPO',
    'PRIMEMMLKEP1': 'EPO',
    'PRIMEMMLKEP2': 'EPO',
    'PRIMEMMLM': 'EPO',
    'PRIMEMMLR': 'EPO',
    'PRIMEMMMH': 'EPO',
    'PRIMEMMMR': 'EPO',
    'PRIMEMMNV': 'EPO',
    'PRIMEMMPLT': 'EPO',
    'PRIMEMMPMC': 'EPO',
    'PRIMEMMPPOEN': 'PPO',
    'PRIMEMMPPOLAP': 'PPO',
    'PRIMEMMPPOLHCEN': 'PPO',
    'PRIMEMMPPOLLCEN': 'PPO',
    'PRIMEMMPPOUH': 'PPO',
    'PRIMEMMPPOLH': 'PPO',
    'PRIMEMMPPOUL': 'PPO',
    'PRIMEMMRHRI': 'EPO',
    'PRIMEMMRV': 'EPO',
    'PRIMEMMSB': 'EPO',
    'PRIMEMMSBPLT': 'EPO',
    'PRIMEMMSJH': 'EPO',
    'PRIMEMMSM': 'EPO',
    'PRIMEMMSMMSMRMC': 'EPO',
    'PRIMEMMSMMSMRMCP': 'PPO',
    'PRIMEMMSR': 'EPO',
    'PRIMEMMSTCL': 'EPO',
    'PRIMEMMST': 'EPO',
    'PRIMEMMSTPPO': 'PPO',
    'PRIMEMMVAL': 'VALUE',
    'PRIMEMMSMECN': 'EPO',
    'PRIMEMMSMECW': 'EPO',
    'PRIMEMMCIR': 'EPO',
    'PRIMEMMIUOE': 'EPO',
    'PRIMEMMJNESO': 'EPO',
    'PRIMELBH': 'EPO',
    'PRIMEMMEPO3': 'EPO',
    'PRIMEMMEPOESU': 'EPO',
    'PRIMEMMEPOLEUN': 'EPO',
    'PRIMEMMEPPLUS': 'EPO',
    'PRIMEMMVALUE': 'VALUE',
    'PRIMEPOPRE21': 'EPO',
    'PRIMESFMCVAL': 'VALUE',
    'PRIMEASCIL': 'EPO',
    'PRIMEASCILEPO': 'EPO',
    'PRIMEMMIL': 'VALUE',
    'PRIMEMMILVALUE': 'VALUE',
    'PRIMEINAIL': 'EPO',
    'PRIMEINAILVALUE': 'VALUE',
    'PRIMEMMSMEPLUS': 'EPO',
    'PRIMEMMSMUN': 'EPO'
}

def normalize_tier(raw_tier):
    """
    Normalize raw tier text to standard format for Excel template
    Returns one of: 'EE', 'EE & Spouse', 'EE & Child', 'EE & Children', 'EE & Family'
    """
    if pd.isna(raw_tier):
        return 'EE'  # Default to EE for template compatibility
    
    # Clean the input
    tier_str = str(raw_tier).strip().upper()
    # Normalize separators
    tier_str = tier_str.replace('&', ' AND ').replace('+', ' AND ').replace('/', ' AND ')
    # Remove extra spaces
    tier_str = ' '.join(tier_str.split())
    
    # Employee Only variants
    if tier_str in ['EMP', 'EE', 'EMPLOYEE ONLY', 'EE ONLY', 'EMPLOYEE', 'E']:
        return 'EE'
    
    # Employee + Spouse variants
    if tier_str in ['ESP', 'EE AND SPOUSE', 'EMPLOYEE AND SPOUSE', 'EE SPOUSE', 
                    'EMPLOYEE SPOUSE', 'ES', 'E AND S']:
        return 'EE & Spouse'
    
    # Employee + 1 Child variants
    if tier_str in ['E1D', 'EE AND CHILD', 'EMPLOYEE AND CHILD', 'EE CHILD',
                    'EE AND 1 CHILD', 'EE AND 1 DEPENDENT']:
        return 'EE & Child'
    
    # Employee + Children variants
    if tier_str in ['ECH', 'EE AND CHILDREN', 'EMPLOYEE AND CHILDREN', 'EE CHILDREN',
                    'EC', 'E AND C']:
        return 'EE & Children'
    
    # Employee + Family variants
    if tier_str in ['FAM', 'FAMILY', 'EE AND FAMILY', 'EMPLOYEE AND FAMILY',
                    'EE FAMILY', 'EF', 'E AND F']:
        return 'EE & Family'
    
    # Default to EE if unknown
    return 'EE'

def infer_plan_type(code):
    """
    Infer plan type from code if not in mapping
    """
    s = str(code).upper() if pd.notna(code) else ''
    if 'PPO' in s: 
        return 'PPO'
    if 'EPO' in s: 
        return 'EPO'
    if 'VAL' in s or 'VALUE' in s: 
        return 'VALUE'
    return 'VALUE'  # Default

def read_and_prepare_data(file_path):
    """
    Read source data and prepare for processing
    """
    df = pd.read_excel(file_path, sheet_name=0)
    
    # Filter to active only
    if 'STATUS' in df.columns:
        original_count = len(df)
        df = df[df['STATUS'].astype(str).str.upper() == 'A'].copy()
        print(f"Filtered to {len(df)} active rows (STATUS='A') from {original_count} total")
    
    # Add facility info
    if 'CLIENT ID' in df.columns:
        df['facility_id'] = df['CLIENT ID']
        df['facility_name'] = df['facility_id'].map(TPA_TO_FACILITY)
        print(f"Mapped {df['facility_name'].notna().sum()} facilities")
    
    return df

def process_enrollment_data_fixed(df):
    """
    Process enrollment data with fixed tier normalization
    Returns data structured for Excel template update
    """
    processed_data = {}
    
    # Filter to subscribers only
    if 'RELATION' in df.columns:
        subscribers = df[df['RELATION'].str.upper() == 'SELF'].copy()
        print(f"Filtered to {len(subscribers)} subscribers (RELATION='SELF')")
    else:
        subscribers = df.copy()
    
    # Normalize tiers directly from BEN CODE
    if 'BEN CODE' in subscribers.columns:
        print("Using BEN CODE column for tier information")
        subscribers['tier'] = subscribers['BEN CODE'].apply(normalize_tier)
    else:
        print("Warning: No BEN CODE column found, defaulting to EE")
        subscribers['tier'] = 'EE'
    
    # Map plan types
    if 'PLAN' in subscribers.columns:
        subscribers['plan_type'] = subscribers['PLAN'].map(PLAN_TO_TYPE).fillna(
            subscribers['PLAN'].apply(infer_plan_type)
        )
    else:
        subscribers['plan_type'] = 'VALUE'
    
    # Show tier distribution
    tier_dist = subscribers['tier'].value_counts()
    print(f"\nTier Distribution:\n{tier_dist}")
    
    # Process each tab and facility
    for tab_name, facilities in FACILITY_MAPPING.items():
        processed_data[tab_name] = {}
        
        for client_id, facility_name in facilities.items():
            # Find data for this facility
            if 'CLIENT ID' in subscribers.columns:
                facility_data = subscribers[subscribers['CLIENT ID'] == client_id].copy()
            else:
                facility_data = pd.DataFrame()  # Empty if no CLIENT ID
            
            if facility_data.empty:
                # Default to zeros
                processed_data[tab_name][facility_name] = {
                    plan: {tier: 0 for tier in ['EE', 'EE & Spouse', 'EE & Child', 'EE & Children', 'EE & Family']}
                    for plan in ['EPO', 'PPO', 'VALUE']
                }
                continue
            
            # Count enrollments by plan type and tier
            enrollment_counts = (facility_data
                .groupby(['plan_type', 'tier'])
                .size()
                .unstack(fill_value=0)
                .reindex(columns=['EE', 'EE & Spouse', 'EE & Child', 'EE & Children', 'EE & Family'], fill_value=0)
                .to_dict('index')
            )
            
            # Structure the result
            result = {}
            for plan in ['EPO', 'PPO', 'VALUE']:
                if plan in enrollment_counts:
                    result[plan] = enrollment_counts[plan]
                else:
                    result[plan] = {tier: 0 for tier in ['EE', 'EE & Spouse', 'EE & Child', 'EE & Children', 'EE & Family']}
            
            processed_data[tab_name][facility_name] = result
    
    return processed_data

def find_facility_location(ws, facility_name, start_row=1, max_row=1000):
    """
    Find where a facility's data begins in the template
    """
    for row in range(start_row, min(max_row, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and facility_name in str(cell_value):
                return row, col
    return None, None

def find_section_start(ws, anchor_row, keywords=('EPO',)):
    """
    Find where a section (EPO, PPO, VALUE) starts
    """
    max_r = min(ws.max_row, anchor_row + 25)
    max_c = min(ws.max_column, 10)
    
    for r in range(anchor_row, max_r + 1):
        for c in range(1, max_c + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and any(k in val.upper() for k in keywords):
                return r
    return None

def update_plan_section_by_position(ws, start_row, col, tier_data):
    """
    Fill in enrollment numbers in the template
    """
    # Map tier names to their row positions
    tier_rows = {
        'EE': 0,
        'EE & Spouse': 1,
        'EE & Child': 2,
        'EE & Children': 3,
        'EE & Family': 4
    }
    
    # Check if template uses combined Child/Children format
    tier_label_col = col - 1
    row2_label = ws.cell(row=start_row + 2, column=tier_label_col).value
    
    if row2_label and 'Child(ren)' in str(row2_label):
        # Combined format
        tier_rows = {
            'EE': 0,
            'EE & Spouse': 1,
            'EE & Child': 2,
            'EE & Children': 2,
            'EE & Family': 3
        }
    
    for tier, row_offset in tier_rows.items():
        if tier in tier_data:
            current_value = ws.cell(row=start_row + row_offset, column=col).value or 0
            if tier == 'EE & Children' and row_offset == tier_rows.get('EE & Child', -1):
                ws.cell(row=start_row + row_offset, column=col).value = current_value + tier_data[tier]
            else:
                ws.cell(row=start_row + row_offset, column=col).value = tier_data[tier]

def update_destination_file(destination_path, processed_data, output_path=None):
    """
    Update the Excel template with enrollment counts
    """
    wb = load_workbook(destination_path)
    
    for tab_name, facilities_data in processed_data.items():
        if tab_name not in wb.sheetnames:
            print(f"Warning: Tab '{tab_name}' not found in destination file")
            continue
            
        ws = wb[tab_name]
        
        for facility_name, plan_data in facilities_data.items():
            # Find where this facility's section starts
            facility_row, facility_col = find_facility_location(ws, facility_name)
            
            if not facility_row:
                print(f"Warning: Could not find '{facility_name}' in tab '{tab_name}'")
                continue
            
            # From facility location: 3 columns over, 1 row down is where numbers go
            enrollment_col = facility_col + 3
            
            print(f"  Found '{facility_name}' at {get_column_letter(facility_col)}{facility_row}")
            
            # Update EPO section
            epo_row = find_section_start(ws, facility_row, ('EPO',))
            if epo_row and 'EPO' in plan_data:
                update_plan_section_by_position(ws, epo_row, enrollment_col, plan_data['EPO'])
            
            # Update PPO section if exists
            ppo_row = find_section_start(ws, facility_row, ('PPO',))
            if ppo_row and 'PPO' in plan_data:
                update_plan_section_by_position(ws, ppo_row, enrollment_col, plan_data['PPO'])
            
            # Update VALUE section
            value_row = find_section_start(ws, facility_row, ('VALUE',))
            if value_row and 'VALUE' in plan_data:
                update_plan_section_by_position(ws, value_row, enrollment_col, plan_data['VALUE'])
    
    # Save the updated workbook
    if output_path:
        wb.save(output_path)
        print(f"✓ Excel file updated: {output_path}")
    else:
        output_path = destination_path.replace('.xlsx', '_updated.xlsx')
        wb.save(output_path)
        print(f"✓ Excel file saved as: {output_path}")

def main():
    """
    Main execution function
    """
    # FILE PATHS - UPDATED FOR LOCAL TESTING
    source_file = r"C:\Users\becas\Prime_EFR\data\input\source_data.xlsx"
    destination_file = r"C:\Users\becas\Prime_EFR\data\input\Prime Enrollment Funding by Facility for August.xlsx"
    output_file = r"C:\Users\becas\Prime_EFR\data\input\Prime Enrollment Funding by Facility for August.xlsx"  # Same as destination
    summary_csv = r"C:\Users\becas\Prime_EFR\output\enrollment_summary.csv"
    
    try:
        print("="*60)
        print("ENROLLMENT AUTOMATION - COMPLETE VERSION")
        print("="*60)
        print(f"Source: {source_file}")
        print(f"Template: {destination_file}")
        print(f"Output: {output_file}")
        
        # Step 1: Read source data
        print("\n1. Reading source data...")
        df = read_and_prepare_data(source_file)
        
        # Step 2: Process enrollment data with fix
        print("\n2. Processing enrollment data...")
        processed_data = process_enrollment_data_fixed(df)
        
        # Step 3: Create summary CSV
        print("\n3. Creating summary report...")
        summary_rows = []
        for tab, facilities in processed_data.items():
            for facility, plans in facilities.items():
                for plan_type, tiers in plans.items():
                    for tier, count in tiers.items():
                        if count > 0:
                            summary_rows.append({
                                'Tab': tab,
                                'Facility': facility,
                                'Plan Type': plan_type,
                                'Tier': tier,
                                'Count': count
                            })
        
        if summary_rows:
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_csv(summary_csv, index=False)
            print(f"✓ Summary saved to: {summary_csv}")
        
        # Step 4: Update Excel template
        print("\n4. Updating Excel template...")
        update_destination_file(destination_file, processed_data, output_file)
        
        # Step 5: Final summary
        print("\n" + "="*60)
        print("✓ PROCESSING COMPLETE!")
        print("="*60)
        print(f"✓ Source data processed: {source_file}")
        print(f"✓ Excel updated: {output_file}")
        print(f"✓ Summary CSV: {summary_csv}")
        print(f"✓ Total tabs processed: {len(processed_data)}")
        
        # Show sample results
        total_enrollments = sum(
            sum(sum(tiers.values()) for tiers in plans.values())
            for facilities in processed_data.values()
            for plans in facilities.values()
        )
        print(f"✓ Total enrollments processed: {total_enrollments}")
        
    except FileNotFoundError as e:
        print(f"\n❌ ERROR: File not found - {e}")
        print("\nPlease ensure these files exist:")
        print(f"  - Source: {source_file}")
        print(f"  - Template: {destination_file}")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()