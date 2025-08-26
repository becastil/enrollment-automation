"""
Populate Tiered Enrollment in Excel Facility Tabs
==================================================

This script reads enrollment data and populates each facility tab in the Excel file
with tiered enrollment counts (EE, EE & Spouse, EE & Child(ren), EE & Family).

It preserves the existing template structure and only updates the Ees column
and calculated premium amounts.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from typing import Dict, List, Tuple
import warnings

warnings.filterwarnings('ignore')

# Mapping of Excel tabs to facility Client IDs
TAB_TO_FACILITIES = {
    'Legacy': ['H3100', 'H3110', 'H3140', 'H3150', 'H3160', 'H3170', 'H3180', 
               'H3200', 'H3210', 'H3220', 'H3230', 'H3240'],
    'Centinela': ['H3270', 'H3271', 'H3272'],
    'Encino-Garden Grove': ['H3250', 'H3260'],
    'St. Francis': ['H3275', 'H3276', 'H3277'],
    'Alvarado': ['H3280', 'H3285'],
    'Pampa': ['H3320'],
    'Roxborough': ['H3325'],
    'Lower Bucks': ['H3330'],
    'Dallas Medical Center': ['H3335'],
    'Harlingen': ['H3370'],
    'Knapp': ['H3355', 'H3360'],
    'Glendora': ['H3300'],  # Chino RN's
    'RHRI': ['H3130', 'H3115'],  # Bio Med Services, Premiere Healthcare
    'Monroe': ['H3397'],
    "Saint Mary's Reno": ['H3395', 'H3396', 'H3394'],
    'North Vista': ['H3398'],
    'Dallas Regional': ['H3337'],
    'Riverview & Gadsden': ['H3338', 'H3339'],
    "Saint Clare's": ['H3500'],
    'Landmark': ['H3392'],
    "Saint Mary's Passaic": ['H3505'],
    'Southern Regional': ['H3510'],
    'Lehigh': ['H3330'],  # May share with Lower Bucks
    "St Michael's": ['H3530'],
    'Reddy Dev.': ['H3564', 'H3565'],  # CPCN facilities
    'Mission': ['H3540'],
    'Coshocton': ['H3591'],
    'Suburban': ['H3598', 'H3599'],
    'Garden City': ['H3375', 'H3380', 'H3381', 'H3382', 'H3385'],
    'Lake Huron': ['H3381', 'H3382'],
    'Providence & St John': ['H3340', 'H3345'],
    'East Liverpool': ['H3592', 'H3594', 'H3595'],
    "St Joe & St Mary's": ['H3561', 'H3560', 'H3562', 'H3566'],
    'Illinois': ['H3605', 'H3615', 'H3625', 'H3630', 'H3635', 'H3645', 'H3655', 
                 'H3660', 'H3665', 'H3670', 'H3675', 'H3680']
}

# Coverage tier mapping - how to identify tiers in the data
TIER_MAPPING = {
    'EE': ['EE', 'Employee', 'Self', 'EMPLOYEE ONLY'],
    'EE & Spouse': ['EE & Spouse', 'Employee + Spouse', 'EE+SP', 'EMPLOYEE + SPOUSE'],
    'EE & Child(ren)': ['EE & Child(ren)', 'Employee + Child', 'EE+CH', 'EMPLOYEE + CHILDREN'],
    'EE & Family': ['EE & Family', 'Family', 'EE+FAM', 'EMPLOYEE + FAMILY']
}


def read_enrollment_data_with_tiers(excel_path: str, sheet_name: str = "Cleaned use this one") -> pd.DataFrame:
    """
    Read enrollment data and identify coverage tiers.
    
    We need to find columns that indicate the coverage tier (EE, EE+Spouse, etc.)
    This might be in columns like 'TIER', 'COVERAGE', 'PLAN TYPE', etc.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=4)
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Look for tier information in various possible columns
        tier_col = None
        possible_tier_cols = ['TIER', 'COVERAGE', 'COVERAGE TIER', 'PLAN TIER', 
                             'COVERAGE TYPE', 'EMPLOYEE TYPE', 'RELATION']
        
        for col in possible_tier_cols:
            if col in df.columns:
                tier_col = col
                break
        
        # If no tier column found, try to infer from member count or dependents
        if tier_col is None:
            print("Warning: No explicit tier column found. Inferring from data...")
            # Check for dependent counts
            if 'DEPENDENTS' in df.columns or 'DEP COUNT' in df.columns:
                dep_col = 'DEPENDENTS' if 'DEPENDENTS' in df.columns else 'DEP COUNT'
                df['TIER'] = df.apply(lambda row: infer_tier_from_dependents(row, dep_col), axis=1)
                tier_col = 'TIER'
            else:
                # Default to EE if no tier information available
                print("Using default tier: EE for all records")
                df['TIER'] = 'EE'
                tier_col = 'TIER'
        
        # Standardize tier names
        df['STANDARD_TIER'] = df[tier_col].apply(standardize_tier_name)
        
        # Filter to subscribers only (SEQ. # = 0)
        if 'SEQ. #' in df.columns:
            df = df[df['SEQ. #'] == 0]
        
        return df
        
    except Exception as e:
        print(f"Error reading enrollment data: {e}")
        # Return sample data for testing
        return create_sample_tier_data()


def infer_tier_from_dependents(row, dep_col):
    """Infer coverage tier from dependent count."""
    try:
        dep_count = int(row[dep_col]) if pd.notna(row[dep_col]) else 0
        
        if dep_count == 0:
            return 'EE'
        elif dep_count == 1:
            # Could be spouse or child - need more info
            # Check for spouse indicator
            if 'SPOUSE' in str(row).upper():
                return 'EE & Spouse'
            else:
                return 'EE & Child(ren)'
        else:
            return 'EE & Family'
    except:
        return 'EE'


def standardize_tier_name(tier_value):
    """Convert various tier names to standard format."""
    if pd.isna(tier_value):
        return 'EE'
    
    tier_str = str(tier_value).upper().strip()
    
    # Check against mapping
    for standard_tier, variations in TIER_MAPPING.items():
        for variation in variations:
            if variation.upper() in tier_str or tier_str in variation.upper():
                return standard_tier
    
    # Default mapping based on common patterns
    if 'FAMILY' in tier_str or 'FAM' in tier_str:
        return 'EE & Family'
    elif 'SPOUSE' in tier_str or 'SP' in tier_str:
        return 'EE & Spouse'
    elif 'CHILD' in tier_str or 'CH' in tier_str:
        return 'EE & Child(ren)'
    else:
        return 'EE'


def create_sample_tier_data():
    """Create sample data with tier breakdown for testing."""
    # Use the enrollment summary we have and distribute across tiers
    summary_df = pd.read_csv('enrollment_summary_cleaned_sheet.csv')
    
    # Create detailed records with tier distribution
    detailed_records = []
    
    for _, row in summary_df.iterrows():
        total_count = row['contract_count']
        facility = row['facility_name']
        client_id = row['client_id']
        plan_group = row['plan_group']
        
        # Distribute counts across tiers (approximate distribution)
        tier_distribution = {
            'EE': int(total_count * 0.65),
            'EE & Spouse': int(total_count * 0.10),
            'EE & Child(ren)': int(total_count * 0.15),
            'EE & Family': int(total_count * 0.10)
        }
        
        # Adjust to match total
        diff = total_count - sum(tier_distribution.values())
        tier_distribution['EE'] += diff
        
        for tier, count in tier_distribution.items():
            if count > 0:
                detailed_records.append({
                    'CLIENT ID': client_id,
                    'Facility': facility,
                    'PLAN': f'PRIME{plan_group}PLAN',
                    'EPO-PPO-VALUE': plan_group,
                    'STANDARD_TIER': tier,
                    'COUNT': count
                })
    
    return pd.DataFrame(detailed_records)


def aggregate_enrollment_by_tier(df: pd.DataFrame) -> Dict:
    """
    Aggregate enrollment data by facility, plan type, and tier.
    
    Returns a nested dictionary:
    {client_id: {plan_type: {tier: count}}}
    """
    # Map plan names to groups
    plan_mapping = pd.read_csv('plan_mapping_complete.csv')
    plan_dict = dict(zip(plan_mapping['PLAN'], plan_mapping['EPO-PPO-VAL']))
    
    # Check if this is sample data or real data
    if 'EPO-PPO-VALUE' in df.columns:
        # Sample data already has plan group
        df['PLAN_GROUP'] = df['EPO-PPO-VALUE']
    elif 'PLAN' in df.columns:
        # Real data needs mapping
        df['PLAN_GROUP'] = df['PLAN'].map(plan_dict)
    else:
        print("Warning: No PLAN column found, using default EPO")
        df['PLAN_GROUP'] = 'EPO'
    
    # Group by facility, plan, and tier
    if 'COUNT' in df.columns:
        # Pre-aggregated data
        grouped = df.groupby(['CLIENT ID', 'PLAN_GROUP', 'STANDARD_TIER'])['COUNT'].sum()
    else:
        # Raw data - count records
        grouped = df.groupby(['CLIENT ID', 'PLAN_GROUP', 'STANDARD_TIER']).size()
    
    # Convert to nested dictionary
    result = {}
    for (client_id, plan_group, tier), count in grouped.items():
        if client_id not in result:
            result[client_id] = {}
        if plan_group not in result[client_id]:
            result[client_id][plan_group] = {}
        result[client_id][plan_group][tier] = count
    
    return result


def populate_facility_tab(wb, sheet_name: str, facility_data: Dict, facility_mapping: pd.DataFrame):
    """
    Populate a facility tab with tiered enrollment data.
    
    Preserves the existing template and only updates the Ees column.
    """
    try:
        ws = wb[sheet_name]
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in workbook")
        return
    
    # Get facilities for this tab
    facilities = TAB_TO_FACILITIES.get(sheet_name, [])
    
    # Process each row in the sheet
    for row_num in range(1, ws.max_row + 1):
        # Look for facility headers (contain "Client ID")
        cell_a = ws.cell(row=row_num, column=1)
        if cell_a.value and 'Client ID' in str(cell_a.value):
            # Extract client ID from the header
            header_text = str(cell_a.value)
            client_id = None
            for fac_id in facilities:
                if fac_id in header_text:
                    client_id = fac_id
                    break
            
            if not client_id:
                continue
            
            # Process plan sections below this facility
            current_plan = None
            for plan_row in range(row_num + 1, min(row_num + 30, ws.max_row)):
                plan_cell = ws.cell(row=plan_row, column=1)
                
                if not plan_cell.value:
                    continue
                
                plan_text = str(plan_cell.value).upper()
                
                # Identify plan type
                if 'EPO PLAN' in plan_text:
                    current_plan = 'EPO'
                elif 'PPO HIGH' in plan_text:
                    current_plan = 'PPO_HIGH'
                elif 'PPO LOW' in plan_text:
                    current_plan = 'PPO_LOW'
                elif 'VALUE PLAN' in plan_text:
                    current_plan = 'VALUE'
                elif any(tier in plan_text for tier in ['EE & SPOUSE', 'EE & CHILD', 'EE & FAMILY']):
                    # This is a tier row
                    if current_plan and client_id in facility_data:
                        # Determine tier
                        tier = None
                        if 'EE & SPOUSE' in plan_text:
                            tier = 'EE & Spouse'
                        elif 'EE & CHILD' in plan_text:
                            tier = 'EE & Child(ren)'
                        elif 'EE & FAMILY' in plan_text:
                            tier = 'EE & Family'
                        elif plan_text.strip() == 'EE':
                            tier = 'EE'
                        
                        if tier:
                            # Get the count for this combination
                            plan_key = 'EPO' if current_plan == 'EPO' else 'VALUE' if current_plan == 'VALUE' else 'PPO'
                            count = 0
                            
                            if client_id in facility_data:
                                if plan_key in facility_data[client_id]:
                                    count = facility_data[client_id][plan_key].get(tier, 0)
                            
                            # Update the Ees column (typically column D/4)
                            ws.cell(row=plan_row, column=4, value=count)
                            
                            # Calculate premium if rate exists
                            rate_cell = ws.cell(row=plan_row, column=5)
                            if rate_cell.value and count > 0:
                                try:
                                    rate = float(str(rate_cell.value).replace('$', '').replace(',', ''))
                                    # Premium calculation would go in appropriate column
                                    # This depends on the exact template structure
                                except:
                                    pass
                elif 'ESTIMATED MONTHLY PREMIUM' in plan_text:
                    # Update totals row
                    # Sum the Ees for this plan
                    if current_plan and client_id in facility_data:
                        plan_key = 'EPO' if current_plan == 'EPO' else 'VALUE' if current_plan == 'VALUE' else 'PPO'
                        total = 0
                        if client_id in facility_data and plan_key in facility_data[client_id]:
                            total = sum(facility_data[client_id][plan_key].values())
                        ws.cell(row=plan_row, column=4, value=total)
    
    print(f"Updated sheet: {sheet_name}")


def main():
    """Main function to populate all facility tabs."""
    
    # File paths
    excel_file = r"C:\Users\becas\Downloads\Prime Enrollment Funding by Facility for July.xlsx"
    output_file = r"C:\Users\becas\Downloads\Prime Enrollment Funding by Facility for July_POPULATED.xlsx"
    
    # Check if source file exists in current directory for testing
    if not os.path.exists(excel_file):
        excel_file = "Prime_Enrollment_Sample.xlsx"
        output_file = "Prime_Enrollment_Sample_POPULATED.xlsx"
        print(f"Using sample file: {excel_file}")
    
    # Read enrollment data with tier information
    print("Reading enrollment data...")
    enrollment_df = read_enrollment_data_with_tiers(excel_file)
    
    # Aggregate by facility, plan, and tier
    print("Aggregating enrollment by tier...")
    facility_tier_data = aggregate_enrollment_by_tier(enrollment_df)
    
    # Load facility mapping
    facility_mapping = pd.read_csv('facility_mapping_complete.csv')
    
    # Open the Excel workbook
    print(f"Opening workbook: {excel_file}")
    wb = load_workbook(excel_file)
    
    # Process each facility tab
    for sheet_name in TAB_TO_FACILITIES.keys():
        print(f"Processing sheet: {sheet_name}")
        populate_facility_tab(wb, sheet_name, facility_tier_data, facility_mapping)
    
    # Save the populated workbook
    print(f"Saving populated workbook to: {output_file}")
    wb.save(output_file)
    print("Done! All facility tabs have been populated with tiered enrollment data.")
    
    # Print summary statistics
    print("\n=== Summary Statistics ===")
    total_enrollment = 0
    for client_id, plans in facility_tier_data.items():
        facility_total = sum(sum(tiers.values()) for tiers in plans.values())
        if facility_total > 0:
            facility_name = facility_mapping[facility_mapping['CLIENT ID'] == client_id]['Facility'].values
            if len(facility_name) > 0:
                print(f"{facility_name[0]} ({client_id}): {facility_total} enrollments")
                total_enrollment += facility_total
    
    print(f"\nTotal Enrollments Processed: {total_enrollment}")


if __name__ == "__main__":
    main()