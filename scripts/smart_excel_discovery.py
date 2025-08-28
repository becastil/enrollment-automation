#!/usr/bin/env python3
"""
Smart Excel Auto-Discovery System
=================================

This module dynamically discovers where to place enrollment data in Excel templates
by scanning for Client IDs, plan names, and tier structures - mimicking Ctrl+F functionality.

Key Features:
- Automatically finds Client ID headers in ANY column
- Auto-detects column layout or accepts configuration
- Discovers all associated plan sections (1 to N plans)
- Dynamically maps tier and value column locations
- Handles variable numbers of plans per facility
- Works with any template column structure
"""

import os
import sys
import json
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
import re
from typing import Dict, List, Tuple, Optional, Any

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from enrollment_automation_v6 import CID_TO_TAB, load_block_aggregations


class TemplateAnalyzer:
    """Analyzes Excel templates to discover structure and cell locations"""
    
    def __init__(self, template_path: str, column_config: Dict = None):
        """
        Initialize with template file path and optional column configuration.
        
        Args:
            template_path: Path to Excel template
            column_config: Optional dictionary specifying columns:
                {
                    'client_id_columns': ['D', 'E'],  # Columns to search for Client IDs
                    'tier_column': 'F',                # Column containing tier names
                    'value_column': 'G'                # Column for enrollment values
                }
        """
        self.template_path = template_path
        self.workbook = None
        self.discovery_cache = {}
        self.column_config = column_config or {}
        self.detected_columns = {}  # Will store auto-detected column layout
        
    def load_template(self):
        """Load the Excel template"""
        try:
            self.workbook = load_workbook(self.template_path, read_only=False, keep_vba=True)
            print(f"✓ Loaded template: {self.template_path}")
            return True
        except Exception as e:
            print(f"✗ Error loading template: {e}")
            return False
    
    def auto_detect_columns(self, worksheet) -> Dict[str, Any]:
        """
        Auto-detect which columns contain Client IDs, tiers, and values.
        
        Returns:
            Dictionary with detected column structure
        """
        detected = {
            'client_id_columns': [],
            'tier_column': None,
            'value_column': None,
            'plan_column': None
        }
        
        # Scan first 50 rows across columns A-Z
        max_rows_to_scan = min(50, worksheet.max_row)
        columns_to_scan = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[:min(26, worksheet.max_column)]
        
        client_id_pattern_count = {}
        tier_pattern_count = {}
        numeric_pattern_count = {}
        
        for col in columns_to_scan:
            client_ids_found = 0
            tiers_found = 0
            numbers_found = 0
            
            for row in range(1, max_rows_to_scan + 1):
                cell_value = self._get_cell_value(worksheet, f'{col}{row}')
                
                if cell_value:
                    # Check for Client ID patterns
                    if self._looks_like_client_id(cell_value):
                        client_ids_found += 1
                    
                    # Check for tier patterns
                    if self._looks_like_tier(cell_value):
                        tiers_found += 1
                    
                    # Check for numeric values
                    try:
                        float(str(cell_value).replace(',', ''))
                        numbers_found += 1
                    except:
                        pass
            
            # Record pattern counts
            if client_ids_found > 0:
                client_id_pattern_count[col] = client_ids_found
            if tiers_found > 0:
                tier_pattern_count[col] = tiers_found
            if numbers_found > 0:
                numeric_pattern_count[col] = numbers_found
        
        # Determine client ID columns (may be multiple)
        if client_id_pattern_count:
            # Sort by count and take columns with significant matches
            sorted_cols = sorted(client_id_pattern_count.items(), key=lambda x: x[1], reverse=True)
            detected['client_id_columns'] = [col for col, count in sorted_cols if count >= 1]
        
        # Determine tier column (usually has most tier matches)
        if tier_pattern_count:
            detected['tier_column'] = max(tier_pattern_count.items(), key=lambda x: x[1])[0]
            
            # Value column is typically next to tier column
            if detected['tier_column']:
                tier_col_index = ord(detected['tier_column']) - ord('A')
                next_col_index = tier_col_index + 1
                if next_col_index < len(columns_to_scan):
                    detected['value_column'] = chr(ord('A') + next_col_index)
        
        # Plan column is often same as client ID column
        if detected['client_id_columns']:
            detected['plan_column'] = detected['client_id_columns'][0]
        
        print(f"  Auto-detected columns: {detected}")
        return detected
    
    def _looks_like_client_id(self, value: str) -> bool:
        """Check if value looks like a Client ID or facility header"""
        if not value:
            return False
        
        value_upper = str(value).upper().strip()
        
        # Direct Client ID patterns
        if re.match(r'^H\d{4}', value_upper):  # H#### pattern
            return True
        
        # Client ID text patterns
        if 'CLIENT ID' in value_upper:
            return True
        
        # Facility name indicators
        facility_indicators = ['HOSPITAL', 'MEDICAL CENTER', 'CLINIC', 'HEALTH']
        return any(indicator in value_upper for indicator in facility_indicators)
    
    def _looks_like_tier(self, value: str) -> bool:
        """Check if value looks like a tier name"""
        if not value:
            return False
        
        value_upper = str(value).upper().strip()
        
        tier_patterns = [
            'EE ONLY', 'EE+', 'EMPLOYEE', 
            'SPOUSE', 'CHILD', 'FAMILY', 'DEPENDENT',
            'EMP', 'ESP', 'ECH', 'FAM', 'E1D'
        ]
        
        return any(pattern in value_upper for pattern in tier_patterns)
    
    def discover_facility_structure(self, worksheet, client_id: str) -> Dict:
        """
        Discover the structure for a specific facility in a worksheet.
        
        Returns a dictionary mapping with discovered cell locations.
        """
        structure = {
            'header_cell': None,
            'header_column': None,
            'plans': []
        }
        
        # Auto-detect columns if not already done
        if not self.detected_columns:
            self.detected_columns = self.auto_detect_columns(worksheet)
        
        # Get columns to use (from config or auto-detected)
        columns_to_use = self._get_columns_to_use()
        
        # Search for Client ID in detected columns
        header_location = self._find_client_id_header(worksheet, client_id, columns_to_use)
        if not header_location:
            return structure
        
        header_col, header_row = header_location
        structure['header_cell'] = f'{header_col}{header_row}'
        structure['header_column'] = header_col
        print(f"  Found {client_id} at {header_col}{header_row}")
        
        # Scan downward to find all plan sections
        current_row = header_row + 1
        max_scan_rows = 50  # Reasonable limit to prevent infinite scanning
        
        # Use detected or configured plan column (often same as header column)
        plan_col = columns_to_use.get('plan_column', header_col)
        
        while current_row < header_row + max_scan_rows:
            # Check if this row has a plan name in the plan column
            plan_name = self._get_cell_value(worksheet, f'{plan_col}{current_row}')
            
            if self._is_plan_name(plan_name):
                # Found a plan - now discover its tiers
                plan_info = {
                    'plan_name': plan_name,
                    'plan_cell': f'{plan_col}{current_row}',
                    'tiers': {}
                }
                
                # Scan for tiers starting from the next row
                tier_start_row = current_row + 1
                plan_info['tiers'] = self._discover_plan_tiers(worksheet, tier_start_row, columns_to_use)
                
                structure['plans'].append(plan_info)
                print(f"    Plan: {plan_name} with {len(plan_info['tiers'])} tiers")
                
                # Move past this plan's tiers
                current_row = tier_start_row + len(plan_info['tiers'])
            else:
                # Check if we've hit another Client ID or empty section
                if self._is_client_id_header(plan_name) or not plan_name:
                    break
                current_row += 1
        
        return structure
    
    def _get_columns_to_use(self) -> Dict:
        """Get columns to use from config or auto-detection"""
        columns = {}
        
        # Use configured columns if available, otherwise use detected
        if 'client_id_columns' in self.column_config:
            columns['client_id_columns'] = self.column_config['client_id_columns']
        elif self.detected_columns.get('client_id_columns'):
            columns['client_id_columns'] = self.detected_columns['client_id_columns']
        else:
            # Default fallback - search common columns
            columns['client_id_columns'] = ['D', 'E', 'C', 'B', 'A']
        
        # Tier column
        if 'tier_column' in self.column_config:
            columns['tier_column'] = self.column_config['tier_column']
        elif self.detected_columns.get('tier_column'):
            columns['tier_column'] = self.detected_columns['tier_column']
        else:
            columns['tier_column'] = 'F'  # Default
        
        # Value column
        if 'value_column' in self.column_config:
            columns['value_column'] = self.column_config['value_column']
        elif self.detected_columns.get('value_column'):
            columns['value_column'] = self.detected_columns['value_column']
        else:
            columns['value_column'] = 'G'  # Default
        
        # Plan column (often same as client ID column)
        if 'plan_column' in self.column_config:
            columns['plan_column'] = self.column_config['plan_column']
        elif self.detected_columns.get('plan_column'):
            columns['plan_column'] = self.detected_columns['plan_column']
        elif columns['client_id_columns']:
            columns['plan_column'] = columns['client_id_columns'][0]
        else:
            columns['plan_column'] = 'D'  # Default
        
        return columns
    
    def _find_client_id_header(self, worksheet, client_id: str, columns_to_use: Dict) -> Optional[Tuple[str, int]]:
        """
        Find the Client ID header in any column.
        
        Returns:
            Tuple of (column, row) if found, None otherwise
        """
        # Search patterns - facility names might have variations
        search_patterns = [
            client_id,  # Exact match
            f"Client ID {client_id}",
            f"Client ID: {client_id}",
            f"{client_id} -",  # With facility name
        ]
        
        # Get columns to search
        search_columns = columns_to_use.get('client_id_columns', [])
        
        # If no specific columns, search all columns A-Z
        if not search_columns:
            max_col = min(26, worksheet.max_column)
            search_columns = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ'[:max_col])
        
        # Search in each column
        for col in search_columns:
            for row in range(1, min(worksheet.max_row + 1, 1000)):  # Limit to 1000 rows
                cell_value = self._get_cell_value(worksheet, f'{col}{row}')
                if cell_value:
                    # Check if any pattern matches
                    for pattern in search_patterns:
                        if pattern.upper() in str(cell_value).upper():
                            return (col, row)
        
        return None
    
    def _is_plan_name(self, value: str) -> bool:
        """Check if a value looks like a plan name"""
        if not value:
            return False
        
        value_upper = str(value).upper()
        plan_indicators = [
            'PRIME', 'EPO', 'VALUE', 'PLAN', 
            'SELF-INSURED', 'UNION', 'NON-UNION',
            'UNIFIED', 'SEIU', 'CNA', 'CWA'
        ]
        
        # Must contain at least 2 plan indicators
        matches = sum(1 for indicator in plan_indicators if indicator in value_upper)
        return matches >= 2
    
    def _is_client_id_header(self, value: str) -> bool:
        """Check if a value is another Client ID header"""
        if not value:
            return False
        
        value_upper = str(value).upper()
        return 'CLIENT ID' in value_upper or re.match(r'^H\d{4}', value_upper)
    
    def _discover_plan_tiers(self, worksheet, start_row: int, columns_to_use: Dict) -> Dict[str, str]:
        """Discover tier locations for a plan starting at given row"""
        tiers = {}
        
        # Get tier and value columns from config/detection
        tier_col = columns_to_use.get('tier_column', 'F')
        value_col = columns_to_use.get('value_column', 'G')
        
        # Standard tier names we're looking for
        standard_tiers = [
            'EE Only', 'EE+Spouse', 'EE+Child(ren)', 'EE+Family',
            'EE+1 Dep',  # 5-tier structure
            'EMP', 'ESP', 'ECH', 'E1D', 'FAM'  # Alternative names
        ]
        
        # Scan next few rows for tiers in the tier column
        for offset in range(8):  # Max 8 rows for tiers
            row = start_row + offset
            tier_name = self._get_cell_value(worksheet, f'{tier_col}{row}')
            
            if tier_name:
                # Normalize tier name
                normalized = self._normalize_tier_name(tier_name)
                if normalized:
                    # The enrollment value cell is in the value column
                    tiers[normalized] = f'{value_col}{row}'
        
        return tiers
    
    def _normalize_tier_name(self, tier_name: str) -> Optional[str]:
        """Normalize tier names to standard format"""
        if not tier_name:
            return None
        
        tier_upper = str(tier_name).upper().strip()
        
        # Mapping of variations to standard names
        tier_map = {
            'EE ONLY': 'EE Only',
            'EMP': 'EE Only',
            'EMPLOYEE': 'EE Only',
            'EMPLOYEE ONLY': 'EE Only',
            
            'EE+SPOUSE': 'EE+Spouse',
            'ESP': 'EE+Spouse',
            'EMPLOYEE + SPOUSE': 'EE+Spouse',
            'EE + SPOUSE': 'EE+Spouse',
            
            'EE+CHILD(REN)': 'EE+Child(ren)',
            'EE+CHILDREN': 'EE+Child(ren)',
            'ECH': 'EE+Child(ren)',
            'EMPLOYEE + CHILD': 'EE+Child(ren)',
            'EE + CHILD': 'EE+Child(ren)',
            
            'EE+FAMILY': 'EE+Family',
            'FAM': 'EE+Family',
            'FAMILY': 'EE+Family',
            'EMPLOYEE + FAMILY': 'EE+Family',
            
            'EE+1 DEP': 'EE+1 Dep',
            'E1D': 'EE+1 Dep',
            'EMPLOYEE + 1': 'EE+1 Dep',
            'EE + 1 DEPENDENT': 'EE+1 Dep'
        }
        
        # Check for matches
        for pattern, standard in tier_map.items():
            if pattern in tier_upper:
                return standard
        
        return None
    
    def _get_cell_value(self, worksheet, cell_ref: str) -> Optional[str]:
        """Get value from a cell safely"""
        try:
            cell = worksheet[cell_ref]
            if cell.value is not None:
                return str(cell.value).strip()
        except:
            pass
        return None
    
    def discover_tab_structure(self, tab_name: str) -> Dict:
        """Discover complete structure for a specific tab"""
        if not self.workbook:
            if not self.load_template():
                return {}
        
        if tab_name not in self.workbook.sheetnames:
            print(f"✗ Tab '{tab_name}' not found in template")
            return {}
        
        worksheet = self.workbook[tab_name]
        print(f"\nAnalyzing tab: {tab_name}")
        
        # Get all Client IDs that should be in this tab
        facilities_in_tab = {
            cid: name for cid, name in CID_TO_TAB.items() 
            if name == tab_name
        }
        
        tab_structure = {
            'tab_name': tab_name,
            'facilities': {}
        }
        
        # Discover structure for each facility
        for client_id in facilities_in_tab:
            facility_structure = self.discover_facility_structure(worksheet, client_id)
            if facility_structure['header_cell']:
                tab_structure['facilities'][client_id] = facility_structure
        
        # Cache the discovery
        self.discovery_cache[tab_name] = tab_structure
        
        return tab_structure
    
    def discover_all_tabs(self) -> Dict:
        """Discover structure for all tabs in the template"""
        if not self.workbook:
            if not self.load_template():
                return {}
        
        all_structures = {}
        
        # Get unique tab names from CID_TO_TAB
        unique_tabs = set(CID_TO_TAB.values())
        
        for tab_name in unique_tabs:
            if tab_name in self.workbook.sheetnames:
                structure = self.discover_tab_structure(tab_name)
                if structure and structure['facilities']:
                    all_structures[tab_name] = structure
        
        return all_structures
    
    def save_discovery_map(self, output_path: str = None):
        """Save discovered structure to JSON file"""
        if not output_path:
            output_path = 'config/discovered_cell_mappings.json'
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(self.discovery_cache, f, indent=2)
        
        print(f"\n✓ Saved discovery map to: {output_path}")
        return output_path


class SmartExcelWriter:
    """Writes enrollment data using discovered cell mappings"""
    
    def __init__(self, template_path: str, discovery_map: Dict = None):
        """Initialize with template and optional pre-discovered map"""
        self.template_path = template_path
        self.discovery_map = discovery_map or {}
        self.workbook = None
        self.analyzer = TemplateAnalyzer(template_path)
    
    def load_template(self):
        """Load the Excel template for writing"""
        try:
            self.workbook = load_workbook(self.template_path, keep_vba=True)
            return True
        except Exception as e:
            print(f"Error loading template for writing: {e}")
            return False
    
    def write_enrollment_value(self, tab_name: str, client_id: str, 
                              plan_name: str, tier: str, value: Any) -> bool:
        """Write a single enrollment value using discovery"""
        
        # Get or discover structure for this tab
        if tab_name not in self.discovery_map:
            tab_structure = self.analyzer.discover_tab_structure(tab_name)
            if not tab_structure or not tab_structure['facilities']:
                print(f"✗ Could not discover structure for {tab_name}")
                return False
            self.discovery_map[tab_name] = tab_structure
        
        # Find the specific cell
        tab_data = self.discovery_map[tab_name]
        
        if client_id not in tab_data['facilities']:
            print(f"✗ Client {client_id} not found in {tab_name}")
            return False
        
        facility_data = tab_data['facilities'][client_id]
        
        # Find matching plan
        for plan_info in facility_data['plans']:
            if self._match_plan_name(plan_name, plan_info['plan_name']):
                # Find tier cell
                normalized_tier = self.analyzer._normalize_tier_name(tier)
                if normalized_tier in plan_info['tiers']:
                    cell_ref = plan_info['tiers'][normalized_tier]
                    
                    # Write the value
                    worksheet = self.workbook[tab_name]
                    worksheet[cell_ref] = value
                    
                    print(f"  ✓ Wrote {value} to {cell_ref} ({client_id}/{plan_name}/{tier})")
                    return True
        
        print(f"✗ Could not find cell for {client_id}/{plan_name}/{tier}")
        return False
    
    def _match_plan_name(self, search_name: str, template_name: str) -> bool:
        """Check if plan names match (fuzzy matching)"""
        # Simple containment check - can be enhanced
        search_upper = str(search_name).upper()
        template_upper = str(template_name).upper()
        
        # Check if key parts match
        if 'EPO' in search_upper and 'EPO' in template_upper:
            return True
        if 'VALUE' in search_upper and 'VALUE' in template_upper:
            return True
        
        # More specific matching can be added
        return search_upper in template_upper or template_upper in search_upper
    
    def save_workbook(self, output_path: str = None):
        """Save the workbook with written values"""
        if not output_path:
            output_path = self.template_path.replace('.xlsx', '_updated.xlsx')
        
        self.workbook.save(output_path)
        print(f"\n✓ Saved updated workbook to: {output_path}")
        return output_path


def test_discovery_with_legacy():
    """Test the discovery system with Legacy tab"""
    
    print("\n" + "="*70)
    print("TESTING SMART EXCEL AUTO-DISCOVERY")
    print("="*70)
    
    # Template path (adjust as needed)
    template_path = r"C:\Users\becas\Prime_EFR\templates\enrollment_template.xlsx"
    
    # Check if template exists
    if not os.path.exists(template_path):
        print(f"\n✗ Template not found at: {template_path}")
        print("  Creating mock discovery for demonstration...")
        
        # Mock discovery result for Legacy tab
        mock_discovery = {
            "Legacy": {
                "tab_name": "Legacy",
                "facilities": {
                    "H3210": {
                        "header_cell": "D100",
                        "plans": [
                            {
                                "plan_name": "PRIME EPO PLAN (Self-Insured)",
                                "plan_cell": "D101",
                                "tiers": {
                                    "EE Only": "G102",
                                    "EE+Spouse": "G103",
                                    "EE+Child(ren)": "G104",
                                    "EE+Family": "G105"
                                }
                            },
                            {
                                "plan_name": "PRIME VALUE PLAN (Self-Insured)",
                                "plan_cell": "D106",
                                "tiers": {
                                    "EE Only": "G107",
                                    "EE+Spouse": "G108",
                                    "EE+Child(ren)": "G109",
                                    "EE+Family": "G110"
                                }
                            }
                        ]
                    }
                }
            }
        }
        
        # Save mock discovery
        os.makedirs('config', exist_ok=True)
        with open('config/discovered_cell_mappings.json', 'w') as f:
            json.dump(mock_discovery, f, indent=2)
        
        print("\n✓ Created mock discovery map")
        print("\nMock structure for H3210 (Huntington Beach Hospital):")
        print("  - Found at row 100")
        print("  - 2 plans discovered:")
        print("    1. PRIME EPO PLAN with 4 tiers")
        print("    2. PRIME VALUE PLAN with 4 tiers")
        
        return mock_discovery
    
    # Run actual discovery
    analyzer = TemplateAnalyzer(template_path)
    
    # Test with Legacy tab
    legacy_structure = analyzer.discover_tab_structure('Legacy')
    
    if legacy_structure and legacy_structure['facilities']:
        print(f"\nDiscovered {len(legacy_structure['facilities'])} facilities in Legacy tab")
        
        # Show sample discovery
        for client_id, structure in list(legacy_structure['facilities'].items())[:2]:
            print(f"\n{client_id}:")
            print(f"  Header: {structure['header_cell']}")
            print(f"  Plans: {len(structure['plans'])}")
            for plan in structure['plans']:
                print(f"    - {plan['plan_name']}: {len(plan['tiers'])} tiers")
        
        # Save discovery
        analyzer.save_discovery_map()
        
        return analyzer.discovery_cache
    else:
        print("\n✗ No facilities discovered in Legacy tab")
        return {}


def main():
    """Main function to demonstrate smart discovery"""
    
    print("\n" + "="*70)
    print("SMART EXCEL AUTO-DISCOVERY SYSTEM")
    print("="*70)
    print("\nThis system automatically discovers where to place enrollment data")
    print("by scanning Excel templates for Client IDs, plans, and tiers.")
    
    # Run discovery test
    discovery_map = test_discovery_with_legacy()
    
    if discovery_map:
        print("\n" + "="*70)
        print("DISCOVERY COMPLETE")
        print("="*70)
        print("\nThe system can now:")
        print("  1. Find any Client ID in column D (like Ctrl+F)")
        print("  2. Discover all associated plans below it")
        print("  3. Map tier locations in column F")
        print("  4. Identify where to write values in column G")
        print("\nNo more static cell mappings needed!")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())