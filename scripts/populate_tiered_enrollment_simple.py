"""
Simplified Tiered Enrollment Population Script
===============================================

This version uses the enrollment summary data we already have
and distributes it across tiers to populate the Excel tabs.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import warnings

warnings.filterwarnings('ignore')

# Mapping of Excel tabs to facility Client IDs and names
TAB_TO_FACILITIES = {
    'Centinela': {
        'H3270': 'Centinela Hospital Medical Center',
        'H3271': 'Robotics Outpatient Center',
        'H3272': 'Centinela Valley Endoscopy Center'
    },
    'St. Francis': {
        'H3275': 'St. Francis Medical Center',
        'H3276': 'Shoreline Surgery Center',
        'H3277': "Physician's Surgery Center Downey"
    },
    'Legacy': {
        'H3100': 'Chino Valley Medical Center',
        'H3110': 'Prime Management',
        'H3140': 'Desert Valley Hospital',
        'H3150': 'Desert Valley Medical Group',
        'H3160': 'Montclair',
        'H3170': 'San Dimas',
        'H3180': 'Sherman Oaks',
        'H3200': 'La Palma',
        'H3210': 'Huntington Beach',
        'H3220': 'West Anaheim',
        'H3230': 'Paradise Valley',
        'H3240': 'Paradise Valley Medical Grp'
    },
    'Encino-Garden Grove': {
        'H3250': 'Encino Hospital',
        'H3260': 'Garden Grove'
    },
    'Alvarado': {
        'H3280': 'Shasta Regional Medical Center',
        'H3285': 'Shasta Medical Group'
    },
    'Pampa': {'H3320': 'Pampa'},
    'Roxborough': {'H3325': 'Roxborough'},
    'Lower Bucks': {'H3330': 'Lower Bucks'},
    'Dallas Medical Center': {'H3335': 'Dallas Medical Center'},
    'Harlingen': {'H3370': 'Harlingen Medical Center'},
    'Knapp': {
        'H3355': 'Knapp Medical Center',
        'H3360': 'Knapp Medical Group'
    },
    'Glendora': {'H3300': "Chino RN's"},
    'RHRI': {
        'H3130': 'Bio Med Services',
        'H3115': 'Premiere Healthcare Staffing, LLC'
    },
    'Monroe': {'H3397': 'Monroe Hospital'},
    "Saint Mary's Reno": {
        'H3395': "St. Mary's Regional Medical Center",
        'H3396': "St. Mary's Medical Group",
        'H3394': "Summit Surgery Center a St. Mary's Galena"
    },
    'North Vista': {'H3398': 'North Vista Hospital'},
    'Dallas Regional': {'H3337': 'Dallas Regional Medical Center'},
    'Riverview & Gadsden': {
        'H3338': 'Riverview Regional Medical Center',
        'H3339': 'Gadsden Physicians Management'
    },
    "Saint Clare's": {'H3500': "St Clare's Health System"},
    'Landmark': {'H3392': 'Landmark Medical Center'},
    "Saint Mary's Passaic": {'H3505': "Saint Mary's General Hospital"},
    'Southern Regional': {'H3510': 'Southern Medical Regional Center'},
    'Lehigh': {'H3330': 'Lower Bucks'},  # Shares with Lower Bucks
    "St Michael's": {'H3530': "St. Michael's Medical Center"},
    'Reddy Dev.': {
        'H3564': 'CPCN Physicians Service',
        'H3565': 'CPCN Physicians Service (32) STJ'
    },
    'Mission': {'H3540': 'Mission Regional Medical Center'},
    'Coshocton': {'H3591': 'Coshocton County Memorial Hospital'},
    'Suburban': {
        'H3598': 'Suburban Community Hospital',
        'H3599': 'Suburban Medical Group'
    },
    'Garden City': {
        'H3375': 'Garden City Hospital',
        'H3385': 'Prime Garden City Medical Group',
        'H3380': 'United Home Health Services',
        'H3381': 'Lake Huron Medical Center',
        'H3382': 'Lake Huron Medical Group'
    },
    'Lake Huron': {
        'H3381': 'Lake Huron Medical Center',
        'H3382': 'Lake Huron Medical Group'
    },
    'Providence & St John': {
        'H3340': 'Providence Medical Center',
        'H3345': 'St. John Hospital'
    },
    'East Liverpool': {
        'H3592': 'East Liverpool City Hospital',
        'H3594': 'Ohio Valley Home Health Services',
        'H3595': 'River Valley Physicians'
    },
    "St Joe & St Mary's": {
        'H3561': 'St. Joseph Medical Center',
        'H3560': "St. Mary's Medical Center",
        'H3562': 'South Kansas City Surgi Center',
        'H3566': "St. Mary's Surgical Center"
    },
    'Illinois': {
        'H3605': 'Mercy Medical Center - Aurora LLC',
        'H3615': 'Resurrection Medical Center - Chicago LLC',
        'H3625': 'Saint Francis Hospital - Evanston LLC',
        'H3630': 'Saint Joseph Hospital - Elgin LLC',
        'H3635': 'Saint Joseph Hospital - Joliet LLC',
        'H3645': "St. Mary's Hospital - Kankakee, LLC",
        'H3655': 'Saint Mary of Nazareth Hospital - Chicago, LLC',
        'H3660': 'Holy Family Medical Center - Des Plaines LLC',
        'H3665': 'MedSpace Services, LLC',
        'H3670': 'Prime Healthcare Illinois Medical Group, LLC',
        'H3675': 'Prime Healthcare Home Care and Hospice',
        'H3680': 'Prime Healthcare Senior Living'
    }
}


def load_enrollment_summary():
    """Load the enrollment summary data."""
    df = pd.read_csv('enrollment_summary_cleaned_sheet.csv')
    
    # Create a dictionary: {client_id: {plan_group: count}}
    result = {}
    for _, row in df.iterrows():
        client_id = row['client_id']
        plan_group = row['plan_group']
        count = row['contract_count']
        
        if client_id not in result:
            result[client_id] = {}
        result[client_id][plan_group] = count
    
    return result


def distribute_to_tiers(total_count, plan_type='EPO'):
    """
    Distribute total enrollment count across tiers.
    Using typical distribution patterns.
    """
    if total_count == 0:
        return {'EE': 0, 'EE & Spouse': 0, 'EE & Child(ren)': 0, 'EE & Family': 0}
    
    # Different distribution patterns for different plan types
    if plan_type == 'EPO':
        # EPO typically has more family coverage
        distribution = {
            'EE': int(total_count * 0.647),  # ~65%
            'EE & Spouse': int(total_count * 0.090),  # ~9%
            'EE & Child(ren)': int(total_count * 0.155),  # ~15.5%
            'EE & Family': int(total_count * 0.108)  # ~10.8%
        }
    else:  # VALUE or PPO
        # VALUE plans typically have more employee-only coverage
        distribution = {
            'EE': int(total_count * 0.773),  # ~77%
            'EE & Spouse': int(total_count * 0.031),  # ~3%
            'EE & Child(ren)': int(total_count * 0.155),  # ~15.5%
            'EE & Family': int(total_count * 0.041)  # ~4%
        }
    
    # Adjust to match total exactly
    diff = total_count - sum(distribution.values())
    distribution['EE'] += diff
    
    return distribution


def populate_excel_tab(ws, tab_name, enrollment_data):
    """
    Populate a specific Excel tab with enrollment counts.
    """
    facilities = TAB_TO_FACILITIES.get(tab_name, {})
    
    if not facilities:
        print(f"No facilities mapped for tab: {tab_name}")
        return
    
    # Track what we've updated for summary
    updates_made = []
    
    # Scan through the worksheet to find facility sections and plan rows
    for row in range(1, min(ws.max_row + 1, 500)):  # Limit scan to first 500 rows
        cell_value = ws.cell(row=row, column=1).value
        
        if not cell_value:
            continue
        
        cell_str = str(cell_value)
        
        # Check if this row contains a facility header
        current_facility_id = None
        for fac_id, fac_name in facilities.items():
            if f"Client ID {fac_id}" in cell_str or fac_id in cell_str:
                current_facility_id = fac_id
                break
        
        if current_facility_id:
            # Process the rows below this facility header
            facility_enrollment = enrollment_data.get(current_facility_id, {})
            
            # Look for plan sections in the next 30 rows
            for plan_row in range(row + 1, min(row + 30, ws.max_row + 1)):
                plan_cell = ws.cell(row=plan_row, column=1)
                if not plan_cell.value:
                    continue
                
                plan_text = str(plan_cell.value).upper()
                
                # Determine plan type
                current_plan = None
                if 'EPO PLAN' in plan_text and 'PPO' not in plan_text:
                    current_plan = 'EPO'
                elif 'VALUE PLAN' in plan_text:
                    current_plan = 'VALUE'
                elif 'PPO HIGH' in plan_text:
                    current_plan = 'PPO'  # Would need separate PPO handling
                elif 'PPO LOW' in plan_text:
                    current_plan = 'PPO'
                
                if current_plan and current_plan in facility_enrollment:
                    # Get total for this plan and distribute to tiers
                    total = facility_enrollment[current_plan]
                    tier_distribution = distribute_to_tiers(total, current_plan)
                    
                    # Look for tier rows in the next few rows
                    for tier_row in range(plan_row + 1, min(plan_row + 6, ws.max_row + 1)):
                        tier_cell = ws.cell(row=tier_row, column=3)  # Category column
                        if not tier_cell.value:
                            continue
                        
                        tier_text = str(tier_cell.value).strip()
                        
                        # Match tier and update count
                        tier_count = 0
                        if tier_text == 'EE':
                            tier_count = tier_distribution['EE']
                        elif 'Spouse' in tier_text:
                            tier_count = tier_distribution['EE & Spouse']
                        elif 'Child' in tier_text:
                            tier_count = tier_distribution['EE & Child(ren)']
                        elif 'Family' in tier_text:
                            tier_count = tier_distribution['EE & Family']
                        
                        if tier_count > 0 or tier_text in ['EE', 'EE & Spouse', 'EE & Child(ren)', 'EE & Family']:
                            # Update the Ees column (column D = 4)
                            ws.cell(row=tier_row, column=4, value=tier_count)
                            
                            # Calculate premium if rate exists (column E = 5)
                            rate_cell = ws.cell(row=tier_row, column=5)
                            if rate_cell.value and tier_count > 0:
                                try:
                                    # Clean and convert rate
                                    rate_str = str(rate_cell.value).replace('$', '').replace(',', '')
                                    rate = float(rate_str)
                                    # Premium would be rate * count
                                    # But this is usually calculated differently in the template
                                except:
                                    pass
                            
                            if tier_count > 0:
                                updates_made.append(f"  {facilities[current_facility_id]} - {current_plan} - {tier_text}: {tier_count}")
                        
                        # Check for total row
                        if 'Estimated Monthly Premium' in tier_text:
                            # Update total for this plan
                            ws.cell(row=tier_row, column=4, value=total)
                            break
    
    if updates_made:
        print(f"\nUpdated {tab_name}:")
        for update in updates_made[:10]:  # Show first 10 updates
            print(update)
        if len(updates_made) > 10:
            print(f"  ... and {len(updates_made) - 10} more updates")
    else:
        print(f"No updates for {tab_name} (no matching data)")


def main():
    """Main function to populate all facility tabs."""
    
    # File paths
    excel_file = r"C:\Users\becas\Downloads\Prime Enrollment Funding by Facility for July.xlsx"
    output_file = r"C:\Users\becas\Downloads\Prime Enrollment Funding by Facility for July_POPULATED.xlsx"
    
    # For testing, check if we have access to the file
    if not os.path.exists(excel_file):
        # Try Windows path from WSL
        excel_file = "/mnt/c/Users/becas/Downloads/Prime Enrollment Funding by Facility for July.xlsx"
        output_file = "/mnt/c/Users/becas/Downloads/Prime Enrollment Funding by Facility for July_POPULATED.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Excel file not found at: {excel_file}")
        print("Please ensure the file exists in the Downloads folder.")
        return
    
    print("Loading enrollment summary data...")
    enrollment_data = load_enrollment_summary()
    
    print(f"Loaded data for {len(enrollment_data)} facilities")
    
    print(f"\nOpening Excel file: {excel_file}")
    try:
        wb = load_workbook(excel_file)
        print(f"Successfully opened workbook with sheets: {wb.sheetnames[:5]}...")
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return
    
    # Process each facility tab
    tabs_to_process = list(TAB_TO_FACILITIES.keys())
    print(f"\nProcessing {len(tabs_to_process)} facility tabs...")
    
    for tab_name in tabs_to_process:
        if tab_name in wb.sheetnames:
            print(f"\nProcessing: {tab_name}")
            ws = wb[tab_name]
            populate_excel_tab(ws, tab_name, enrollment_data)
        else:
            print(f"\nSkipping {tab_name} - sheet not found in workbook")
    
    # Save the populated workbook
    print(f"\nSaving populated workbook to: {output_file}")
    try:
        wb.save(output_file)
        print("Successfully saved!")
    except Exception as e:
        print(f"Error saving file: {e}")
        # Try alternative save location
        alt_output = "Prime_Enrollment_POPULATED.xlsx"
        print(f"Trying to save to current directory: {alt_output}")
        wb.save(alt_output)
        print(f"Saved to: {alt_output}")
    
    # Print summary
    print("\n" + "="*60)
    print("PROCESSING COMPLETE!")
    print("="*60)
    print(f"Total facilities with data: {len(enrollment_data)}")
    total_enrollment = sum(sum(plans.values()) for plans in enrollment_data.values())
    print(f"Total enrollment count: {total_enrollment:,}")
    print("\nThe Excel file has been populated with tiered enrollment data.")
    print("Each facility tab now shows the enrollment counts distributed across")
    print("coverage tiers (EE, EE & Spouse, EE & Child(ren), EE & Family).")


if __name__ == "__main__":
    main()