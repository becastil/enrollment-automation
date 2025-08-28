#!/usr/bin/env python3
"""
Diagnose Excel File Corruption Issue
====================================

This script diagnoses why the generated Excel file cannot be opened.
"""

import os
import sys
import openpyxl
from openpyxl import load_workbook, Workbook
import zipfile
from datetime import datetime

def check_file_basics(filepath):
    """Check basic file properties"""
    print(f"\n📁 FILE BASICS")
    print("="*50)
    
    if not os.path.exists(filepath):
        print(f"✗ File does not exist: {filepath}")
        return False
    
    print(f"✓ File exists: {filepath}")
    
    # Check file size
    size = os.path.getsize(filepath)
    print(f"  Size: {size:,} bytes")
    
    if size == 0:
        print("✗ File is empty (0 bytes)")
        return False
    elif size < 1000:
        print("⚠️ File is suspiciously small")
    
    return True

def check_file_is_zip(filepath):
    """Check if file is a valid ZIP (Excel files are ZIP archives)"""
    print(f"\n🔍 ZIP STRUCTURE CHECK")
    print("="*50)
    
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            print("✓ File is a valid ZIP archive")
            
            # List contents
            files = z.namelist()
            print(f"  Contains {len(files)} entries")
            
            # Check for required Excel structure
            required_files = [
                "xl/workbook.xml",
                "[Content_Types].xml",
                "_rels/.rels"
            ]
            
            for req in required_files:
                if req in files:
                    print(f"  ✓ Found: {req}")
                else:
                    print(f"  ✗ Missing: {req}")
                    
            return True
    except zipfile.BadZipFile:
        print("✗ File is not a valid ZIP archive")
        print("  This means the file is corrupted or not an Excel file")
        return False
    except Exception as e:
        print(f"✗ Error checking ZIP: {e}")
        return False

def try_load_with_openpyxl(filepath):
    """Try to load file with openpyxl"""
    print(f"\n📊 OPENPYXL LOAD TEST")
    print("="*50)
    
    try:
        wb = load_workbook(filepath, read_only=True)
        print(f"✓ Successfully loaded with openpyxl")
        print(f"  Sheets: {wb.sheetnames}")
        wb.close()
        return True
    except Exception as e:
        print(f"✗ Failed to load with openpyxl")
        print(f"  Error: {e}")
        return False

def create_simple_test_file():
    """Create a simple test Excel file to verify openpyxl works"""
    print(f"\n🧪 CREATING TEST FILE")
    print("="*50)
    
    test_path = "/mnt/c/Users/becas/Prime_EFR/output/test_simple.xlsx"
    
    try:
        # Create simple workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "Test Value"
        ws["B1"] = 123
        
        # Save it
        os.makedirs(os.path.dirname(test_path), exist_ok=True)
        wb.save(test_path)
        wb.close()
        
        print(f"✓ Created test file: {test_path}")
        
        # Check if it's valid
        if check_file_is_zip(test_path):
            print("✓ Test file is valid Excel format")
        
        return test_path
    except Exception as e:
        print(f"✗ Failed to create test file: {e}")
        return None

def compare_with_template(template_path, output_path):
    """Compare output with original template"""
    print(f"\n🔄 COMPARING WITH TEMPLATE")
    print("="*50)
    
    if os.path.exists(template_path):
        template_size = os.path.getsize(template_path)
        output_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
        
        print(f"Template size: {template_size:,} bytes")
        print(f"Output size:   {output_size:,} bytes")
        
        if output_size < template_size * 0.5:
            print("⚠️ Output is much smaller than template")
            print("  This suggests content may be missing")
        elif output_size > template_size * 2:
            print("⚠️ Output is much larger than template")
            print("  This might indicate duplication")
        else:
            print("✓ File sizes are reasonable")

def suggest_fixes():
    """Suggest possible fixes"""
    print(f"\n💡 SUGGESTED FIXES")
    print("="*50)
    
    print("""
The issue is likely one of:

1. **File not properly closed**: The workbook wasn't properly closed after writing
   → Fix: Added workbook.close() after save

2. **Missing preservation flags**: Excel features not preserved during load
   → Fix: Added data_only=False and keep_links=True to preserve formulas

3. **Template has macros/VBA**: The template has VBA that's not being preserved
   → Fix: Added keep_vba=True flag when loading

4. **Output directory issue**: Path or permission issues
   → Fix: Added proper directory creation and error handling

The fixes have been applied. Please run the script again:
   python scripts/enrollment_writer_v2.py
""")

def main():
    """Main diagnostic function"""
    print("\n" + "="*70)
    print("EXCEL FILE DIAGNOSTIC")
    print("="*70)
    
    # Check the problem file
    problem_file = r"C:\Users\becas\Prime_EFR\output\enrollment_output_20250828_153531.xlsx"
    
    print(f"\nChecking: {problem_file}")
    
    # Run diagnostics
    if check_file_basics(problem_file):
        check_file_is_zip(problem_file)
        try_load_with_openpyxl(problem_file)
    
    # Create test file
    test_file = create_simple_test_file()
    if test_file:
        print(f"\n✅ If the test file opens in Excel, openpyxl is working correctly")
    
    # Compare with template
    template_path = r"C:\Users\becas\Prime_EFR\Prime Enrollment Funding by Facility for August.xlsx"
    compare_with_template(template_path, problem_file)
    
    # Suggest fixes
    suggest_fixes()
    
    print("\n" + "="*70)
    print("DIAGNOSTIC COMPLETE")
    print("="*70)

if __name__ == "__main__":
    sys.exit(main())