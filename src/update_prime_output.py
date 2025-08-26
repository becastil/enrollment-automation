#!/usr/bin/env python3
"""
Excel Data Ingestion and Update Script for Prime Employee Enrollment

This script reads employee enrollment data from a source Excel file and updates
Column D (employee counts by coverage tier) in a target Excel template file,
preserving all existing formatting, formulas, and structure.

Author: Senior Python Data Engineer
Date: 2025
"""

import argparse
import logging
import re
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ==============================================================================
# CONFIGURATION
# ==============================================================================

DEFAULT_CONFIG = {
    "source_path": r"C:\Users\becas\Prime_EFR\data\input\Data_file_prime.xlsx",
    "target_path": r"C:\Users\becas\Prime_EFR\data\input\Prime_output_file.xlsx",
    "source_sheet": 0,  # First sheet by default
    "target_sheet": 0,  # First sheet by default
    "header_row": 1,    # Row number where headers are located
    "key_source": "CLIENT ID",
    "key_target_pattern": r"Client ID ([A-Z0-9]+)",  # Pattern to extract client ID from target
}

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ==============================================================================
# DATA STRUCTURES
# ==============================================================================

class EmployeeCategory:
    """Employee coverage category enumeration"""
    EE = "EE"
    EE_SPOUSE = "EE & Spouse"
    EE_CHILDREN = "EE & Child(ren)"
    EE_FAMILY = "EE & Family"
    TOTAL = "Estimated Monthly Premium"
    
    @classmethod
    def all_categories(cls) -> List[str]:
        """Return all categories in order"""
        return [cls.EE, cls.EE_SPOUSE, cls.EE_CHILDREN, cls.EE_FAMILY, cls.TOTAL]


class PlanType:
    """Plan type enumeration"""
    EPO = "EPO"
    VALUE = "VALUE"


class UpdateResult:
    """Container for update results"""
    def __init__(self):
        self.rows_updated: int = 0
        self.rows_skipped: int = 0
        self.unmatched_clients: List[str] = []
        self.changes: List[Dict[str, Any]] = []
        
    def add_change(self, row: int, client_id: str, plan: str, category: str,
                   old_value: Any, new_value: Any):
        """Record a change made to the target file"""
        self.changes.append({
            "row": row,
            "client_id": client_id,
            "plan": plan,
            "category": category,
            "old_value": old_value,
            "new_value": new_value
        })
        self.rows_updated += 1


# ==============================================================================
# DATA LOADING FUNCTIONS
# ==============================================================================

def load_source_data(file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Load and validate source Excel data
    
    Args:
        file_path: Path to source Excel file
        sheet_name: Name or index of sheet to read
        
    Returns:
        DataFrame with source data
        
    Raises:
        FileNotFoundError: If source file doesn't exist
        ValueError: If required columns are missing
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Source file not found: {file_path}")
    
    logger.info(f"Loading source data from: {file_path}")
    
    # Read Excel file
    try:
        if sheet_name is None:
            df = pd.read_excel(file_path, sheet_name=0)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        logger.error(f"Failed to read source file: {e}")
        raise
    
    # Validate required columns
    required_columns = ['CLIENT ID', 'EMPLOYEE NAME', 'RELATION', 'EPO-PPO-VAL']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        raise ValueError(f"Missing required columns in source: {missing_columns}")
    
    # Clean data
    df['CLIENT ID'] = df['CLIENT ID'].astype(str).str.strip()
    df['EMPLOYEE NAME'] = df['EMPLOYEE NAME'].astype(str).str.strip()
    df['RELATION'] = df['RELATION'].fillna('SELF').astype(str).str.strip().str.upper()
    df['EPO-PPO-VAL'] = df['EPO-PPO-VAL'].astype(str).str.strip().str.upper()
    
    logger.info(f"Loaded {len(df)} records from source file")
    logger.info(f"Found {df['CLIENT ID'].nunique()} unique client IDs")
    
    return df


# ==============================================================================
# EMPLOYEE CATEGORIZATION
# ==============================================================================

def categorize_employees(df: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    Categorize employees by client ID, plan type, and coverage tier
    
    Args:
        df: Source DataFrame with employee data
        
    Returns:
        Nested dictionary: {client_id: {plan_type: {category: count}}}
    """
    logger.info("Categorizing employees by coverage tier...")
    
    results = {}
    
    # Group by client ID
    for client_id in df['CLIENT ID'].unique():
        client_data = df[df['CLIENT ID'] == client_id]
        results[client_id] = {
            PlanType.EPO: {cat: 0 for cat in EmployeeCategory.all_categories()},
            PlanType.VALUE: {cat: 0 for cat in EmployeeCategory.all_categories()}
        }
        
        # Group by employee to determine their coverage tier
        employees = {}
        for _, row in client_data.iterrows():
            emp_name = row['EMPLOYEE NAME']
            relation = row['RELATION']
            plan_type = PlanType.EPO if row['EPO-PPO-VAL'] == 'EPO' else PlanType.VALUE
            
            if emp_name not in employees:
                employees[emp_name] = {
                    'plan_type': plan_type,
                    'relations': []
                }
            
            employees[emp_name]['relations'].append(relation)
        
        # Categorize each employee
        for emp_name, emp_data in employees.items():
            plan_type = emp_data['plan_type']
            relations = emp_data['relations']
            
            # Determine coverage category based on dependents
            has_spouse = any(
                'SPOUSE' in r or 'SP' in r 
                for r in relations if r != 'SELF'
            )
            has_children = any(
                'CHILD' in r or 'CH' in r or 'SON' in r or 'DAUGH' in r
                for r in relations if r != 'SELF'
            )
            
            if has_spouse and has_children:
                category = EmployeeCategory.EE_FAMILY
            elif has_spouse:
                category = EmployeeCategory.EE_SPOUSE
            elif has_children:
                category = EmployeeCategory.EE_CHILDREN
            else:
                category = EmployeeCategory.EE
            
            results[client_id][plan_type][category] += 1
        
        # Calculate totals
        for plan_type in [PlanType.EPO, PlanType.VALUE]:
            total = sum(
                results[client_id][plan_type][cat] 
                for cat in [
                    EmployeeCategory.EE,
                    EmployeeCategory.EE_SPOUSE,
                    EmployeeCategory.EE_CHILDREN,
                    EmployeeCategory.EE_FAMILY
                ]
            )
            results[client_id][plan_type][EmployeeCategory.TOTAL] = total
    
    # Log summary
    total_employees = sum(
        sum(
            plan_data[cat]
            for plan_data in client_data.values()
            for cat in [EmployeeCategory.EE, EmployeeCategory.EE_SPOUSE,
                        EmployeeCategory.EE_CHILDREN, EmployeeCategory.EE_FAMILY]
        )
        for client_data in results.values()
    )
    logger.info(f"Categorized {total_employees} unique employees across all clients")
    
    return results


# ==============================================================================
# TARGET FILE PARSING
# ==============================================================================

def parse_target_structure(wb: Workbook, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parse target Excel file to identify update positions
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Name of sheet to parse
        
    Returns:
        List of dictionaries with row mapping information
    """
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    logger.info(f"Parsing target sheet: {ws.title}")
    
    mappings = []
    current_client = None
    current_plan = None
    
    for row_num in range(1, ws.max_row + 1):
        col_a_value = ws.cell(row=row_num, column=1).value  # Coverage/Carrier
        col_c_value = ws.cell(row=row_num, column=3).value  # Category
        
        # Check for client ID in Column A
        if col_a_value and isinstance(col_a_value, str):
            match = re.search(DEFAULT_CONFIG["key_target_pattern"], col_a_value)
            if match:
                current_client = match.group(1)
                logger.debug(f"Found client ID {current_client} at row {row_num}")
                continue
            
            # Check for plan type
            if current_client:
                if 'EPO' in col_a_value.upper():
                    current_plan = PlanType.EPO
                elif 'VALUE' in col_a_value.upper():
                    current_plan = PlanType.VALUE
        
        # Map category rows
        if current_client and current_plan and col_c_value:
            if col_c_value in EmployeeCategory.all_categories():
                mappings.append({
                    'row': row_num,
                    'client_id': current_client,
                    'plan_type': current_plan,
                    'category': col_c_value
                })
    
    logger.info(f"Found {len(mappings)} update positions in target file")
    
    return mappings


# ==============================================================================
# UPDATE FUNCTIONS
# ==============================================================================

def update_target_file(
    wb: Workbook,
    mappings: List[Dict[str, Any]],
    employee_counts: Dict[str, Dict[str, Dict[str, int]]],
    dry_run: bool = False
) -> UpdateResult:
    """
    Update Column D in target file with employee counts
    
    Args:
        wb: Openpyxl workbook object
        mappings: List of row mappings
        employee_counts: Categorized employee counts
        dry_run: If True, don't actually save changes
        
    Returns:
        UpdateResult object with change summary
    """
    ws = wb.active
    result = UpdateResult()
    
    for mapping in mappings:
        row = mapping['row']
        client_id = mapping['client_id']
        plan_type = mapping['plan_type']
        category = mapping['category']
        
        # Get current value
        current_value = ws.cell(row=row, column=4).value  # Column D
        
        # Get new value
        if client_id in employee_counts:
            new_value = employee_counts[client_id][plan_type][category]
        else:
            if client_id not in result.unmatched_clients:
                result.unmatched_clients.append(client_id)
            result.rows_skipped += 1
            continue
        
        # Update cell if not in dry-run mode
        if not dry_run:
            ws.cell(row=row, column=4).value = new_value
        
        # Record change
        result.add_change(
            row=row,
            client_id=client_id,
            plan=plan_type,
            category=category,
            old_value=current_value,
            new_value=new_value
        )
    
    return result


def create_backup(file_path: Path) -> Path:
    """
    Create a backup of the target file
    
    Args:
        file_path: Path to file to backup
        
    Returns:
        Path to backup file
    """
    backup_path = file_path.parent / f"{file_path.stem}.backup{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    logger.info(f"Created backup: {backup_path}")
    return backup_path


# ==============================================================================
# MAIN WORKFLOW
# ==============================================================================

def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(
        description="Update Prime output file with employee counts from source data"
    )
    parser.add_argument(
        "--source",
        type=str,
        default=DEFAULT_CONFIG["source_path"],
        help="Path to source Excel file"
    )
    parser.add_argument(
        "--target",
        type=str,
        default=DEFAULT_CONFIG["target_path"],
        help="Path to target Excel file to update"
    )
    parser.add_argument(
        "--source-sheet",
        type=str,
        default=None,
        help="Source sheet name (default: first sheet)"
    )
    parser.add_argument(
        "--target-sheet",
        type=str,
        default=None,
        help="Target sheet name (default: first sheet)"
    )
    parser.add_argument(
        "--key-source",
        type=str,
        default=DEFAULT_CONFIG["key_source"],
        help="Column name for client ID in source"
    )
    parser.add_argument(
        "--key-target",
        type=str,
        default=DEFAULT_CONFIG["key_target_pattern"],
        help="Pattern to extract client ID from target"
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=DEFAULT_CONFIG["header_row"],
        help="Header row number in target file"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Perform validation and show changes without saving"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging"
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    
    try:
        # Convert paths
        source_path = Path(args.source)
        target_path = Path(args.target)
        
        # Step 1: Load source data
        logger.info("=" * 60)
        logger.info("Step 1: Loading source data")
        logger.info("=" * 60)
        source_df = load_source_data(source_path, args.source_sheet)
        
        # Step 2: Categorize employees
        logger.info("=" * 60)
        logger.info("Step 2: Categorizing employees")
        logger.info("=" * 60)
        employee_counts = categorize_employees(source_df)
        
        # Step 3: Load and parse target file
        logger.info("=" * 60)
        logger.info("Step 3: Parsing target file structure")
        logger.info("=" * 60)
        
        if not target_path.exists():
            raise FileNotFoundError(f"Target file not found: {target_path}")
        
        wb = openpyxl.load_workbook(target_path, data_only=False)
        mappings = parse_target_structure(wb, args.target_sheet)
        
        # Step 4: Update target file
        logger.info("=" * 60)
        if args.dry_run:
            logger.info("Step 4: DRY RUN - Simulating updates")
        else:
            logger.info("Step 4: Updating target file")
        logger.info("=" * 60)
        
        result = update_target_file(wb, mappings, employee_counts, dry_run=args.dry_run)
        
        # Step 5: Save changes (if not dry run)
        if not args.dry_run:
            # Create backup first
            backup_path = create_backup(target_path)
            
            # Save updated file
            wb.save(target_path)
            logger.info(f"Saved updated file: {target_path}")
        
        # Step 6: Report results
        logger.info("=" * 60)
        logger.info("SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Rows updated: {result.rows_updated}")
        logger.info(f"Rows skipped: {result.rows_skipped}")
        
        if result.unmatched_clients:
            logger.warning(f"Unmatched clients in target: {result.unmatched_clients}")
        
        # Show sample changes
        if result.changes:
            logger.info("\nSample changes (first 5):")
            for change in result.changes[:5]:
                logger.info(
                    f"  Row {change['row']}: {change['client_id']} {change['plan']} "
                    f"{change['category']}: {change['old_value']} → {change['new_value']}"
                )
        
        if args.dry_run:
            logger.info("\nDRY RUN COMPLETE - No changes were saved")
            logger.info(f"Run without --dry-run to apply {result.rows_updated} changes")
        else:
            logger.info("\nUPDATE COMPLETE")
            logger.info(f"Backup saved to: {backup_path}")
        
        return 0
        
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=args.verbose)
        return 1


if __name__ == "__main__":
    sys.exit(main())