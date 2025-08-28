""" 
=================================================================================
ENROLLMENT AUTOMATION V6 - DECLARATIVE PLAN AGGREGATIONS
=================================================================================

This version implements:
- Config-driven block aggregations for multi-block tabs
- Hard 29-sheet allowlist enforcement
- Explicit PLAN code → block routing via sum_of lists
- Per-tab children policy (split vs combined)
- UNASSIGNED guard with gap reporting
- Config linting and validation
- Block-level post-write verification

Control Totals (Ground Truth):
- EMP (EE Only): 14,533
- ESP (EE+Spouse): 2,639
- ECH (EE+Child(ren)): 4,413
- FAM (EE+Family): 3,123
TOTAL: 24,708
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import traceback
import os
import re
import json
import argparse
import sys
import csv
from datetime import datetime
from collections import Counter, defaultdict
from difflib import SequenceMatcher
warnings.filterwarnings('ignore')

# ============= CONSTANTS =============

# CONTROL TOTALS - GROUND TRUTH
CONTROL_TOTALS = {
    "EE Only": 14533,
    "EE+Spouse": 2639,
    "EE+Child(ren)": 4413,
    "EE+Family": 3123
}
TIER_ORDER = ["EE Only", "EE+Spouse", "EE+Child(ren)", "EE+Family"]
CONTROL_TOTAL = sum(CONTROL_TOTALS.values())  # 24,708

# HARD ALLOWLIST - Only these 29 sheets will be processed
ALLOWED_TABS = [
    "Centinela", "Coshocton", "Dallas Medical Center", "Dallas Regional",
    "East Liverpool", "Encino-Garden Grove", "Garden City", "Harlingen",
    "Illinois", "Knapp", "Lake Huron", "Landmark", "Legacy", "Lower Bucks",
    "Mission", "Monroe", "North Vista", "Pampa", "Providence & St John",
    "Riverview & Gadsden", "Roxborough", "Saint Clare's", "Saint Mary's Passaic",
    "Saint Mary's Reno", "Southern Regional", "St Joe & St Mary's",
    "St Michael's", "St. Francis", "Suburban"
]

# Excluded CLIENT IDs
EXCLUDED_CIDS = ["H3310"]  # Alvarado - explicitly excluded

# ============= GLOBAL STATE =============
waterfall_stages = []
unknown_tiers_tracker = Counter()
unknown_plans_tracker = Counter()
unassigned_plans = []  # Track plans not in any sum_of list
removed_rows_samples = {}
WRITE_LOG_ROWS = []
PROCESSED_SHEETS = set()
SHEETS_WITH_WRITES = set()  # Only sheets with non-zero writes

# ============= CONFIG MANAGEMENT =============

def load_config(config_path):
    """Load JSON config file"""
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            return json.load(f)
    return {}

def save_config(config_path, data):
    """Save JSON config file"""
    os.makedirs(os.path.dirname(config_path), exist_ok=True)
    with open(config_path, 'w') as f:
        json.dump(data, f, indent=2)

def load_plan_mappings():
    """Load PLAN to plan_group mappings"""
    config = load_config('config/plan_mappings.json')
    return config.get('mappings', {})

def load_block_aggregations():
    """Load block aggregation rules"""
    return load_config('config/block_aggregations.json')

def normalize_tab_name(name):
    """Normalize tab name for matching (handle punctuation/ampersands)"""
    if pd.isna(name):
        return ''
    # Normalize common variations
    normalized = str(name).strip()
    normalized = normalized.replace('&', 'and')
    normalized = normalized.replace("'", '')
    normalized = normalized.replace('.', '')
    normalized = normalized.replace('  ', ' ')
    return normalized

def lint_block_aggregations(block_config, source_plans):
    """
    Validate block aggregation config with enhanced validation
    - No duplicate PLAN codes across blocks for same (cid, plan_type)
    - No empty sum_of lists
    - All referenced PLANs exist in source
    - Special validation for multi-block facilities like St. Michael's
    """
    issues = []
    multi_block_facilities = {}  # Track facilities with multiple blocks
    
    for tab, clients in block_config.items():
        if tab.startswith('_'):  # Skip metadata keys
            continue
        
        for client_id, plan_types in clients.items():
            if client_id.startswith('_'):
                continue
                
            for plan_type, blocks in plan_types.items():
                if plan_type.startswith('_'):
                    continue
                
                # Track multi-block facilities
                if len(blocks) > 1:
                    multi_block_facilities[f"{tab}/{client_id}/{plan_type}"] = len(blocks)
                
                seen_plans = set()
                plan_to_blocks = defaultdict(list)  # Track which blocks use which plans
                
                for block_label, block_def in blocks.items():
                    if 'sum_of' not in block_def:
                        issues.append(f"{tab}/{client_id}/{plan_type}/{block_label}: missing sum_of")
                        continue
                    
                    sum_of = block_def['sum_of']
                    if not sum_of:
                        issues.append(f"{tab}/{client_id}/{plan_type}/{block_label}: empty sum_of")
                    
                    for plan_code in sum_of:
                        plan_to_blocks[plan_code].append(block_label)
                        
                        if plan_code in seen_plans:
                            issues.append(f"{tab}/{client_id}/{plan_type}: CRITICAL - duplicate PLAN '{plan_code}' in blocks: {', '.join(plan_to_blocks[plan_code])}")
                        seen_plans.add(plan_code)
                        
                        if plan_code not in source_plans:
                            issues.append(f"{tab}/{client_id}/{plan_type}/{block_label}: PLAN '{plan_code}' not in source")
    
    # Log multi-block facilities for awareness
    if multi_block_facilities:
        print(f"Multi-block facilities detected: {multi_block_facilities}")
        # Special validation for St. Michael's
        if "St Michael's/H3530/EPO" in multi_block_facilities:
            print("St. Michael's Medical Center has 5 EPO blocks - validating for duplicates...")
    
    return issues

# ============= HELPER FUNCTIONS =============

def clean_key(value):
    """Clean a key value for matching"""
    if pd.isna(value):
        return ''
    return str(value).strip().upper()

def is_active(status):
    """Check if status indicates active enrollment (including COBRA)"""
    if pd.isna(status):
        return False
    s = str(status).strip().upper()
    return s in ['A', 'ACTIVE', 'ACT', 'C', 'COBRA', 'COB']

def is_subscriber(relation):
    """Check if relation indicates subscriber/employee"""
    if pd.isna(relation):
        return False
    r = str(relation).strip().upper()
    return r in ['SELF', 'EE', 'EMPLOYEE', 'SUBSCRIBER', 'SUB', 'EMP', 'S']

def normalize_tier_strict(raw_tier, use_five_tier=False):
    """
    Strictly normalize raw tier text
    For 4-tier: Returns 'EE Only', 'EE+Spouse', 'EE+Child(ren)', 'EE+Family', or 'UNKNOWN'
    For 5-tier: Returns 'EE Only', 'EE+Spouse', 'EE+1 Dep', 'EE+Child', 'EE+Family', or 'UNKNOWN'
    """
    global unknown_tiers_tracker
    
    if pd.isna(raw_tier):
        unknown_tiers_tracker['<NaN>'] += 1
        return 'UNKNOWN'
    
    # Clean the input
    tier_str = str(raw_tier).strip().upper()
    
    # Direct mapping for BEN CODE values
    if tier_str == 'EMP':
        return 'EE Only'
    elif tier_str == 'ESP':
        return 'EE+Spouse'
    elif tier_str == 'FAM':
        return 'EE+Family'
    elif tier_str == 'E1D':
        # E1D handling depends on tier mode
        return 'EE+1 Dep' if use_five_tier else 'EE+Child(ren)'
    elif tier_str == 'ECH':
        # ECH handling depends on tier mode
        return 'EE+Child' if use_five_tier else 'EE+Child(ren)'
    
    # Normalize separators for more complex variants
    for sep in ['&', '+', '/', ' AND ', '  ']:
        tier_str = tier_str.replace(sep, ' ')
    
    # Collapse multiple spaces
    tier_str = ' '.join(tier_str.split())
    
    # Employee Only variants
    ee_only_variants = [
        'EE', 'EMPLOYEE ONLY', 'EE ONLY', 'EMPLOYEE', 'E',
        'SELF ONLY', 'SELF', 'SUBSCRIBER ONLY', 'SUBSCRIBER',
        'EMP ONLY', 'E ONLY'
    ]
    if tier_str in ee_only_variants:
        return 'EE Only'
    
    # Employee + Spouse variants
    spouse_variants = [
        'EE SPOUSE', 'EMPLOYEE SPOUSE', 'ES', 'E S',
        'EMP SPOUSE', 'EMP SP', 'EMPLOYEE SP', 'EE SP'
    ]
    if tier_str in spouse_variants:
        return 'EE+Spouse'
    
    # Employee + Child(ren) variants - handled based on tier mode
    child_variants = [
        'EE CHILD', 'EMPLOYEE CHILD', 'EE CHILDREN',
        'EMPLOYEE CHILDREN', 'EC', 'E C', 'E CHILD', 'E CHILDREN',
        'EMP CHILD', 'EMP CHILDREN', 'CHILD', 'CHILDREN',
        'EE CHILD REN', 'CHILD REN', 'E1C', 'E2C',
        'EE 1 DEP', 'EE DEP', 'EE DEPS', 'EE 1', 'EE 2'
    ]
    if tier_str in child_variants:
        return 'EE+Child(ren)'
    
    # Employee + Family variants
    family_variants = [
        'EE FAMILY', 'EMPLOYEE FAMILY', 'FAMILY', 'E F',
        'EMP FAMILY', 'EMP FAM', 'EF', 'E FAM'
    ]
    if tier_str in family_variants:
        return 'EE+Family'
    
    # Track unknown
    unknown_tiers_tracker[tier_str] += 1
    return 'UNKNOWN'

def create_employee_group(df):
    """
    Create consistent employee grouping with NaN handling
    """
    def safe_group(row):
        client_id = row.get('CLIENT ID', '')
        if pd.isna(client_id):
            client_id = f"UNKNOWN_{row.name}"
        else:
            client_id = str(client_id).strip()
        
        # Try DEP SSN first, then EMPLOYEE #
        dep_ssn = row.get('DEP SSN', '')
        emp_num = row.get('EMPLOYEE #', '')
        
        if not pd.isna(dep_ssn) and str(dep_ssn).strip():
            suffix = str(dep_ssn).strip()
        elif not pd.isna(emp_num) and str(emp_num).strip():
            suffix = str(emp_num).strip()
        else:
            suffix = f"GROUP_{row.name}"
        
        return f"{client_id}_{suffix}"
    
    df['EMPLOYEE_GROUP'] = df.apply(safe_group, axis=1)
    return df

# ============= FACILITY MAPPING =============

# CLIENT ID to Tab mapping
CID_TO_TAB = {
    # Legacy tab (multiple facilities)
    'H3100': 'Legacy', 'H3105': 'Legacy', 'H3110': 'Legacy', 'H3115': 'Legacy',
    'H3120': 'Legacy', 'H3130': 'Legacy', 'H3140': 'Legacy', 'H3150': 'Legacy',
    'H3160': 'Legacy', 'H3170': 'Legacy', 'H3180': 'Legacy', 'H3190': 'Legacy',
    'H3200': 'Legacy', 'H3210': 'Legacy', 'H3230': 'Legacy', 'H3240': 'Legacy',
    'H3280': 'Legacy', 'H3285': 'Legacy', 'H3290': 'Legacy', 'H3300': 'Legacy',
    
    # Encino-Garden Grove tab
    'H3220': 'Encino-Garden Grove', 'H3250': 'Encino-Garden Grove', 'H3260': 'Encino-Garden Grove',
    
    # Centinela tab
    'H3270': 'Centinela', 'H3271': 'Centinela', 'H3272': 'Centinela',
    
    # St. Francis tab
    'H3275': 'St. Francis', 'H3276': 'St. Francis', 'H3277': 'St. Francis',
    
    # Individual facility tabs
    'H3320': 'Pampa',
    'H3325': 'Roxborough',
    'H3330': 'Lower Bucks',
    'H3335': 'Dallas Medical Center',
    'H3337': 'Dallas Regional',
    'H3338': 'Riverview & Gadsden',
    'H3339': 'Riverview & Gadsden',
    'H3340': 'Providence & St John',
    'H3345': 'Providence & St John',
    'H3355': 'Knapp',
    'H3360': 'Knapp',
    'H3370': 'Harlingen',
    'H3375': 'Garden City',
    'H3380': 'Garden City',
    'H3385': 'Garden City',
    'H3381': 'Lake Huron',
    'H3382': 'Lake Huron',
    'H3392': 'Landmark',
    'H3394': "Saint Mary's Reno",
    'H3395': "Saint Mary's Reno",
    'H3396': "Saint Mary's Reno",
    'H3397': 'Monroe',
    'H3398': 'North Vista',
    'H3500': "Saint Clare's",
    'H3505': "Saint Mary's Passaic",
    'H3510': 'Southern Regional',
    'H3530': 'St Michael\'s',
    'H3540': 'Mission',
    'H3591': 'Coshocton',
    'H3592': 'East Liverpool',
    'H3594': 'Ohio Valley HHC',
    'H3595': 'River Valley Pri.',
    'H3596': "St. Mary's Medical",
    'H3598': 'Suburban',
    'H3599': 'Suburban',
    
    # Illinois facilities
    'H3605': 'Illinois', 'H3615': 'Illinois', 'H3625': 'Illinois', 'H3630': 'Illinois',
    'H3635': 'Illinois', 'H3645': 'Illinois', 'H3655': 'Illinois', 'H3660': 'Illinois',
    'H3665': 'Illinois', 'H3670': 'Illinois', 'H3675': 'Illinois', 'H3680': 'Illinois'
}

# ============= LOGGING FUNCTIONS =============

def _log(sheet, client_id, plan_type, block_label, tier_label, cell, value, reason='normal', detection_mode='mapped'):
    """Enhanced logging with block_label and reason"""
    timestamp = datetime.now().isoformat()
    WRITE_LOG_ROWS.append([
        timestamp, sheet, client_id, plan_type, block_label, 
        tier_label, cell, int(value), reason, detection_mode
    ])

def log_stage(stage_name, df, prev_df=None):
    """Log a stage in the waterfall with tier counts"""
    global waterfall_stages, removed_rows_samples
    
    tier_counts = {}
    if 'tier' in df.columns:
        tier_dist = df['tier'].value_counts()
        for tier in TIER_ORDER:
            tier_counts[tier] = tier_dist.get(tier, 0)
        tier_counts['UNKNOWN'] = tier_dist.get('UNKNOWN', 0)
        tier_counts['Other'] = len(df) - sum(tier_counts.values())
    else:
        tier_counts = {tier: 0 for tier in TIER_ORDER}
        tier_counts['UNKNOWN'] = 0
        tier_counts['Other'] = len(df)
    
    delta = sum(tier_counts.get(tier, 0) for tier in TIER_ORDER) - CONTROL_TOTAL
    
    stage_info = {
        'stage': stage_name,
        'total_rows': len(df),
        'tier_counts': tier_counts,
        'delta_to_control': delta
    }
    waterfall_stages.append(stage_info)
    
    return df

# ============= DATA PROCESSING =============

def read_and_prepare_data(file_path, plan_mappings):
    """Read source data and prepare with all transformations"""
    
    # Stage 1: Read raw
    df = pd.read_excel(file_path, sheet_name=0)
    df['original_index'] = df.index
    df = log_stage('read_raw', df)
    
    # Stage 2: Clean keys
    if 'CLIENT ID' in df.columns:
        df['CLIENT ID'] = df['CLIENT ID'].apply(clean_key)
    if 'BEN CODE' in df.columns:
        df['BEN CODE'] = df['BEN CODE'].apply(clean_key)
    if 'PLAN' in df.columns:
        df['PLAN_CLEAN'] = df['PLAN'].apply(clean_key)
    df = log_stage('clean_keys', df)
    
    # Stage 3: Exclude specific CIDs
    prev_df = df.copy()
    df = df[~df['CLIENT ID'].isin(EXCLUDED_CIDS)].copy()
    if len(prev_df) > len(df):
        print(f"Excluded {len(prev_df) - len(df)} rows for CIDs: {EXCLUDED_CIDS}")
    df = log_stage('exclude_cids', df, prev_df)
    
    # Stage 4: Status filter (including COBRA)
    prev_df = df.copy()
    if 'STATUS' in df.columns:
        df['is_active'] = df['STATUS'].apply(is_active)
        df = df[df['is_active']].copy()
    df = log_stage('status_filter', df, prev_df)
    
    # Stage 5: Relation filter (subscribers only)
    prev_df = df.copy()
    if 'RELATION' in df.columns:
        df['is_subscriber'] = df['RELATION'].apply(is_subscriber)
        df = df[df['is_subscriber']].copy()
    df = log_stage('relation_filter', df, prev_df)
    
    # Stage 6: Tab mapping
    if 'CLIENT ID' in df.columns:
        df['tab_name'] = df['CLIENT ID'].map(CID_TO_TAB)
        df['tab_name'] = df['tab_name'].fillna('UNKNOWN')
        df['tab_normalized'] = df['tab_name'].apply(normalize_tab_name)
    df = log_stage('tab_map', df)
    
    # Stage 7: Tier normalization with 5-tier fix
    # Use CALCULATED BEN CODE for 5-tier tabs, BEN CODE for others
    # CRITICAL FIX: Prevent double-counting in 5-tier tabs
    if 'CLIENT ID' in df.columns:
        # Define 5-tier tabs
        FIVE_TIER_TABS = ['Encino-Garden Grove', 'North Vista']
        
        # Add processing flag to prevent duplicate counting
        df['processed'] = False
        
        def get_appropriate_ben_code(row):
            """Choose between BEN CODE and CALCULATED BEN CODE based on tab
            FIX: Use ONLY CALCULATED BEN CODE for 5-tier tabs to prevent double-counting"""
            tab_name = row.get('tab_name', '')
            if tab_name in FIVE_TIER_TABS:
                # For 5-tier tabs, ONLY use CALCULATED BEN CODE if present
                if 'CALCULATED BEN CODE' in row and pd.notna(row.get('CALCULATED BEN CODE')):
                    return row['CALCULATED BEN CODE']
                # Fallback to BEN CODE only if CALCULATED is missing
                elif 'BEN CODE' in row and pd.notna(row.get('BEN CODE')):
                    return row['BEN CODE']
            else:
                # For 4-tier tabs, use BEN CODE
                if 'BEN CODE' in row and pd.notna(row.get('BEN CODE')):
                    return row['BEN CODE']
            return None
        
        # Create unified ben_code column (from fix_5tier_enrollment.py)
        df['unified_ben_code'] = df.apply(get_appropriate_ben_code, axis=1)
        
        def normalize_with_context(row):
            """Normalize tier based on tab context"""
            ben_code = row.get('unified_ben_code')
            tab_name = row.get('tab_name', '')
            use_five_tier = tab_name in FIVE_TIER_TABS
            return normalize_tier_strict(ben_code, use_five_tier)
        
        df['tier'] = df.apply(normalize_with_context, axis=1)
    elif 'BEN CODE' in df.columns:
        df['tier'] = df['BEN CODE'].apply(normalize_tier_strict)
    else:
        df['tier'] = 'UNKNOWN'
    df = log_stage('tier_normalized', df)
    
    # Stage 8: Plan group mapping
    if 'PLAN_CLEAN' in df.columns:
        df['plan_group'] = df['PLAN_CLEAN'].map(plan_mappings)
        df['plan_group'] = df['plan_group'].fillna('UNKNOWN')
    else:
        df['plan_group'] = 'UNKNOWN'
    df = log_stage('plan_grouped', df)
    
    # Stage 9: Create employee groups
    df = create_employee_group(df)
    df = log_stage('employee_grouped', df)
    
    # Stage 10: Enhanced Deduplication with duplicate tracking
    # Add unique identifier to prevent duplicate aggregation
    if 'EE ID' in df.columns:
        df['enrollment_id'] = df['CLIENT ID'].astype(str) + '_' + df['EE ID'].astype(str) + '_' + df['PLAN'].astype(str)
    else:
        df['enrollment_id'] = df['CLIENT ID'].astype(str) + '_' + df['EMPLOYEE_GROUP'].astype(str) + '_' + df['PLAN'].astype(str)
    
    # Track duplicates before removal
    duplicate_mask = df.duplicated(subset=['CLIENT ID', 'EMPLOYEE_GROUP', 'PLAN'], keep=False)
    duplicate_count = duplicate_mask.sum()
    if duplicate_count > 0:
        print(f"Warning: Found {duplicate_count} duplicate enrollments - removing duplicates")
    
    df_deduped = df.drop_duplicates(
        subset=['CLIENT ID', 'EMPLOYEE_GROUP', 'PLAN'],
        keep='first'
    ).copy()
    df = log_stage('deduplicated', df_deduped, df)
    
    return df

def build_tier_data_from_source(df, block_aggregations, allow_ppo=False):
    """
    Build tier data with block-level aggregation based on config
    Supports both 4-tier and 5-tier structures
    """
    global unassigned_plans
    
    # Define 5-tier tabs
    FIVE_TIER_TABS = ['Encino-Garden Grove', 'North Vista']
    
    tier_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        'EE Only': 0, 'EE+Spouse': 0, 'EE+1 Dep': 0, 'EE+Child': 0, 'EE+Child(ren)': 0, 'EE+Children': 0, 'EE+Family': 0
    })))
    
    unassigned_plans = []
    block_diagnostics = defaultdict(lambda: defaultdict(Counter))
    
    for _, row in df.iterrows():
        client_id = row['CLIENT ID']
        plan_group = row['plan_group']
        plan_code = row.get('PLAN_CLEAN', '')
        tier = row['tier']
        tab_name = row['tab_name']
        
        # Skip non-EPO/VALUE unless allow_ppo
        if plan_group not in ['EPO', 'VALUE']:
            if plan_group == 'PPO' and not allow_ppo:
                print(f"WARNING: PPO plan found for {client_id}: {plan_code} (use --allow-ppo to include)")
            continue
        
        # Skip UNKNOWN tiers
        if tier == 'UNKNOWN':
            continue
        
        # Find block label for this row
        block_label = None
        
        # Check if this tab/client has block aggregation rules
        if tab_name in block_aggregations and client_id in block_aggregations[tab_name]:
            client_config = block_aggregations[tab_name][client_id]
            
            if plan_group in client_config:
                plan_config = client_config[plan_group]
                
                # Find which block this PLAN code belongs to
                for block_name, block_def in plan_config.items():
                    if block_name.startswith('_'):  # Skip metadata
                        continue
                    
                    if 'sum_of' in block_def and plan_code in block_def['sum_of']:
                        block_label = block_name
                        block_diagnostics[(tab_name, client_id, plan_group)][block_name][plan_code] += 1
                        break
                
                if not block_label:
                    # PLAN not in any sum_of list for this client
                    unassigned_plans.append((tab_name, client_id, plan_group, plan_code))
                    block_label = 'UNASSIGNED'
        else:
            # No block config for this tab/client - use default single block
            block_label = f"{plan_group} (Default)"
        
        # Aggregate tier counts
        # Store directly - we'll handle 4-tier vs 5-tier mapping later
        tier_data[client_id][plan_group][block_label][tier] += 1
    
    # Print block diagnostics
    print("\n" + "="*80)
    print("BLOCK AGGREGATION DIAGNOSTICS")
    print("="*80)
    for (tab_name, client_id, plan_group), blocks in sorted(block_diagnostics.items()):
        print(f"\n{tab_name} / {client_id} / {plan_group}:")
        for block_label, plan_counts in blocks.items():
            total = sum(plan_counts.values())
            plans_list = ', '.join([f"{p}({c})" for p, c in plan_counts.most_common()])
            print(f"  {block_label}: {total} total → {plans_list}")
    
    # Convert to regular dict
    return {k: {p: {b: dict(v) for b, v in blocks.items()} 
                for p, blocks in plans.items()} 
            for k, plans in tier_data.items()}

def check_unassigned_plans(allow_unassigned=False):
    """Check for unassigned plans and fail if any exist"""
    global unassigned_plans
    
    if not unassigned_plans:
        return True
    
    # Count unassigned by unique combinations
    unassigned_counter = Counter(unassigned_plans)
    
    print("\n" + "="*80)
    print("❌ UNASSIGNED PLANS DETECTED - Mapping Gap Table")
    print("="*80)
    print(f"{'Tab':<25} {'Client':<10} {'Type':<8} {'PLAN Code':<30} {'Count':<10}")
    print("-"*80)
    
    for (tab, client_id, plan_type, plan_code), count in unassigned_counter.most_common():
        print(f"{tab:<25} {client_id:<10} {plan_type:<8} {plan_code:<30} {count:<10}")
    
    if not allow_unassigned:
        print("\n❌ Cannot proceed with unassigned plans. Add to block_aggregations.json")
        return False
    
    return True

def assert_control_from_tier_data(tier_data):
    """Assert global control totals (excluding UNKNOWN)"""
    totals = Counter({'EE Only': 0, 'EE+Spouse': 0, 'EE+Child(ren)': 0, 'EE+Family': 0})
    
    for client_plans in tier_data.values():
        for plan_blocks in client_plans.values():
            for block_counts in plan_blocks.values():
                totals['EE Only'] += int(block_counts.get('EE Only', 0))
                totals['EE+Spouse'] += int(block_counts.get('EE+Spouse', 0))
                # For 5-tier tabs: EE+1 Dep and EE+Child are separate
                # For 4-tier tabs: EE+Child(ren) combines both
                child_total = (int(block_counts.get('EE+Child', 0)) + 
                              int(block_counts.get('EE+Children', 0)) + 
                              int(block_counts.get('EE+1 Dep', 0)) +
                              int(block_counts.get('EE+Child(ren)', 0)))
                totals['EE+Child(ren)'] += child_total
                totals['EE+Family'] += int(block_counts.get('EE+Family', 0))
    
    deltas = {k: totals[k] - CONTROL_TOTALS[k] for k in CONTROL_TOTALS}
    ok = all(v == 0 for v in deltas.values())
    
    print("\n" + "="*80)
    print("CONTROL CHECK (from tier_data)")
    print(f"Totals: {dict(totals)}")
    print(f"Deltas: {deltas}")
    print(f"Status: {'✅ PASS' if ok else '❌ FAIL'}")
    print("="*80)
    
    return ok

def print_diagnostics(df, plan_mappings):
    """Print diagnostic tables"""
    
    # BEN CODE mapping table
    print("\n" + "="*80)
    print("BEN CODE → TIER MAPPING TABLE")
    print("="*80)
    if 'BEN CODE' in df.columns:
        ben_code_counts = df['BEN CODE'].value_counts()
        for ben_code, count in ben_code_counts.head(20).items():
            normalized = normalize_tier_strict(ben_code)
            print(f"{ben_code:20s} → {normalized:20s} ({count:,} rows)")
    
    # Unknown tracking
    if unknown_tiers_tracker:
        print("\n" + "="*80)
        print("UNKNOWN TIERS")
        print("="*80)
        for tier, count in unknown_tiers_tracker.most_common(10):
            print(f"  {tier}: {count} occurrences")

# ============= WRITE FUNCTIONS =============

def get_children_policy(tab_name, client_id, block_aggregations):
    """Get children policy for a specific tab/client"""
    default_policy = block_aggregations.get('_default_children_policy', 'combined')
    
    if tab_name in block_aggregations:
        tab_config = block_aggregations[tab_name]
        
        # Check client-specific policy
        if client_id in tab_config:
            client_policy = tab_config[client_id].get('_children_policy')
            if client_policy:
                return client_policy
        
        # Check tab-level policy
        tab_policy = tab_config.get('_children_policy')
        if tab_policy:
            return tab_policy
    
    return default_policy

def get_actual_sheet_name(wb, desired_name):
    """
    Find the actual sheet name in the workbook that matches the desired name
    Uses fuzzy matching to handle variations in punctuation, spacing, etc.
    """
    # Direct match first
    if desired_name in wb.sheetnames:
        return desired_name
    
    # Known mappings for common variations (now aligned with ALLOWED_TABS)
    sheet_mappings = {
        # These are already in correct format in ALLOWED_TABS
        # Keeping for backward compatibility if needed
    }
    
    if desired_name in sheet_mappings:
        actual = sheet_mappings[desired_name]
        if actual in wb.sheetnames:
            return actual
    
    # Try normalized matching
    normalized_desired = normalize_tab_name(desired_name)
    for sheet in wb.sheetnames:
        if normalize_tab_name(sheet) == normalized_desired:
            return sheet
    
    return None

def write_to_specific_sheet(wb, sheet_name, write_map, tier_data, block_aggregations):
    """
    Write tier counts to specific sheet with block-level matching
    """
    
    # Check allowlist
    normalized_sheet = normalize_tab_name(sheet_name)
    if normalized_sheet not in [normalize_tab_name(t) for t in ALLOWED_TABS]:
        print(f"⚠️ Skipped '{sheet_name}' - not in allowlist")
        return []
    
    # Find the actual sheet name in workbook
    actual_sheet_name = get_actual_sheet_name(wb, sheet_name)
    if not actual_sheet_name:
        print(f"❌ ERROR: '{sheet_name}' sheet not found in workbook!")
        return []
    
    ws = wb[actual_sheet_name]
    PROCESSED_SHEETS.add(sheet_name)
    write_log = []
    seen_blocks = set()  # Track (client_id, plan_type, block_label) for dedupe
    has_non_zero_write = False
    
    print(f"\nWriting to {sheet_name} sheet...")
    
    for entry in write_map:
        client_id = entry['client_id']
        plan_type = entry['plan']
        cells = entry['cells']
        label = entry.get('label', '')
        block_label = label  # Use label as block identifier
        
        # Multi-block dedupe with block_label
        key = (client_id, plan_type, block_label)
        if key in seen_blocks:
            # Zero duplicate blocks
            for cell in cells.values():
                ws[cell] = 0
                _log(sheet_name, client_id, plan_type, block_label, 'DUPLICATE-ZERO', 
                     cell, 0, 'duplicate', 'mapped')
            print(f"  ⚠️ Skipped duplicate block: {key}")
            continue
        seen_blocks.add(key)
        
        # Define 5-tier tabs
        FIVE_TIER_TABS = ['Encino-Garden Grove', 'North Vista']
        is_five_tier = sheet_name in FIVE_TIER_TABS
        
        # Get tier counts for this block
        tier_counts = {'EE Only': 0, 'EE+Spouse': 0, 'EE+1 Dep': 0, 'EE+Child': 0, 
                      'EE+Children': 0, 'EE+Child(ren)': 0, 'EE+Family': 0}
        
        # Find matching block in tier_data
        if client_id in tier_data and plan_type in tier_data[client_id]:
            blocks = tier_data[client_id][plan_type]
            
            # Try exact match first
            if block_label in blocks:
                for tier, count in blocks[block_label].items():
                    tier_counts[tier] = int(count)
            else:
                # If only one block exists for this client/plan, use it regardless of label
                if len(blocks) == 1:
                    block_key = list(blocks.keys())[0]
                    for tier, count in blocks[block_key].items():
                        tier_counts[tier] = int(count)
                else:
                    # Try to find a matching block by partial match
                    for block_key, block_counts in blocks.items():
                        if block_label and block_key and (block_label in block_key or block_key in block_label):
                            for tier, count in block_counts.items():
                                tier_counts[tier] = int(count)
                            break
        
        # Apply children policy based on tier structure
        children_policy = get_children_policy(sheet_name, client_id, block_aggregations)
        
        if is_five_tier:
            # For 5-tier tabs (Encino-Garden Grove, North Vista)
            # E1D stays separate as EE+1 Dep, ECH stays as EE+Child
            # Map E1D to the EE+Children cell row in Excel
            pass  # Keep tiers as-is for 5-tier structure
        else:
            # For 4-tier tabs, combine child tiers
            if children_policy == 'combined':
                combined_children = (tier_counts.get('EE+Child', 0) + 
                                   tier_counts.get('EE+Children', 0) + 
                                   tier_counts.get('EE+Child(ren)', 0) +
                                   tier_counts.get('EE+1 Dep', 0))  # Include E1D in combined
                tier_counts['EE+Children'] = combined_children
                tier_counts['EE+Child'] = 0
                tier_counts['EE+Child(ren)'] = 0
                tier_counts['EE+1 Dep'] = 0
        
        # Zero-fill all cells first
        for cell in cells.values():
            ws[cell] = 0
        
        # Write values
        written_total = 0
        for tier_label, cell in cells.items():
            # Map label variants based on tier structure
            if is_five_tier:
                # 5-tier mapping for Encino-Garden Grove and North Vista
                if 'EE' in tier_label and 'Spouse' in tier_label:
                    value = tier_counts.get('EE+Spouse', 0)
                elif 'EE' in tier_label and 'Children' in tier_label:
                    # E1D maps to "EE & Children" row
                    value = tier_counts.get('EE+1 Dep', 0)
                elif 'EE' in tier_label and 'Child' in tier_label and 'Children' not in tier_label:
                    # ECH maps to "EE & Child" row
                    value = tier_counts.get('EE+Child', 0)
                elif 'EE' in tier_label and 'Family' in tier_label:
                    value = tier_counts.get('EE+Family', 0)
                elif tier_label in ['EE', 'EEs', 'Employees', 'EE Only']:
                    value = tier_counts.get('EE Only', 0)
                else:
                    value = tier_counts.get(tier_label, 0)
            else:
                # 4-tier mapping for all other tabs
                if 'EE' in tier_label and 'Spouse' in tier_label:
                    value = tier_counts.get('EE+Spouse', 0)
                elif 'EE' in tier_label and ('Child' in tier_label or 'Children' in tier_label):
                    if children_policy == 'split' and 'Children' in tier_label:
                        value = tier_counts.get('EE+Children', 0)
                    elif children_policy == 'split' and 'Child' in tier_label and 'Children' not in tier_label:
                        value = tier_counts.get('EE+Child', 0)
                    else:
                        # Combined child tiers
                        value = tier_counts.get('EE+Children', 0) + tier_counts.get('EE+Child(ren)', 0)
                elif 'EE' in tier_label and 'Family' in tier_label:
                    value = tier_counts.get('EE+Family', 0)
                elif tier_label in ['EE', 'EEs', 'Employees', 'EE Only']:
                    value = tier_counts.get('EE Only', 0)
                else:
                    value = tier_counts.get(tier_label, 0)
            
            ws[cell] = int(value)
            written_total += int(value)
            if value > 0:
                has_non_zero_write = True
            
            _log(sheet_name, client_id, plan_type, block_label, tier_label, 
                 cell, value, 'normal', 'mapped')
            write_log.append((sheet_name, client_id, plan_type, block_label, cell, value))
        
        if written_total > 0:
            print(f"  ✓ {client_id} {plan_type} ({block_label}): {written_total} total")
    
    if has_non_zero_write:
        SHEETS_WITH_WRITES.add(sheet_name)
    
    return write_log

# Import write maps from separate file
from write_maps import SHEET_WRITE_MAPS

def perform_comprehensive_writeback(workbook_path, tier_data, block_aggregations, output_path=None, dry_run=False):
    """Perform comprehensive write-back to all configured sheets"""
    
    global WRITE_LOG_ROWS, PROCESSED_SHEETS, SHEETS_WITH_WRITES
    WRITE_LOG_ROWS = []
    PROCESSED_SHEETS = set()
    SHEETS_WITH_WRITES = set()
    
    if not output_path:
        base_name = os.path.splitext(workbook_path)[0]
        output_path = f"{base_name}_updated.xlsx"
    
    print("\n" + "="*80)
    print("COMPREHENSIVE ENROLLMENT WRITE-BACK")
    print("="*80)
    
    # Pre-write control assertion
    if not assert_control_from_tier_data(tier_data):
        print("❌ Control totals mismatch - aborting write-back")
        return None
    
    # Load workbook
    print(f"\nOpening workbook: {workbook_path}")
    try:
        wb = load_workbook(workbook_path)
    except FileNotFoundError:
        print(f"❌ ERROR: Workbook not found: {workbook_path}")
        return None
    except PermissionError:
        print("❌ ERROR: Workbook is open in Excel. Please close and retry.")
        return None
    
    all_write_logs = []
    
    # Process each sheet with its write map
    for sheet_name, write_map in SHEET_WRITE_MAPS.items():
        if sheet_name in ALLOWED_TABS or normalize_tab_name(sheet_name) in [normalize_tab_name(t) for t in ALLOWED_TABS]:
            logs = write_to_specific_sheet(wb, sheet_name, write_map, tier_data, block_aggregations)
            all_write_logs.extend(logs)
    
    # Save workbook
    if not dry_run:
        print(f"\nSaving to: {output_path}")
        wb.save(output_path)
    else:
        print("\n[DRY RUN - not saved]")
    
    # Save write log CSV
    os.makedirs('output', exist_ok=True)
    write_log_path = 'output/write_log.csv'
    
    with open(write_log_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['timestamp', 'sheet', 'client_id', 'plan_type', 'block_label',
                        'tier_label', 'cell', 'value', 'reason', 'detection_mode'])
        writer.writerows(WRITE_LOG_ROWS)
    
    # Summary
    print("\n" + "="*80)
    print("WRITE SUMMARY")
    print("="*80)
    print(f"Total writes: {len(WRITE_LOG_ROWS)}")
    print(f"Sheets processed: {len(PROCESSED_SHEETS)}")
    print(f"Sheets with writes: {len(SHEETS_WITH_WRITES)} → {', '.join(sorted(SHEETS_WITH_WRITES))}")
    print(f"Sheets skipped: {len(ALLOWED_TABS) - len(PROCESSED_SHEETS)}")
    print(f"Write log: {write_log_path}")
    print(f"Output file: {output_path if not dry_run else '[DRY RUN]'}")
    
    # Post-write verification
    verify_writes_per_block(tier_data)
    
    return output_path

def verify_writes_per_block(tier_data):
    """Post-write verification at block level"""
    
    print("\n" + "="*80)
    print("POST-WRITE VERIFICATION (Per Block)")
    print("="*80)
    
    by_block = Counter()
    
    # Sum written values by (client_id, plan_type, block_label)
    for row in WRITE_LOG_ROWS:
        _, sheet, client_id, plan_type, block_label, tier_label, cell, value, reason, _ = row
        # Skip duplicate entries
        if reason != 'duplicate':
            by_block[(client_id, plan_type, block_label)] += int(value)
    
    # Compare to source
    mismatches = []
    for client_id, plan_types in tier_data.items():
        for plan_type, blocks in plan_types.items():
            for block_label, counts in blocks.items():
                src_sum = sum(int(v) for v in counts.values())
                sheet_sum = by_block.get((client_id, plan_type, block_label), 0)
                
                if src_sum != sheet_sum and plan_type in ['EPO', 'VALUE']:
                    mismatches.append((client_id, plan_type, block_label, src_sum, sheet_sum))
    
    if mismatches:
        print("⚠️ Mismatches found:")
        for client_id, plan_type, block_label, src, sheet in mismatches[:20]:
            print(f"  {client_id} {plan_type} ({block_label}): source={src} vs sheet={sheet}")
    else:
        print("✅ All block totals match!")

# ============= MAIN FUNCTION =============

def main(args):
    """Main execution function"""
    
    # Reset global state
    global waterfall_stages, unknown_tiers_tracker, unknown_plans_tracker
    global unassigned_plans, removed_rows_samples, WRITE_LOG_ROWS
    global PROCESSED_SHEETS, SHEETS_WITH_WRITES
    
    waterfall_stages = []
    unknown_tiers_tracker = Counter()
    unknown_plans_tracker = Counter()
    unassigned_plans = []
    removed_rows_samples = {}
    WRITE_LOG_ROWS = []
    PROCESSED_SHEETS = set()
    SHEETS_WITH_WRITES = set()
    
    print("="*80)
    print("ENROLLMENT AUTOMATION V6 - DECLARATIVE PLAN AGGREGATIONS")
    print("="*80)
    
    # Load configurations
    plan_mappings = load_plan_mappings()
    block_aggregations = load_block_aggregations()
    
    print(f"\nLoaded {len(plan_mappings)} plan mappings")
    print(f"Loaded block aggregations for {len([k for k in block_aggregations if not k.startswith('_')])} tabs")
    
    # Process data
    print("\nProcessing source data...")
    df = read_and_prepare_data(args.source, plan_mappings)
    
    # Print diagnostics
    print_diagnostics(df, plan_mappings)
    
    # Lint block aggregations
    source_plans = set(df['PLAN_CLEAN'].unique()) if 'PLAN_CLEAN' in df.columns else set()
    lint_issues = lint_block_aggregations(block_aggregations, source_plans)
    
    if lint_issues and args.strict_control:
        print("\n❌ BLOCK AGGREGATION CONFIG ISSUES:")
        for issue in lint_issues[:20]:
            print(f"  - {issue}")
        if not args.allow_config_issues:
            print("\nCannot proceed with config issues. Fix block_aggregations.json")
            return 1
    
    # Build tier data with block aggregation
    print("\nBuilding tier aggregations with block routing...")
    tier_data = build_tier_data_from_source(df, block_aggregations, args.allow_ppo)
    
    # Check for unassigned plans
    if not check_unassigned_plans(args.allow_unassigned):
        return 1
    
    # Perform write-back
    output_path = perform_comprehensive_writeback(
        args.workbook, 
        tier_data,
        block_aggregations,
        args.output,
        args.dry_run
    )
    
    if output_path:
        print(f"\n✅ SUCCESS: Enrollment automation complete!")
        return 0
    else:
        print(f"\n❌ FAILED: Check errors above")
        return 1

# ============= CLI SETUP =============

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Enrollment Automation V6 - Declarative Aggregations')
    parser.add_argument('--source', default='C:\\Users\\becas\\Prime_EFR\\data\\input\\source_data.xlsx',
                       help='Source data file path')
    parser.add_argument('--workbook', 
                       default='C:\\Users\\becas\\Prime_EFR\\Prime Enrollment Funding by Facility for August.xlsx',
                       help='Template workbook path')
    parser.add_argument('--output', help='Output file path (optional)')
    parser.add_argument('--strict-control', action='store_true', default=True,
                       help='Enforce strict control totals')
    parser.add_argument('--no-strict-control', dest='strict_control', action='store_false',
                       help='Disable strict control totals')
    parser.add_argument('--dry-run', action='store_true',
                       help='Run without saving Excel file')
    parser.add_argument('--allow-ppo', action='store_true',
                       help='Allow PPO plans (default: fail on PPO)')
    parser.add_argument('--allow-unassigned', action='store_true',
                       help='Allow unassigned PLAN codes')
    parser.add_argument('--allow-config-issues', action='store_true',
                       help='Allow config validation issues')
    
    args = parser.parse_args()
    
    sys.exit(main(args))